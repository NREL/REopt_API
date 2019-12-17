from django.db import models
import uuid
import os
import datetime, tzlocal
from openpyxl import load_workbook
from reo.models import ScenarioModel, SiteModel, PVModel, WindModel, GeneratorModel, StorageModel, FinancialModel, ElectricTariffModel, LoadProfileModel
from reo.src.data_manager import big_number
from reo.log_levels import log

class ProForma(models.Model):

    scenariomodel = models.OneToOneField(
        ScenarioModel,
        on_delete=models.CASCADE,
        default=0,
        to_field='id',
        blank=True,
        primary_key=True
    )
    uuid = models.UUIDField(default=uuid.uuid4, null=False)
    spreadsheet_created = models.DateTimeField(null=True)
    created = models.DateTimeField(auto_now_add=True)
    updated = models.DateTimeField(auto_now=True)
    
    @classmethod
    def create(cls, scenariomodel, **kwargs ):
        pf = cls(scenariomodel = scenariomodel, **kwargs)

        file_dir = os.path.dirname(pf.output_file)
        
        if not os.path.exists(file_dir):
            os.mkdir(file_dir)
        pf.save()        
        return pf
    
    @property
    def sheet_io(self):
        return "Inputs and Outputs"

    @property
    def file_template(self):
        return os.path.join('proforma',"REoptCashFlowTemplate.xlsm")

    @property
    def output_file_name(self):
        return "ProForma.xlsm"

    @property
    def output_file(self):
        folder = os.path.join(os.getcwd(),'static', 'files', str(self.uuid))
        if not os.path.exists(folder):
            os.mkdir(folder)
        return os.path.join(folder, self.output_file_name)
          
    def generate_spreadsheet(self):

        log.info("Generating proforma spreadsheet")

        scenario = self.scenariomodel
        batt = StorageModel.objects.filter(run_uuid=scenario.run_uuid).first()
        pv = PVModel.objects.filter(run_uuid=scenario.run_uuid).first()
        wind = WindModel.objects.filter(run_uuid=scenario.run_uuid).first()
        generator = GeneratorModel.objects.filter(run_uuid=scenario.run_uuid).first()
        electric_tariff = ElectricTariffModel.objects.filter(run_uuid=scenario.run_uuid).first()
        financial = FinancialModel.objects.filter(run_uuid=scenario.run_uuid).first()
        loadprofile = LoadProfileModel.objects.filter(run_uuid=scenario.run_uuid).first()

        # Open file for reading
        wb = load_workbook(self.file_template, read_only=False, keep_vba=True)

        # Open Inputs Sheet
        ws = wb.get_sheet_by_name(self.sheet_io)

        # handle None's
        pv_installed_kw = pv.size_kw or 0
        if pv_installed_kw > 0 and pv.existing_kw > 0:
            pv_installed_kw -= pv.existing_kw
        batt_size_kw = batt.size_kw or 0
        batt_size_kwh = batt.size_kwh or 0
        wind_installed_kw = wind.size_kw or 0
        generator_installed_kw = generator.size_kw or 0
        if generator_installed_kw > 0 and generator.existing_kw > 0:
            generator_installed_kw -= generator.existing_kw

        pv_installed_cost_us_dollars_per_kw = pv.installed_cost_us_dollars_per_kw or 0
        batt_installed_cost_us_dollars_per_kw = batt.installed_cost_us_dollars_per_kw or 0
        batt_installed_cost_us_dollars_per_kwh = batt.installed_cost_us_dollars_per_kwh or 0
        wind_installed_cost_us_dollars_per_kw = wind.installed_cost_us_dollars_per_kw or 0
        generator_installed_cost_us_dollars_per_kw = generator.installed_cost_us_dollars_per_kw or 0
        pv_energy = pv.year_one_energy_produced_kwh or 0
        wind_energy = wind.year_one_energy_produced_kwh or 0
        generator_energy = generator.year_one_energy_produced_kwh or 0

        pv_cost = pv_installed_kw * pv_installed_cost_us_dollars_per_kw
        batt_cost = batt_size_kw * batt_installed_cost_us_dollars_per_kw \
                    + batt_size_kwh * batt_installed_cost_us_dollars_per_kwh
        wind_cost = wind_installed_kw * wind_installed_cost_us_dollars_per_kw
        generator_cost = generator_installed_kw * generator_installed_cost_us_dollars_per_kw
        diesel_fuel_used_cost = generator.diesel_fuel_cost_us_dollars_per_gallon * generator.fuel_used_gal

        # System Design
        ws['B3'] = pv_installed_kw
        ws['B4'] = pv.existing_kw or 0
        ws['B5'] = pv.degradation_pct * 100
        ws['B6'] = wind_installed_kw
        ws['B7'] = generator_installed_kw
        ws['B8'] = generator.existing_kw or 0
        ws['B9'] = batt.size_kw or 0
        ws['B10'] = batt.size_kwh or 0

        # Year 1 Results
        ws['B13'] = electric_tariff.year_one_bill_bau_us_dollars or 0
        ws['B14'] = electric_tariff.year_one_bill_us_dollars or 0
        ws['B15'] = electric_tariff.year_one_export_benefit_us_dollars or 0
        ws['B16'] = pv_energy
        ws['B17'] = wind_energy
        ws['B18'] = generator_energy
        ws['B19'] = wind_energy + pv_energy + generator_energy

        # System Costs
        ws['B22'] = pv_cost + batt_cost + wind_cost + generator_cost
        ws['B23'] = pv_cost
        ws['B24'] = wind_cost
        ws['B25'] = generator_cost
        ws['B26'] = batt_cost
        ws['B28'] = pv.om_cost_us_dollars_per_kw
        ws['B29'] = wind.om_cost_us_dollars_per_kw
        ws['B30'] = generator.om_cost_us_dollars_per_kw
        ws['B31'] = generator.om_cost_us_dollars_per_kwh
        ws['B32'] = diesel_fuel_used_cost
        ws['B33'] = batt.replace_cost_us_dollars_per_kw
        ws['B34'] = batt.inverter_replacement_year
        ws['B35'] = batt.replace_cost_us_dollars_per_kwh
        ws['B36'] = batt.battery_replacement_year

        # Analysis Parameters
        ws['B44'] = financial.analysis_years
        ws['B45'] = financial.om_cost_escalation_pct * 100
        ws['B46'] = financial.escalation_pct * 100
        ws['B47'] = financial.offtaker_discount_pct * 100

        # Tax rates
        ws['B50'] = financial.offtaker_tax_pct * 100

        # PV Tax Credits and Incentives
        ws['B55'] = pv.federal_itc_pct * 100
        ws['C55'] = ''
        ws['B60'] = pv.state_ibi_pct * 100
        ws['C60'] = pv.state_ibi_max_us_dollars
        ws['B61'] = pv.utility_ibi_pct * 100
        ws['C61'] = pv.utility_ibi_max_us_dollars
        ws['B63'] = pv.federal_rebate_us_dollars_per_kw * 0.001
        ws['C63'] = ''
        ws['B64'] = pv.state_rebate_us_dollars_per_kw * 0.001
        ws['C64'] = pv.state_rebate_max_us_dollars
        ws['B65'] = pv.utility_rebate_us_dollars_per_kw * 0.001
        ws['C65'] = pv.utility_rebate_max_us_dollars
        ws['B67'] = pv.pbi_us_dollars_per_kwh
        ws['C67'] = pv.pbi_max_us_dollars
        ws['E67'] = pv.pbi_years
        ws['F67'] = pv.pbi_system_max_kw

        # Wind Tax Credits and Incentives
        ws['B72'] = wind.federal_itc_pct * 100
        ws['C72'] = ''
        ws['B77'] = wind.state_ibi_pct * 100
        ws['C77'] = wind.state_ibi_max_us_dollars
        ws['B78'] = wind.utility_ibi_pct * 100
        ws['C78'] = wind.utility_ibi_max_us_dollars
        ws['B80'] = wind.federal_rebate_us_dollars_per_kw * 0.001
        ws['C80'] = ''
        ws['B81'] = wind.state_rebate_us_dollars_per_kw * 0.001
        ws['C81'] = wind.state_rebate_max_us_dollars
        ws['B82'] = wind.utility_rebate_us_dollars_per_kw * 0.001
        ws['C82'] = wind.utility_rebate_max_us_dollars
        ws['B84'] = wind.pbi_us_dollars_per_kwh
        ws['C84'] = wind.pbi_max_us_dollars
        ws['E84'] = wind.pbi_years
        ws['F84'] = wind.pbi_system_max_kw

        # Battery Tax Credits and Incentives
        ws['B89'] = batt.total_itc_pct * 100
        ws['C89'] = big_number  # max itc
        ws['B94'] = 0  # state ITC
        ws['C94'] = big_number  # state ITC max
        ws['B95'] = 0  # utility ITC
        ws['C95'] = big_number  # utility ITC max
        ws['B97'] = batt.total_rebate_us_dollars_per_kw * 0.001
        ws['C97'] = big_number  # max rebate
        ws['B98'] = 0  # state rebate
        ws['C98'] = big_number  # max state rebate
        ws['B99'] = 0  # utility rebate
        ws['C99'] = big_number  # max utility rebate

        # Depreciation
        if pv.macrs_option_years > 0:
            ws['B102'] = pv.macrs_option_years
            ws['B103'] = pv.macrs_bonus_pct
        
        elif pv.macrs_option_years == 0:
            ws['B102'] = "None"
            ws['B103'] = 0

        if batt.macrs_option_years > 0:
            ws['C102'] = batt.macrs_option_years
            ws['C103'] = batt.macrs_bonus_pct
        
        elif batt.macrs_option_years == 0:
            ws['C102'] = "None"
            ws['C103'] = 0

        if wind.macrs_option_years > 0:
            ws['D102'] = wind.macrs_option_years
            ws['D103'] = wind.macrs_bonus_pct

        elif wind.macrs_option_years == 0:
            ws['D102'] = "None"
            ws['D103'] = 0

        # this value is being introduced in proforma for modeling variable o&m and diesel cost
        # for back-up generator. Currently in REopt Lite webtool, diesel generator only runs for
        # outage. This flag is used to determine outage occurrences in the analysis period.
        ws['D117'] = int(not loadprofile.outage_is_major_event)

        # Save
        wb.save(self.output_file)

        self.spreadsheet_created = tzlocal.get_localzone().localize(datetime.datetime.now())
        
        return True

