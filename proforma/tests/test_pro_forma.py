# *********************************************************************************
# REopt, Copyright (c) 2019-2020, Alliance for Sustainable Energy, LLC.
# All rights reserved.
#
# Redistribution and use in source and binary forms, with or without modification,
# are permitted provided that the following conditions are met:
#
# Redistributions of source code must retain the above copyright notice, this list
# of conditions and the following disclaimer.
#
# Redistributions in binary form must reproduce the above copyright notice, this
# list of conditions and the following disclaimer in the documentation and/or other
# materials provided with the distribution.
#
# Neither the name of the copyright holder nor the names of its contributors may be
# used to endorse or promote products derived from this software without specific
# prior written permission.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" AND
# ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED
# WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE DISCLAIMED.
# IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE FOR ANY DIRECT,
# INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING,
# BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES; LOSS OF USE,
# DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY THEORY OF
# LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT (INCLUDING NEGLIGENCE
# OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED
# OF THE POSSIBILITY OF SUCH DAMAGE.
# *********************************************************************************
from openpyxl import load_workbook
import tzlocal
import json
import datetime
import os
from collections import defaultdict
from django.test import TestCase
from tastypie.test import ResourceTestCaseMixin
from unittest import skip
from proforma.models import ProForma
from reo.models import ScenarioModel


def now():
    return tzlocal.get_localzone().localize(datetime.datetime.now())


class CashFlowTest(ResourceTestCaseMixin, TestCase):
    """
    These tests just ensure that input cells in Proforma are filled in with the appropriate values.
    Unfotunately, we cannot automatically validate calculated values in the spreadsheet. So, the only way to truely
    validate the Proforma is to visually compare the Proforma LCC's and NPV against the API's results.
    """

    REopt_tol = 5e-5

    def setUp(self):
        super(CashFlowTest, self).setUp()
        self.example_reopt_request_data = json.loads(open('proforma/tests/test_data.json').read())
        self.submit_url = '/v1/job/'
        self.results_url = '/v1/job/<run_uuid>/results/'
        self.proforma_url = '/v1/job/<run_uuid>/proforma/'

    def get_response(self, data):
        initial_post = self.api_client.post(self.submit_url, format='json', data=data)
        uuid = json.loads(initial_post.content)['run_uuid']
        response = json.loads(self.api_client.get(self.results_url.replace('<run_uuid>', str(uuid))).content)
        return response

    def test_full_tech_mix(self):
        """
        Does not test Wind
        """
        run_output = self.get_response(self.example_reopt_request_data)
        uuid = run_output['outputs']['Scenario']['run_uuid']
        mapping = self.get_mapping(run_output, uuid)

        idx = 0
        for a, b in mapping:
            msg = "Failed at: " + str(a) + " Value: " + str(a.value) + "!= " + str(b) + " run_uuid: " + str(uuid)
            self.assertAlmostEqual(float(a.value), b, places=2, msg=msg)
            idx += 1

    @skip("HSDS wind api barely works")
    def test_wind(self):
        self.example_reopt_request_data = json.loads(open('proforma/tests/wind_bug.json').read())
        run_output = self.get_response(self.example_reopt_request_data)
        uuid = run_output['outputs']['Scenario']['run_uuid']

        self.assertGreater(run_output["outputs"]["Scenario"]["Profile"]["pre_setup_scenario_seconds"], 0)
        self.assertGreater(run_output["outputs"]["Scenario"]["Profile"]["setup_scenario_seconds"], 0)
        self.assertGreater(run_output["outputs"]["Scenario"]["Profile"]["reopt_seconds"], 0)
        self.assertGreater(run_output["outputs"]["Scenario"]["Profile"]["reopt_bau_seconds"], 0)
        self.assertGreater(run_output["outputs"]["Scenario"]["Profile"]["parse_run_outputs_seconds"], 0)

        mapping = self.get_mapping(run_output, uuid)

        idx = 0
        for a, b in mapping:
            msg = "Failed at idx: " + str(idx) + " Value: " + str(a.value) + "!= " + str(b) + " run_uuid: " + str(uuid)
            self.assertAlmostEqual(float(a.value), b, places=2, msg=msg)
            idx += 1

    def test_bad_run_uuid(self):
        run_uuid = "5"
        response = json.loads(self.api_client.get(self.results_url.replace('<run_uuid>', str(run_uuid))).content)
        # status = response['outputs']['Scenario']['status']
        self.assertDictEqual(response,
                             {u'outputs': {u'Scenario': {u'status': u'error'}},
                              u'messages': {u'error': u'badly formed hexadecimal UUID string'}})

    def get_mapping(self, run_output, uuid):
        """
        Create a list of lists for Proforma and API values to compare
        :param run_output: dict, API response for job
        :param uuid: job UUID
        :return: list of lists, inner list have pairs of values, where the first value is from the Excel ProForma and
            the second value is from the API database.
        """

        start_time = now()
        response = self.api_client.get(self.proforma_url.replace('<run_uuid>', str(uuid)))
        self.assertEqual(response.status_code, 200)

        scenario = ScenarioModel.objects.get(run_uuid=uuid)
        pf = ProForma.objects.get(scenariomodel=scenario)

        self.assertTrue(os.path.exists(pf.output_file))
        self.assertTrue(pf.spreadsheet_created < now() and pf.spreadsheet_created > start_time)

        wb = load_workbook(pf.output_file, read_only=False, keep_vba=True,data_only=True)
        ws = wb.get_sheet_by_name(pf.sheet_io)

        pv_in = run_output['inputs']['Scenario']['Site'].get('PV', defaultdict(int))
        batt_in = run_output['inputs']['Scenario']['Site'].get('Storage', defaultdict(int))
        wind_in = run_output['inputs']['Scenario']['Site'].get('Wind', defaultdict(int))
        generator_in = run_output['inputs']['Scenario']['Site'].get('Generator', defaultdict(int))
        finance_in = run_output['inputs']['Scenario']['Site']['Financial']
        loadprofile_in = run_output['inputs']['Scenario']['Site']['LoadProfile']

        pv_out = run_output['outputs']['Scenario']['Site'].get('PV', defaultdict(int))
        batt_out = run_output['outputs']['Scenario']['Site'].get('Storage', defaultdict(int))
        wind_out = run_output['outputs']['Scenario']['Site'].get('Wind', defaultdict(int))
        generator_out = run_output['outputs']['Scenario']['Site'].get('Generator', defaultdict(int))
        tariff_out = run_output['outputs']['Scenario']['Site']['ElectricTariff']
        finance_out = run_output['outputs']['Scenario']['Site']['Financial']

        adjusted_pv_kw =pv_out["size_kw"] - pv_in["existing_kw"]
        # with new generator speed up logic, since there is no outage planned in this test-case
        # generator will not be modeled (even if existing kw is non-zero)
        # adjusted_generator_kw = generator_out["size_kw"] - generator_in["existing_kw"]
        adjusted_generator_kw = 0
        # diesel_fuel_used_cost = generator_in["diesel_fuel_cost_us_dollars_per_gallon"]
        #                         * generator_out["fuel_used_gal"]
        diesel_fuel_used_cost = 0

        # Note, cannot evaluate LCC, LCC_BAU, and NPV, since that would require having openpxl evaluate formulas in excel
        mapping = [
            [ws['B3'], adjusted_pv_kw],
            [ws['B4'], pv_in["existing_kw"]],
            [ws['B5'], pv_in["degradation_pct"] * 100],
            [ws['B6'], wind_out["size_kw"]],
            [ws['B7'], adjusted_generator_kw],
            [ws['B8'], generator_in["existing_kw"]],
            [ws['B9'], batt_out["size_kw"]],
            [ws['B10'], batt_out["size_kwh"]],
            [ws['B13'], tariff_out["year_one_bill_bau_us_dollars"]],
            [ws['B14'], tariff_out["year_one_bill_us_dollars"]],
            [ws['B15'], tariff_out["year_one_export_benefit_us_dollars"]],
            [ws['B16'], pv_out["year_one_energy_produced_kwh"]],
            [ws['B17'], wind_out["year_one_energy_produced_kwh"]],
            [ws['B18'], generator_out["year_one_energy_produced_kwh"]],
            [ws['B19'], pv_out["year_one_energy_produced_kwh"] + wind_out["year_one_energy_produced_kwh"] + generator_out["year_one_energy_produced_kwh"]],

            [ws['B22'], adjusted_pv_kw * pv_in["installed_cost_us_dollars_per_kw"] +
             wind_out["size_kw"] * wind_in["installed_cost_us_dollars_per_kw"] +
             batt_out["size_kw"] * batt_in["installed_cost_us_dollars_per_kw"] +
             batt_out["size_kwh"] * batt_in["installed_cost_us_dollars_per_kwh"] +
             adjusted_generator_kw * generator_in["installed_cost_us_dollars_per_kw"]],
            [ws['B23'], adjusted_pv_kw * pv_in["installed_cost_us_dollars_per_kw"]],
            [ws['B24'], wind_out["size_kw"] * wind_in["installed_cost_us_dollars_per_kw"]],
            [ws['B25'], adjusted_generator_kw * generator_in["installed_cost_us_dollars_per_kw"]],
            [ws['B26'], batt_out["size_kw"] * batt_in["installed_cost_us_dollars_per_kw"] +
             batt_out["size_kwh"] * batt_in["installed_cost_us_dollars_per_kwh"]]]

        if 'PV' in run_output['inputs']['Scenario']['Site'].keys():
            mapping.append([ws['B28'], pv_in["om_cost_us_dollars_per_kw"]])
        if 'Wind' in run_output['inputs']['Scenario']['Site'].keys():
            mapping.append([ws['B29'], wind_in["om_cost_us_dollars_per_kw"]])

        gen_data = [[ws['B30'], generator_in["om_cost_us_dollars_per_kw"]],
                    [ws['B31'], generator_in["om_cost_us_dollars_per_kwh"]],
                    [ws['B32'], diesel_fuel_used_cost]]
        mapping.extend(gen_data)

        batt_input = [[ws['B33'], batt_in["replace_cost_us_dollars_per_kw"]],
                      [ws['B34'], batt_in["inverter_replacement_year"]],
                      [ws['B35'], batt_in["replace_cost_us_dollars_per_kwh"]],
                      [ws['B36'], batt_in["battery_replacement_year"]]]
        if 'Storage' in run_output['inputs']['Scenario']['Site'].keys():
            mapping.extend(batt_input)

        finance_inputs = [[ws['B44'], finance_in["analysis_years"]],
                         [ws['B45'], finance_in["om_cost_escalation_pct"] * 100],
                         [ws['B46'], finance_in["escalation_pct"] * 100],
                         [ws['B47'], finance_in["offtaker_discount_pct"] * 100],
                         [ws['B50'], finance_in["offtaker_tax_pct"] * 100]]
        mapping.extend(finance_inputs)

        pv_inputs = [[ws['B55'], pv_in["federal_itc_pct"] * 100],
                     [ws['B60'], pv_in["state_ibi_pct"] * 100],
                     [ws['C60'], pv_in["state_ibi_max_us_dollars"]],
                     [ws['B61'], pv_in["utility_ibi_pct"] * 100],
                     [ws['C61'], pv_in["utility_ibi_max_us_dollars"]],
                     [ws['B63'], pv_in["federal_rebate_us_dollars_per_kw"] * 0.001],
                     [ws['B64'], pv_in["state_rebate_us_dollars_per_kw"] * 0.001],
                     [ws['C64'], pv_in["state_rebate_max_us_dollars"]],
                     [ws['B65'], pv_in["utility_rebate_us_dollars_per_kw"] * 0.001],
                     [ws['C65'], pv_in["utility_rebate_max_us_dollars"]],
                     [ws['B67'], pv_in["pbi_us_dollars_per_kwh"]],
                     [ws['C67'], pv_in["pbi_max_us_dollars"]],
                     [ws['E67'], pv_in["pbi_years"]],
                     [ws['F67'], pv_in["pbi_system_max_kw"]]]

        if 'PV' in run_output['inputs']['Scenario']['Site'].keys():
            mapping.extend(pv_inputs)

        wind_inputs = [[ws['B72'], wind_in["federal_itc_pct"] * 100],
                       [ws['B77'], wind_in["state_ibi_pct"] * 100],
                       [ws['C77'], wind_in["state_ibi_max_us_dollars"]],
                       [ws['B78'], wind_in["utility_ibi_pct"] * 100],
                       [ws['C78'], wind_in["utility_ibi_max_us_dollars"]],
                       [ws['B80'], wind_in["federal_rebate_us_dollars_per_kw"] * 0.001],
                       [ws['B81'], wind_in["state_rebate_us_dollars_per_kw"] * 0.001],
                       [ws['C81'], wind_in["state_rebate_max_us_dollars"]],
                       [ws['B82'], wind_in["utility_rebate_us_dollars_per_kw"] * 0.001],
                       [ws['C82'], wind_in["utility_rebate_max_us_dollars"]],
                       [ws['B84'], wind_in["pbi_us_dollars_per_kwh"]],
                       [ws['C84'], wind_in["pbi_max_us_dollars"]],
                       [ws['E84'], wind_in["pbi_years"]],
                       [ws['F84'], wind_in["pbi_system_max_kw"]]]
        if 'Wind' in run_output['inputs']['Scenario']['Site'].keys():
            mapping.extend(wind_inputs)

        batt_inputs = [[ws['B89'], batt_in["total_itc_pct"] * 100],
                       [ws['B97'], batt_in["total_rebate_us_dollars_per_kw"] * 0.001],
                       [ws['C102'], batt_in["macrs_option_years"]],
                       [ws['C103'], batt_in["macrs_bonus_pct"]]]

        if 'Storage' in run_output['inputs']['Scenario']['Site'].keys():
            mapping.extend(batt_inputs)

        pv_inputs = [[ws['B102'], pv_in["macrs_option_years"]],
                     [ws['B103'], pv_in["macrs_bonus_pct"]]]

        if 'PV' in run_output['inputs']['Scenario']['Site'].keys():
            mapping.extend(pv_inputs)

        wind_inputs = [[ws['D102'], wind_in["macrs_option_years"]],
                       [ws['D103'], wind_in["macrs_bonus_pct"]]]

        if 'Wind' in run_output['inputs']['Scenario']['Site'].keys():
            mapping.extend(wind_inputs)

        mapping.extend([[ws['D118'], int(not loadprofile_in["outage_is_major_event"])]])
        return mapping
