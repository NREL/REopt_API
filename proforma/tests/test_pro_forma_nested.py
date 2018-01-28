from openpyxl import load_workbook
import tzlocal
import json
import datetime
import os
from django.test import TestCase
from tastypie.test import ResourceTestCaseMixin
from django.test import Client
from proforma.models import ProForma
from reo.models import ScenarioModel

def now():
    return tzlocal.get_localzone().localize(datetime.datetime.now())


class EntryResourceTest(ResourceTestCaseMixin, TestCase):

    REopt_tol = 5e-5

    def setUp(self):
        super(EntryResourceTest, self).setUp()

        self.example_reopt_request_data = json.loads(open('proforma/tests/test_data_nested.json').read())
        self.submit_url = '/api/v1/reopt/'
        self.results_url = '/reopt/results/?run_uuid='

    def get_response(self, data):
        initial_post = self.api_client.post(self.submit_url, format='json', data=data)
        uuid = json.loads(initial_post.content)['run_uuid']

        response = json.loads(self.api_client.get(self.results_url + uuid).content)

        # # the following is not needed b/c we test the app with Celery tasks in "eager" mode
        # # i.e. asynchronously. If we move to testing the API, then the while loop is needed
        # status = response['outputs']['Scenario']['status']
        # while status == "Optimizing...":
        #     time.sleep(2)
        #     response = json.loads(self.api_client.get(self.results_url + uuid).content)
        #     status = response['outputs']['Scenario']['status']

        return response

    def test_creation(self):
        run_output = self.get_response(self.example_reopt_request_data)

        class ClassAttributes:
            def __init__(self, dictionary):
                for k, v in dictionary.items():
                    setattr(self, k, v)

        scenario_in = ClassAttributes(run_output['inputs']['Scenario'])
        pv_in = ClassAttributes(run_output['inputs']['Scenario']['Site']['PV'])
        batt_in = ClassAttributes(run_output['inputs']['Scenario']['Site']['Storage'])
        tariff_in = ClassAttributes(run_output['inputs']['Scenario']['Site']['ElectricTariff'])
        finance_in = ClassAttributes(run_output['inputs']['Scenario']['Site']['Financial'])

        scenario_out = ClassAttributes(run_output['outputs']['Scenario'])
        pv_out = ClassAttributes(run_output['outputs']['Scenario']['Site']['PV'])
        batt_out = ClassAttributes(run_output['outputs']['Scenario']['Site']['Storage'])
        tariff_out = ClassAttributes(run_output['outputs']['Scenario']['Site']['ElectricTariff'])
        finance_out = ClassAttributes(run_output['outputs']['Scenario']['Site']['Financial'])

        uuid = run_output['outputs']['Scenario']['run_uuid']

        start_time = now()
        response = Client().get('/proforma/?run_uuid='+uuid)
        
        self.assertEqual(response.status_code,200)

        scenario = ScenarioModel.objects.get(run_uuid=uuid)
        pf = ProForma.objects.get(scenariomodel=scenario)

        self.assertTrue(os.path.exists(pf.output_file))
        self.assertTrue(pf.spreadsheet_created < now() and pf.spreadsheet_created > start_time)

        wb = load_workbook(pf.output_file, read_only=False, keep_vba=True)
        ws = wb.get_sheet_by_name(pf.sheet_io)

        mapping = [
        [ws['B3'], pv_out.size_kw],
        [ws['B4'], pv_in.degradation_pct * 100],
        [ws['B5'], batt_out.size_kw],
        [ws['B6'], batt_out.size_kwh],
        [ws['B9'], tariff_out.year_one_bill_bau_us_dollars],
        [ws['B10'], tariff_out.year_one_bill_us_dollars],
        [ws['B11'], tariff_out.year_one_export_benefit_us_dollars],
        [ws['B12'], pv_out.year_one_energy_produced_kwh],
        [ws['B15'], pv_out.size_kw * pv_in.installed_cost_us_dollars_per_kw +
                    batt_out.size_kw * batt_in.installed_cost_us_dollars_per_kw +
                    batt_out.size_kwh * batt_in.installed_cost_us_dollars_per_kwh],
        [ws['B16'], pv_out.size_kw * pv_in.installed_cost_us_dollars_per_kw],
        [ws['B17'], batt_out.size_kw * batt_in.installed_cost_us_dollars_per_kw +
                    batt_out.size_kwh * batt_in.installed_cost_us_dollars_per_kwh],
        [ws['B19'], pv_in.om_cost_us_dollars_per_kw],
        [ws['B20'], batt_in.replace_cost_us_dollars_per_kw],
        [ws['B21'], batt_in.inverter_replacement_year],
        [ws['B22'], batt_in.replace_cost_us_dollars_per_kwh],
        [ws['B23'], batt_in.battery_replacement_year],
        [ws['B31'], finance_in.analysis_years],
        [ws['B32'], finance_in.om_cost_escalation_pct * 100],
        [ws['B33'], finance_in.escalation_pct * 100],
        [ws['B34'], finance_in.offtaker_discount_pct * 100],
        [ws['B37'], finance_in.offtaker_tax_pct * 100],
        [ws['B42'], pv_in.federal_itc_pct * 100],
        [ws['B47'], pv_in.state_ibi_pct * 100],
        [ws['C47'], pv_in.state_ibi_max_us_dollars],
        [ws['B48'], pv_in.utility_ibi_pct * 100],
        [ws['C48'], pv_in.utility_ibi_max_us_dollars],
        [ws['B50'], pv_in.federal_rebate_us_dollars_per_kw * 0.001],
        [ws['B51'], pv_in.state_rebate_us_dollars_per_kw * 0.001],
        [ws['C51'], pv_in.state_rebate_max_us_dollars],
        [ws['B52'], pv_in.utility_rebate_us_dollars_per_kw * 0.001],
        [ws['C52'], pv_in.utility_rebate_max_us_dollars],
        [ws['B54'], pv_in.pbi_us_dollars_per_kwh],
        [ws['C54'], pv_in.pbi_max_us_dollars],
        [ws['E54'], pv_in.pbi_years],
        [ws['F54'], pv_in.pbi_system_max_kw],
        [ws['B59'], batt_in.total_itc_pct * 100],
        [ws['B67'], batt_in.total_rebate_us_dollars_per_kw * 0.001],
        [ws['B72'], pv_in.macrs_option_years],
        [ws['B73'], pv_in.macrs_bonus_pct],
        [ws['C72'], batt_in.macrs_option_years],
        [ws['C73'], batt_in.macrs_bonus_pct]
        ]

        idx = 0
        for a, b in mapping:
            msg = "Failed at idx: " + str(idx) + " Value: " + str(a.value) + "!= " + str(b)
            self.assertAlmostEqual(float(a.value), b, places=2, msg=msg)
            idx += 1

    def test_bad_run_uuid(self):
        run_uuid = "5"
        response = json.loads(self.api_client.get(self.results_url + run_uuid).content)
        # status = response['outputs']['Scenario']['status']
        self.assertDictEqual(response,
                             {u'outputs': {u'Scenario': {u'status': u'error'}},
                              u'messages': {u'error': u'badly formed hexadecimal UUID string'}})


