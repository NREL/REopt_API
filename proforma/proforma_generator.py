import os
import numpy as np
from openpyxl.styles import PatternFill, Border, Font, Side, Alignment
from reo.models import SiteModel, LoadProfileModel, PVModel, WindModel, GeneratorModel, StorageModel, FinancialModel, \
    ElectricTariffModel, CHPModel, AbsorptionChillerModel, HotTESModel, ColdTESModel, FuelTariffModel, BoilerModel
from openpyxl import load_workbook
from reo.src.data_manager import big_number
from reo.nested_inputs import macrs_five_year, macrs_seven_year

one_party_workbook = os.path.join('proforma', 'REoptCashFlowTemplateOneParty.xlsx')
two_party_workbook = os.path.join('proforma', 'REoptCashFlowTemplateTwoParty.xlsx')


def generate_proforma(scenariomodel, output_file_path):
    """
    In this method we first collect and organize the data from the data bases. Then we set up styles and methods to
    apply styles to columms within a row.
    We use a cursor (current_row) to start at the top of the Inputs and Outputs sheet (referred to as ws) then work down
    filling in values, styles and cell references as necessary.
    We next move down the Developer and Host Cash Flow sheets (referred to as dcs and hcs respectively), filling in only
    references (no hard coded values other than year references).
    Finally, we return to the Inputs and Outputs sheet to fill in LCC, NPV enires.

    As the cursor is being updated we captures rows and cells that need to be later used as variables, or in the case of
    PV's (which can have multiple values), as dictionary entries.
    Examples:

    line 242 wind_size_kw_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)

    line 352 pv_cell_locations[i]["pv_om_cost_us_dollars_per_kw_cell"] = "\'{}\'!B{}".format(inandout_sheet_name,
                                                                                             current_row)

    When copying and pasting sections of code, pay attention to when the current_row cursor is being updated and any
    reference variables that need to be captured.
    """

    ####################################################################################################################
    # Get Data
    ####################################################################################################################

    scenario = scenariomodel
    batt = StorageModel.objects.filter(run_uuid=scenario.run_uuid).first()
    pvs = PVModel.objects.filter(run_uuid=scenario.run_uuid)
    wind = WindModel.objects.filter(run_uuid=scenario.run_uuid).first()
    generator = GeneratorModel.objects.filter(run_uuid=scenario.run_uuid).first()
    electric_tariff = ElectricTariffModel.objects.filter(run_uuid=scenario.run_uuid).first()
    financial = FinancialModel.objects.filter(run_uuid=scenario.run_uuid).first()
    site = SiteModel.objects.filter(run_uuid=scenario.run_uuid).first()
    load = LoadProfileModel.objects.filter(run_uuid=scenario.run_uuid).first()
    chp = CHPModel.objects.filter(run_uuid=scenario.run_uuid).first()
    absorption_chiller = AbsorptionChillerModel.objects.filter(run_uuid=scenario.run_uuid).first()
    cold_tes = HotTESModel.objects.filter(run_uuid=scenario.run_uuid).first()
    hot_tes = ColdTESModel.objects.filter(run_uuid=scenario.run_uuid).first()
    fuel_tariff = FuelTariffModel.objects.filter(run_uuid=scenario.run_uuid).first()
    boiler = BoilerModel.objects.filter(run_uuid=scenario.run_uuid).first()

    # Open file for reading
    if financial.two_party_ownership is True:
        wb = load_workbook(two_party_workbook, read_only=False, keep_vba=True)
        developer_cashflow_sheet_name = 'Developer Cash Flow'
        host_cashflow_sheet_name = 'Host Cash Flow'
    else:
        wb = load_workbook(one_party_workbook, read_only=False, keep_vba=True)
        developer_cashflow_sheet_name = 'Optimal Cash Flow'
        host_cashflow_sheet_name = 'BAU Cash Flow'

    # Open Inputs Sheet
    ws = wb.get_sheet_by_name('Inputs and Outputs')
    inandout_sheet_name = 'Inputs and Outputs'

    # handle None's
    pv_cell_locations = {}
    pv_data = []
    for i, pv in enumerate(pvs):
        pv_cell_locations[i] = {}
        pv_data_entry = {}
        pv_data_entry["pv"] = pv
        pv_data_entry["name"] = 'Site {}'.format(i)
        pv_data_entry["pv_installed_kw"] = pv.size_kw or 0
        if pv_data_entry["pv_installed_kw"] > 0 and pv.existing_kw > 0:
            pv_data_entry["pv_installed_kw"] -= pv.existing_kw
        pv_data_entry["pv_installed_cost_us_dollars_per_kw"] = pv.installed_cost_us_dollars_per_kw or 0
        pv_data_entry["pv_cost"] = pv_data_entry["pv_installed_kw"] \
                                   * pv_data_entry["pv_installed_cost_us_dollars_per_kw"]
        pv_data_entry["pv_energy"] = pv.year_one_energy_produced_kwh or 0
        pv_data_entry["pv_energy_bau"] = pv.year_one_energy_produced_bau_kwh or 0
        pv_data_entry["pv_existing_kw"] = pv.existing_kw or 0
        pv_data.append(pv_data_entry)

    if len(pv_data) == 1:
        pv_data[0]['name'] = 'PV'
    else:
        for pv in pv_data:
            pv['name'] = "PV ({})".format(pv['name'])

    batt_size_kw = batt.size_kw or 0
    batt_size_kwh = batt.size_kwh or 0
    batt_installed_cost_us_dollars_per_kw = batt.installed_cost_us_dollars_per_kw or 0
    batt_installed_cost_us_dollars_per_kwh = batt.installed_cost_us_dollars_per_kwh or 0
    batt_cost = batt_size_kw * batt_installed_cost_us_dollars_per_kw \
                + batt_size_kwh * batt_installed_cost_us_dollars_per_kwh

    wind_installed_cost_us_dollars_per_kw = wind.installed_cost_us_dollars_per_kw or 0
    wind_installed_kw = wind.size_kw or 0
    wind_energy = wind.year_one_energy_produced_kwh or 0
    wind_cost = wind_installed_kw * wind_installed_cost_us_dollars_per_kw

    generator_installed_kw = generator.size_kw or 0
    generator_existing_kw = generator.existing_kw or 0
    gen_fuel_used_gal = generator.fuel_used_gal or 0
    gen_fuel_used_gal_bau = generator.fuel_used_gal_bau or 0
    if generator_installed_kw > 0 and generator_existing_kw > 0:
        generator_installed_kw -= generator_existing_kw
    generator_installed_cost_us_dollars_per_kw = generator.installed_cost_us_dollars_per_kw or 0
    generator_energy = generator.year_one_energy_produced_kwh or 0
    generator_cost = generator_installed_kw * generator_installed_cost_us_dollars_per_kw
    diesel_fuel_used_cost = (generator.diesel_fuel_cost_us_dollars_per_gallon or 0) * gen_fuel_used_gal
    diesel_fuel_used_cost_bau = (generator.diesel_fuel_cost_us_dollars_per_gallon or 0) * gen_fuel_used_gal_bau

    ####################################################################################################################
    # Set up styling
    ####################################################################################################################

    title_font = Font(name='Calibri', size=10, bold=True, italic=False, vertAlign=None, underline='none', strike=False,
                      color='000000')
    title_fill = PatternFill(fill_type='solid', start_color='C2C5CC', end_color='C2C5CC')
    title_border_left = Border(left=Side(border_style='thin', color='000000'),
                               right=Side(border_style=None, color=None),
                               top=Side(border_style='thin', color='000000'),
                               bottom=Side(border_style='thin', color='000000'))
    title_border_right = Border(left=Side(border_style=None, color=None),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
    title_border_middle = Border(left=Side(border_style=None, color=None), right=Side(border_style=None, color=None),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))

    calculated_fill = PatternFill(fill_type='solid', start_color='aec9eb', end_color='aec9eb')
    grey_fill = PatternFill(fill_type='solid', start_color='C2C5CC', end_color='C2C5CC')

    attribute_font = Font(name='Calibri', size=10, bold=False, italic=False, vertAlign=None, underline='none',
                          strike=False, color='000000')
    attribute_fill = PatternFill(fill_type=None)
    attribute_border_left = Border(left=Side(border_style='thin', color='000000'),
                                   right=Side(border_style=None, color=None),
                                   top=Side(border_style='thin', color='000000'),
                                   bottom=Side(border_style='thin', color='000000'))
    attribute_border_left_and_right = Border(left=Side(border_style='thin', color='000000'),
                                             right=Side(border_style='thin', color=None),
                                             top=Side(border_style='thin', color='000000'),
                                             bottom=Side(border_style='thin', color='000000'))
    attribute_border_right = Border(left=Side(border_style=None, color=None),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))
    attribute_border_middle = Border(left=Side(border_style=None, color=None),
                                     right=Side(border_style=None, color=None),
                                     top=Side(border_style='thin', color='000000'),
                                     bottom=Side(border_style='thin', color='000000'))

    border_top_and_bottom = Border(left=Side(border_style=None, color='000000'),
                                   right=Side(border_style=None, color=None),
                                   top=Side(border_style='thin', color='000000'),
                                   bottom=Side(border_style='thin', color='000000'))
    border_top = Border(left=Side(border_style=None, color='000000'), right=Side(border_style=None, color=None),
                        top=Side(border_style='thin', color='000000'), bottom=Side(border_style=None, color='000000'))

    no_border = Border(left=Side(border_style=None, color='000000'), right=Side(border_style=None, color=None),
                       top=Side(border_style=None, color='000000'), bottom=Side(border_style=None, color='000000'))
    bold_font = Font(name='Calibri', size=10, bold=True, italic=False, vertAlign=None, underline='none', strike=False,
                     color='000000')
    one_tab_indent = Alignment(indent=1)
    two_tab_indent = Alignment(indent=2)
    center_align = Alignment(horizontal='center')
    right_align = Alignment(horizontal='right')
    current_row = 2

    number_of_alphabet_iterations = round(financial.analysis_years/26) + 1
    base_upper_case_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                          'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    upper_case_letters = []                         
    for i in range(number_of_alphabet_iterations):
        if i == 0:
            upper_case_letters += base_upper_case_letters
        else:
            upper_case_letters += [base_upper_case_letters[i-1] + l for l  in base_upper_case_letters]

    macrs_schedule = {5: macrs_five_year,
                      7: macrs_seven_year}

    def make_title_row(ws, current_row, length=2, offset=0, alignment=None):
        for i in range(offset, offset + length):
            if i == offset:
                ws['{}{}'.format(upper_case_letters[i], current_row)].font = title_font
                ws['{}{}'.format(upper_case_letters[i], current_row)].fill = title_fill
                ws['{}{}'.format(upper_case_letters[i], current_row)].border = title_border_left
            elif i == length - 1:
                ws['{}{}'.format(upper_case_letters[i], current_row)].alignment = alignment or center_align
                ws['{}{}'.format(upper_case_letters[i], current_row)].fill = title_fill
                ws['{}{}'.format(upper_case_letters[i], current_row)].font = title_font
                ws['{}{}'.format(upper_case_letters[i], current_row)].border = title_border_right
            else:
                ws['{}{}'.format(upper_case_letters[i], current_row)].alignment = alignment or center_align
                ws['{}{}'.format(upper_case_letters[i], current_row)].fill = title_fill
                ws['{}{}'.format(upper_case_letters[i], current_row)].font = title_font
                ws['{}{}'.format(upper_case_letters[i], current_row)].border = title_border_middle

    def make_attribute_row(ws, current_row, length=2, bold_cols=[], offset=0, precision=1, alignment=None, border=None,
                           number_format=None, fill=None):
        for i in range(offset, offset + length):
            if i == offset:
                ws['{}{}'.format(upper_case_letters[i], current_row)].font = attribute_font
                ws['{}{}'.format(upper_case_letters[i], current_row)].fill = fill or attribute_fill
                ws['{}{}'.format(upper_case_letters[i], current_row)].border = border or attribute_border_left
            elif i == (length - 1):
                ws['{}{}'.format(upper_case_letters[i], current_row)].font = attribute_font
                ws['{}{}'.format(upper_case_letters[i], current_row)].fill = fill or attribute_fill
                ws['{}{}'.format(upper_case_letters[i], current_row)].alignment = alignment or center_align
                ws['{}{}'.format(upper_case_letters[i], current_row)].border = border or attribute_border_right
            else:
                ws['{}{}'.format(upper_case_letters[i], current_row)].fill = fill or attribute_fill
                ws['{}{}'.format(upper_case_letters[i], current_row)].font = attribute_font
                ws['{}{}'.format(upper_case_letters[i], current_row)].border = border or attribute_border_left_and_right
                ws['{}{}'.format(upper_case_letters[i], current_row)].alignment = alignment or center_align
            value = ws['{}{}'.format(upper_case_letters[i], current_row)].value

            if type(value) == float or type(value) == int:
                if value < 10 and value > 0 and value % 1 == 0:
                    pass
                else:
                    ws['{}{}'.format(upper_case_letters[i], current_row)].alignment = alignment or right_align
                    ws['{}{}'.format(upper_case_letters[i], current_row)].number_format = number_format or '#,##0'
                    if value > 100000000:
                        ws['{}{}'.format(upper_case_letters[i],
                                         current_row)].number_format = number_format or '0.00E+00'
                    if value < 10 and value > 0:
                        ws['{}{}'.format(upper_case_letters[i], current_row)].number_format = \
                            number_format or '0.' + (precision * '0')
            if i in bold_cols:
                ws['{}{}'.format(upper_case_letters[i], current_row)].font = bold_font
            if number_format is not None:
                ws['{}{}'.format(upper_case_letters[i], current_row)].number_format = number_format

    def fill_cols(ws, series, current_row, fill):
        for i in series:
            ws['{}{}'.format(upper_case_letters[i], current_row)].fill = fill

    def fill_border(ws, series, current_row, border):
        for i in series:
            ws['{}{}'.format(upper_case_letters[i], current_row)].border = border

    def fill_alignment(ws, series, current_row, alignment):
        for i in series:
            ws['{}{}'.format(upper_case_letters[i], current_row)].alignment = alignment

    ####################################################################################################################
    ####################################################################################################################
    # INPUTS AND OUTPUTS SHEET
    ####################################################################################################################
    ####################################################################################################################

    ####################################################################################################################
    # System Design
    ####################################################################################################################

    ws['A{}'.format(current_row)] = "OPTIMAL SYSTEM DESIGN (with existing)"
    make_title_row(ws, current_row)
    current_row += 1

    for i, pv in enumerate(pv_data):  # each pv is a dict created above (not PV django model)
        ws['A{}'.format(current_row)] = "{} Nameplate capacity (kW), purchased".format(pv['name'])
        ws['B{}'.format(current_row)] = pv["pv_installed_kw"]
        pv_cell_locations[i]["pv_size_kw_cell"] = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        make_attribute_row(ws, current_row, alignment=right_align)
        current_row += 1
        ws['A{}'.format(current_row)] = "{} Nameplate capacity (kW), existing".format(pv['name'])
        pv_cell_locations[i]["pv_existing_kw_cell"] = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        ws['B{}'.format(current_row)] = pv_data_entry["pv"].existing_kw or 0  # TODO: replace pv_data_entry?
        make_attribute_row(ws, current_row, alignment=right_align)
        current_row += 1
        ws['A{}'.format(current_row)] = "{} degradation rate (%/year)".format(pv['name'])
        ws['B{}'.format(current_row)] = pv_data_entry["pv"].degradation_pct * 100  # TODO: replace pv_data_entry?
        pv_cell_locations[i]["pv_degradation_rate_cell"] = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        make_attribute_row(ws, current_row, alignment=right_align)
        current_row += 1
        ws['A{}'.format(current_row)] = "{} LCOE of New Capacity ($/kWh), nominal".format(pv['name'])
        pv_cell_locations[i]["pv_lcoe_cell"] = 'B{}'.format(current_row)
        make_attribute_row(ws, current_row, alignment=right_align)
        current_row += 1
     

    ws['A{}'.format(current_row)] = "Wind Nameplate capacity (kW), purchased"
    ws['B{}'.format(current_row)] = wind_installed_kw
    wind_size_kw_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, alignment=right_align)
    current_row += 1

    ws['A{}'.format(current_row)] = "Wind LCOE ($/kWh), nominal"
    wind_lcoe_cell = 'B{}'.format(current_row)
    make_attribute_row(ws, current_row, alignment=right_align)

    current_row += 1
    ws['A{}'.format(current_row)] = "Backup Generator Nameplate capacity (kW), purchased"
    ws['B{}'.format(current_row)] = generator_installed_kw
    generator_size_kw_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, alignment=right_align)

    current_row += 1
    ws['A{}'.format(current_row)] = "Backup Generator Nameplate capacity (kW), existing"
    ws['B{}'.format(current_row)] = generator_existing_kw
    generator_existing_kw_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, alignment=right_align)

    current_row += 1
    ws['A{}'.format(current_row)] = "Battery power (kW)"
    ws['B{}'.format(current_row)] = batt.size_kw or 0
    batt_size_kw_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, alignment=right_align)

    current_row += 1
    ws['A{}'.format(current_row)] = "Battery capacity (kWh)"
    ws['B{}'.format(current_row)] = batt.size_kwh or 0
    batt_size_kwh_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, alignment=right_align)

    current_row += 1
    ws['A{}'.format(current_row)] = "CHP capacity (kW)"
    ws['B{}'.format(current_row)] = chp.size_kw or 0
    chp_size_kw_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, alignment=right_align)

    current_row += 1
    ws['A{}'.format(current_row)] = "Absorption chiller capacity (tons)"
    ws['B{}'.format(current_row)] = absorption_chiller.size_ton or 0
    absorption_chiller_size_ton_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, alignment=right_align)

    current_row += 1
    ws['A{}'.format(current_row)] = "Chilled water TES capacity (gallons)"
    ws['B{}'.format(current_row)] = cold_tes.size_gal or 0
    cold_tes_size_gal_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, alignment=right_align)
    
    current_row += 1
    ws['A{}'.format(current_row)] = "Hot water TES capacity (gallons)"
    ws['B{}'.format(current_row)] = hot_tes.size_gal or 0
    hot_tes_size_gal_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, alignment=right_align)
    current_row += 1
    current_row += 1

    ####################################################################################################################
    # Present value of annual Results
    ####################################################################################################################

    ws['A{}'.format(current_row)] = "ANNUAL RESULTS"
    make_title_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Present value of annual Business as Usual electric utility bill ($/year)"
    ws['B{}'.format(current_row)] = electric_tariff.year_one_bill_bau_us_dollars or 0
    year_one_bau_bill_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Present value of annual Business as Usual export credits ($/year)"
    ws['B{}'.format(current_row)] = -1 * electric_tariff.year_one_export_benefit_bau_us_dollars or 0
    year_one_bau_credits_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Present value of annual Optimal electric utility bill($/year)"
    ws['B{}'.format(current_row)] = electric_tariff.year_one_bill_us_dollars or 0
    year_one_bill_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Present value of annual Optimal export credits ($/year)"
    ws['B{}'.format(current_row)] = -1 * electric_tariff.year_one_export_benefit_us_dollars or 0
    year_one_credits_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1

    for i, pv in enumerate(pv_data):
        ws['A{}'.format(current_row)] = "Existing {} electricity produced (kWh), Year 1".format(pv['name'])
        ws['B{}'.format(current_row)] = pv["pv_energy_bau"]
        pv_cell_locations[i]["pv_energy_bau_cell"] = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        make_attribute_row(ws, current_row)
        current_row += 1

        ws['A{}'.format(current_row)] = "Total {} optimal electricity produced (kWh), Year 1".format(pv['name'])
        ws['B{}'.format(current_row)] = pv["pv_energy"]
        pv_cell_locations[i]["pv_energy_cell"] = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        make_attribute_row(ws, current_row)
        current_row += 1

        ws['A{}'.format(current_row)] = "{} optimal electricity produced (kWh), Annual Average".format(pv['name'])
        ws['B{}'.format(current_row)] = pv["pv"].average_yearly_energy_produced_kwh
        pv_cell_locations[i]["pv_avg_energy_cell"] = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        make_attribute_row(ws, current_row)
        current_row += 1

    ws['A{}'.format(current_row)] = "Nominal annual optimal wind electricity produced (kWh/year)"
    ws['B{}'.format(current_row)] = wind_energy
    wind_energy_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Nominal annual optimal backup generator electricity produced (kWh/year)"
    ws['B{}'.format(current_row)] = generator_energy
    generator_energy_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "CHP annual optimal electricity produced (kWh/year)"
    ws['B{}'.format(current_row)] = chp.year_one_electric_energy_produced_kwh or 0
    chp_electric_energy_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)


    current_row += 1
    ws['A{}'.format(current_row)] = "CHP annual runtime (hours/year)"
    ws['B{}'.format(current_row)] = sum(np.array(chp.year_one_electric_production_series_kw or []) > 0) / (scenario.time_steps_per_hour or 1)
    chp_runtime_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, alignment=right_align)

    current_row += 1
    ws['A{}'.format(current_row)] = "Total optimal electricity produced (kWh/year)"
    ws['B{}'.format(current_row)] = wind_energy + generator_energy + sum([pv['pv_energy'] for pv in pv_data]) + (chp.year_one_electric_energy_produced_kwh or 0)
    make_attribute_row(ws, current_row)
    
    current_row += 1
    ws['A{}'.format(current_row)] = "Present value of annual Business as Usual boiler fuels utility bill ($/year)"
    ws['B{}'.format(current_row)] = fuel_tariff.year_one_boiler_fuel_cost_bau_us_dollars or 0
    boiler_fuel_bill_bau_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)
    
    current_row += 1
    ws['A{}'.format(current_row)] = "Present value of annual Optimal boiler fuels utility bill ($/year)"
    ws['B{}'.format(current_row)] = fuel_tariff.year_one_boiler_fuel_cost_us_dollars or 0
    boiler_fuel_bill_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)
    
    current_row += 1
    ws['A{}'.format(current_row)] = "Present value of annual CHP fuels utility bill ($/year)"
    ws['B{}'.format(current_row)] = fuel_tariff.year_one_chp_fuel_cost_us_dollars or 0
    chp_fuel_bill_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)
    
    current_row += 1
    ws['A{}'.format(current_row)] = "CHP annual optimal thermal energy produced (MMBtu/year)"
    ws['B{}'.format(current_row)] = chp.year_one_thermal_energy_produced_mmbtu or 0
    chp_thermal_energy_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Percent electricity from on-site renewable resources"
    ws['B{}'.format(current_row)] = '=ROUND(({wind_energy_cell} + SUM({pv_cells}))/{total_energy},3)'.format(\
        wind_energy_cell=wind_energy_cell, pv_cells=','.join([pv_cell_locations[i]["pv_avg_energy_cell"] for i in range(len(pvs))]), total_energy=load.annual_calculated_kwh)
    make_attribute_row(ws, current_row, alignment=right_align,
                           number_format='##%')
    
    current_row += 1
    ws['A{}'.format(current_row)] = "Percent reduction in annual electricity bill"
    ws['B{}'.format(current_row)] = '=IF({bau_bill}=0,"N/A",ROUND(({bau_bill} - {optimal_bill})/{bau_bill},2))'.format(
        bau_bill=year_one_bau_bill_cell, 
        optimal_bill=year_one_bill_cell)
    make_attribute_row(ws, current_row, alignment=right_align,
                           number_format='##%')
    
    current_row += 1
    ws['A{}'.format(current_row)] = "Percent reduction in annual fuels bill"
    ws['B{}'.format(current_row)] = '==IF({bau_bill}=0,"N/A",ROUND(({bau_bill} - SUM({optimal_boiler_bill},{optimal_chp_bill}))/{bau_bill},2))'.format(
        bau_bill=boiler_fuel_bill_bau_cell, 
        optimal_boiler_bill=boiler_fuel_bill_cell,
        optimal_chp_bill=chp_fuel_bill_cell
        )
    make_attribute_row(ws, current_row, alignment=right_align,
                           number_format='##%')
    current_row += 1
    ws['A{}'.format(current_row)] = "Year one total site carbon dioxide emissions (lb CO2 equivalent)"
    ws['B{}'.format(current_row)] = site.year_one_emissions_lb_C02
    make_attribute_row(ws, current_row)
    current_row += 1

    ws['A{}'.format(current_row)] = "Year one total site carbon dioxide emissions BAU (lb CO2 equivalent)"
    ws['B{}'.format(current_row)] = site.year_one_emissions_bau_lb_C02
    make_attribute_row(ws, current_row)
    current_row += 1
    
    ws['A{}'.format(current_row)] = "Year one total carbon dioxide emissions from electric utility purchases (lb CO2 equivalent)"
    ws['B{}'.format(current_row)] = electric_tariff.year_one_emissions_lb_C02
    make_attribute_row(ws, current_row)
    current_row += 1
    
    ws['A{}'.format(current_row)] = "Year one total carbon dioxide emissions from electric utility purchases BAU (lb CO2 equivalent)"
    ws['B{}'.format(current_row)] = electric_tariff.year_one_emissions_bau_lb_C02
    make_attribute_row(ws, current_row)
    current_row += 1
    
    ws['A{}'.format(current_row)] = "Year one total carbon dioxide emissions from backup generator use (lb CO2 equivalent)"
    ws['B{}'.format(current_row)] = generator.year_one_emissions_lb_C02
    make_attribute_row(ws, current_row)
    current_row += 1
    
    ws['A{}'.format(current_row)] = "Year one total carbon dioxide emissions from backup generator use BAU (lb CO2 equivalent)"
    ws['B{}'.format(current_row)] = generator.year_one_emissions_bau_lb_C02
    make_attribute_row(ws, current_row)
    current_row += 1

    ws['A{}'.format(current_row)] = "Year one total carbon dioxide emissions from boiler fuels (lb CO2 equivalent)"
    ws['B{}'.format(current_row)] = boiler.year_one_emissions_lb_C02
    make_attribute_row(ws, current_row)
    current_row += 1

    ws['A{}'.format(current_row)] = "Year one total carbon dioxide emissions from boiler fuels BAU (lb CO2 equivalent)"
    ws['B{}'.format(current_row)] = boiler.year_one_emissions_bau_lb_C02
    make_attribute_row(ws, current_row)
    current_row += 1

    ws['A{}'.format(current_row)] = "Year one total carbon dioxide emissions from CHP fuels (lb CO2 equivalent)"
    ws['B{}'.format(current_row)] = chp.year_one_emissions_lb_C02
    make_attribute_row(ws, current_row)
    current_row += 1
    current_row += 1

    ####################################################################################################################
    # System Costs
    ####################################################################################################################

    ws['A{}'.format(current_row)] = "SYSTEM COSTS"
    make_title_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Total Installed Cost ($)"
    ws['B{}'.format(current_row)] = financial.initial_capital_costs
    installed_costs_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    for i, pv in enumerate(pv_data):
        ws['A{}'.format(current_row)] = "{} Installed Cost ($)".format(pv['name'])
        ws['B{}'.format(current_row)] = pv["pv_cost"]
        pv_cell_locations[i]["pv_cost_cell"] = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        make_attribute_row(ws, current_row)
        current_row += 1

    ws['A{}'.format(current_row)] = "Wind Installed Cost ($)"
    ws['B{}'.format(current_row)] = wind_cost
    wind_cost_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Backup generator Installed Cost ($)"
    ws['B{}'.format(current_row)] = generator_cost
    generator_cost_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Battery Installed Cost ($)"
    ws['B{}'.format(current_row)] = batt_cost
    batt_cost_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "CHP Installed Cost ($)"
    ws['B{}'.format(current_row)] = (chp.installed_cost_us_dollars_per_kw or 0) * (chp.size_kw or 0)
    chp_cost_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Absorption Chiller Installed Cost ($)"
    ws['B{}'.format(current_row)] = (absorption_chiller.installed_cost_us_dollars_per_ton or 0) * (absorption_chiller.size_ton or 0)
    absorption_chiller_cost_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Hot water TES Installed Cost ($)"
    ws['B{}'.format(current_row)] = (hot_tes.installed_cost_us_dollars_per_gal or 0) * (hot_tes.size_gal or 0)
    hot_tes_cost_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Chilled water TES Installed Cost ($)"
    ws['B{}'.format(current_row)] = (cold_tes.installed_cost_us_dollars_per_gal or 0) * (cold_tes.size_gal or 0)
    cold_tes_cost_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Operation and Maintenance (O&M)"
    make_attribute_row(ws, current_row)
    ws['A{}'.format(current_row)].font = bold_font

    current_row += 1
    for i, pv in enumerate(pv_data):
        ws['A{}'.format(current_row)] = "Fixed {} O&M ($/kW-yr)".format(pv['name'])
        ws['A{}'.format(current_row)].alignment = one_tab_indent
        ws['B{}'.format(current_row)] = pv['pv'].om_cost_us_dollars_per_kw
        pv_cell_locations[i]["pv_om_cost_us_dollars_per_kw_cell"] = "\'{}\'!B{}".format(inandout_sheet_name,
                                                                                        current_row)
        make_attribute_row(ws, current_row)
        current_row += 1

    ws['A{}'.format(current_row)] = "Fixed Wind O&M ($/kW-yr)"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = wind.om_cost_us_dollars_per_kw
    wind_om_cost_us_dollars_per_kw_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Fixed Backup Generator O&M ($/kW-yr)"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = generator.om_cost_us_dollars_per_kw
    generator_om_cost_us_dollars_per_kw_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)
    # TODO: report year 1 O&M costs from REopt instead of calculating them?

    current_row += 1
    ws['A{}'.format(current_row)] = "Variable Backup Generator O&M ($/kWh)"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = generator.om_cost_us_dollars_per_kwh
    generator_om_cost_us_dollars_per_kwh_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Diesel fuel used cost ($)"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = diesel_fuel_used_cost
    diesel_fuel_used_cost_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, alignment=right_align)

    current_row += 1
    ws['A{}'.format(current_row)] = "Diesel BAU fuel used cost ($)"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = diesel_fuel_used_cost_bau
    gen_fuel_used_cost_bau_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Battery replacement cost ($/kW)"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = batt.replace_cost_us_dollars_per_kw
    batt_replace_cost_us_dollars_per_kw_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Battery kW replacement year"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = batt.inverter_replacement_year
    batt_replace_year_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Battery replacement cost ($/kWh)"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = batt.replace_cost_us_dollars_per_kwh
    batt_replace_cost_us_dollars_per_kwh_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Battery kWh replacement year"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = batt.battery_replacement_year
    make_attribute_row(ws, current_row)
    
    current_row += 1
    ws['A{}'.format(current_row)] = "Fixed CHP O&M cost ($/kW-yr)"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = chp.om_cost_us_dollars_per_kw or 0
    chp_om_cost_us_dollars_per_kw_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Variable CHP O&M cost ($/kWh)"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = chp.om_cost_us_dollars_per_kwh or 0
    chp_om_cost_us_dollars_per_kwh_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)
    
    current_row += 1
    ws['A{}'.format(current_row)] = "Runtime CHP O&M cost ($/kW-rated/run-hour)"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = chp.om_cost_us_dollars_per_hr_per_kw_rated or 0
    chp_om_cost_us_dollars_per_kw_rated__cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, number_format='0.000')
    
    current_row += 1
    ws['A{}'.format(current_row)] = "Fixed Absorption Chiller O&M cost ($/ton-yr)"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = absorption_chiller.om_cost_us_dollars_per_ton or 0
    absorption_chiller_om_cost_us_dollars_per_ton_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)
    
    current_row += 1
    ws['A{}'.format(current_row)] = "Fixed Chilled water TES O&M cost ($/gallon-year)"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = cold_tes.om_cost_us_dollars_per_gal or 0 
    cold_tes_om_cost_us_dollars_per_gal_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)
    
    current_row += 1
    ws['A{}'.format(current_row)] = "Fixed Hot water TES O&M cost ($/gallon-year)"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = hot_tes.om_cost_us_dollars_per_gal or 0
    hot_tes_om_cost_us_dollars_per_gal_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)
    current_row += 1
    current_row += 1

    ####################################################################################################################
    # Analysis Parameters
    ####################################################################################################################

    ws['A{}'.format(current_row)] = "ANALYSIS PARAMETERS"
    make_title_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Analysis period (years)"
    ws['B{}'.format(current_row)] = financial.analysis_years
    analysis_period_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Nominal O&M cost escalation rate (%/year)"
    ws['B{}'.format(current_row)] = financial.om_cost_escalation_pct * 100
    om_escalation_rate_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, alignment=right_align)

    current_row += 1
    ws['A{}'.format(current_row)] = "Nominal electric utility cost escalation rate (%/year)"
    ws['B{}'.format(current_row)] = financial.escalation_pct * 100
    escalation_pct_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Nominal boiler fuel cost escalation rate (%/year)"
    ws['B{}'.format(current_row)] = financial.boiler_fuel_escalation_pct * 100
    boiler_escalation_pct_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Nominal CHP fuel cost escalation rate (%/year)"
    ws['B{}'.format(current_row)] = financial.chp_fuel_escalation_pct * 100
    chp_escalation_pct_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Nominal discount rate (%/year)"
    ws['B{}'.format(current_row)] = financial.offtaker_discount_pct * 100
    discount_rate_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    host_discount_rate_cell = discount_rate_cell
    make_attribute_row(ws, current_row)
    
    # NOTE: the following two party logic relies on no lines being inserted here!
    if financial.two_party_ownership:
        ws['A{}'.format(current_row)] = "Nominal Developer discount rate (%/year)"
        ws['B{}'.format(current_row)] = financial.owner_discount_pct * 100
        make_attribute_row(ws, current_row, alignment=right_align)

        current_row += 1
        ws['A{}'.format(current_row)] = "Nominal Host discount rate (%/year)"
        ws['B{}'.format(current_row)] = financial.offtaker_discount_pct * 100
        host_discount_rate_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        make_attribute_row(ws, current_row, alignment=right_align)

    current_row += 1
    current_row += 1

    ####################################################################################################################
    # Tax Rates
    ####################################################################################################################

    ws['A{}'.format(current_row)] = "TAX AND INSURANCE PARAMETERS"
    make_title_row(ws, current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Federal income tax rate (%)"
    ws['B{}'.format(current_row)] = financial.offtaker_tax_pct * 100
    fed_tax_rate_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    host_fed_tax_rate_cell = fed_tax_rate_cell
    make_attribute_row(ws, current_row)

    if financial.two_party_ownership:
        ws['A{}'.format(current_row)] = "Developer Federal income tax rate (%)"
        ws['B{}'.format(current_row)] = financial.owner_tax_pct * 100
        developer_fed_tax_rate_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        make_attribute_row(ws, current_row)

        current_row += 1
        ws['A{}'.format(current_row)] = "Host Federal income tax rate (%)"
        ws['B{}'.format(current_row)] = financial.offtaker_tax_pct * 100
        host_fed_tax_rate_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        make_attribute_row(ws, current_row)

    current_row += 1
    current_row += 1

    ####################################################################################################################
    # PV
    ####################################################################################################################

    for i, pv in enumerate(pv_data):
        ws['A{}'.format(current_row)] = "{} TAX CREDITS".format(pv['name'])
        make_title_row(ws, current_row, length=4)
        current_row += 1
        ws['A{}'.format(current_row)] = "Investment tax credit (ITC)"
        ws['A{}'.format(current_row)].font = bold_font
        ws['D{}'.format(current_row)] = "Reduces depreciation"
        ws['D{}'.format(current_row)].border = attribute_border_left
        ws['D{}'.format(current_row)].font = attribute_font

        current_row += 1
        ws['A{}'.format(current_row)] = "As percentage"
        ws['A{}'.format(current_row)].alignment = one_tab_indent
        ws['B{}'.format(current_row)] = "%"
        ws['C{}'.format(current_row)] = "Maximum"
        ws['D{}'.format(current_row)] = "Federal"
        make_attribute_row(ws, current_row, length=4)

        current_row += 1
        ws['A{}'.format(current_row)] = "Federal"
        ws['A{}'.format(current_row)].alignment = two_tab_indent
        ws['B{}'.format(current_row)] = pv['pv'].federal_itc_pct * 100
        pv_cell_locations[i]["pv_itc_fed_percent_cell"] = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        ws['C{}'.format(current_row)] = big_number
        pv_cell_locations[i]["pv_itc_fed_percent_maxvalue_cell"] = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
        ws['D{}'.format(current_row)] = 'Yes'
        pv_cell_locations[i]["pv_itc_fed_percent_deprbas_fed_cell"] = "\'{}\'!D{}".format(inandout_sheet_name,
                                                                                          current_row)
        make_attribute_row(ws, current_row, length=4)
        current_row += 1
        current_row += 1

        ws['A{}'.format(current_row)] = "{} DIRECT CASH INCENTIVES".format(pv['name'])
        make_title_row(ws, current_row, length=5)
        current_row += 1
        ws['A{}'.format(current_row)] = "Investment based incentive (IBI)"
        ws['D{}'.format(current_row)] = "Incentive is taxable"
        ws['E{}'.format(current_row)] = "Reduces depreciation and ITC basis"
        make_attribute_row(ws, current_row, length=5, bold_cols=[0])

        current_row += 1
        ws['A{}'.format(current_row)] = "As percentage"
        ws['A{}'.format(current_row)].alignment = one_tab_indent
        ws['B{}'.format(current_row)] = "%"
        ws['C{}'.format(current_row)] = "Maximum ($)"
        ws['D{}'.format(current_row)] = "Federal"
        ws['E{}'.format(current_row)] = "Federal"
        make_attribute_row(ws, current_row, length=5)

        current_row += 1
        ws['A{}'.format(current_row)] = "State (% of total installed cost)"
        ws['A{}'.format(current_row)].alignment = two_tab_indent
        ws['B{}'.format(current_row)] = pv['pv'].state_ibi_pct * 100
        pv_cell_locations[i]["pv_state_ibi_pct_cell"] = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        ws['C{}'.format(current_row)] = pv['pv'].state_ibi_max_us_dollars
        pv_cell_locations[i]["pv_state_ibi_max_cell"] = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
        ws['D{}'.format(current_row)] = "No"
        pv_cell_locations[i]["pv_ibi_sta_percent_tax_fed_cell"] = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
        ws['E{}'.format(current_row)] = "Yes"
        pv_cell_locations[i]["pv_ibi_sta_percent_deprbas_fed_cell"] = "\'{}\'!E{}".format(inandout_sheet_name,
                                                                                          current_row)
        make_attribute_row(ws, current_row, length=5)

        current_row += 1
        ws['A{}'.format(current_row)] = "Utility (% of total installed cost)"
        ws['A{}'.format(current_row)].alignment = two_tab_indent
        ws['B{}'.format(current_row)] = pv['pv'].utility_ibi_pct * 100
        pv_cell_locations[i]["pv_utility_ibi_pct_cell"] = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        ws['C{}'.format(current_row)] = pv['pv'].utility_ibi_max_us_dollars
        pv_cell_locations[i]["pv_utility_ibi_max_cell"] = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
        ws['D{}'.format(current_row)] = "No"
        pv_cell_locations[i]["pv_ibi_uti_percent_tax_fed_cell"] = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
        ws['E{}'.format(current_row)] = "Yes"
        pv_cell_locations[i]["pv_ibi_uti_percent_deprbas_fed_cell"] = "\'{}\'!E{}".format(inandout_sheet_name,
                                                                                          current_row)
        make_attribute_row(ws, current_row, length=5)

        current_row += 1
        ws['A{}'.format(current_row)] = "Capacity based incentive (CBI)"
        ws['A{}'.format(current_row)].font = bold_font
        ws['B{}'.format(current_row)] = "Amount ($/W)"
        ws['B{}'.format(current_row)].border = attribute_border_left_and_right
        ws['C{}'.format(current_row)] = "Maximum ($)"
        ws['C{}'.format(current_row)].border = attribute_border_left_and_right
        make_attribute_row(ws, current_row, length=5, bold_cols=[0])

        current_row += 1
        ws['A{}'.format(current_row)] = "Federal ($/W)"
        ws['A{}'.format(current_row)].alignment = two_tab_indent
        ws['B{}'.format(current_row)] = pv['pv'].federal_rebate_us_dollars_per_kw * 0.001
        pv_cell_locations[i]["pv_federal_cbi_cell"] = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        ws['C{}'.format(current_row)] = big_number
        pv_cell_locations[i]["pv_federal_cbi_max_cell"] = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
        ws['D{}'.format(current_row)] = "No"
        pv_cell_locations[i]["pv_cbi_fed_tax_fed_cell"] = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
        ws['E{}'.format(current_row)] = "No"
        pv_cell_locations[i]["pv_cbi_fed_deprbas_fed_cell"] = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
        make_attribute_row(ws, current_row, length=5)

        current_row += 1
        ws['A{}'.format(current_row)] = "State  ($/W)"
        ws['A{}'.format(current_row)].alignment = two_tab_indent
        ws['B{}'.format(current_row)] = pv['pv'].state_rebate_us_dollars_per_kw * 0.001
        pv_cell_locations[i]["pv_state_cbi_cell"] = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        ws['C{}'.format(current_row)] = pv['pv'].state_rebate_max_us_dollars
        pv_cell_locations[i]["pv_state_cbi_max_cell"] = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
        ws['D{}'.format(current_row)] = "No"
        pv_cell_locations[i]["pv_cbi_sta_tax_fed_cell"] = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
        ws['E{}'.format(current_row)] = "Yes"
        pv_cell_locations[i]["pv_cbi_sta_deprbas_fed_cell"] = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
        make_attribute_row(ws, current_row, length=5)

        current_row += 1
        ws['A{}'.format(current_row)] = "Utility  ($/W)"
        ws['A{}'.format(current_row)].alignment = two_tab_indent
        ws['B{}'.format(current_row)] = pv['pv'].utility_rebate_us_dollars_per_kw * 0.001
        pv_cell_locations[i]["pv_utility_cbi_cell"] = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        ws['C{}'.format(current_row)] = pv['pv'].utility_rebate_max_us_dollars
        pv_cell_locations[i]["pv_utility_cbi_max_cell"] = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
        ws['D{}'.format(current_row)] = "No"
        pv_cell_locations[i]["pv_cbi_uti_tax_fed_cell"] = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
        ws['E{}'.format(current_row)] = "Yes"
        pv_cell_locations[i]["pv_cbi_uti_deprbas_fed_cell"] = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
        make_attribute_row(ws, current_row, length=5)

        current_row += 1
        ws['A{}'.format(current_row)] = "Production based incentive (PBI)"
        ws['A{}'.format(current_row)].font = bold_font
        ws['B{}'.format(current_row)] = "Amount ($/kWh)"
        ws['B{}'.format(current_row)].border = attribute_border_left_and_right
        ws['C{}'.format(current_row)] = "Maximum ($/year)"
        ws['C{}'.format(current_row)].border = attribute_border_left_and_right
        ws['D{}'.format(current_row)] = "Federal Taxable"
        ws['D{}'.format(current_row)].border = attribute_border_left_and_right
        ws['E{}'.format(current_row)] = "Term (years)"
        ws['E{}'.format(current_row)].border = attribute_border_left_and_right
        ws['F{}'.format(current_row)] = "System Size Limit (kW)"
        ws['F{}'.format(current_row)].border = attribute_border_left_and_right
        make_attribute_row(ws, current_row, length=6, bold_cols=[0])

        current_row += 1
        ws['A{}'.format(current_row)] = "Combined ($/kWh)"
        ws['A{}'.format(current_row)].alignment = two_tab_indent
        ws['B{}'.format(current_row)] = pv['pv'].pbi_us_dollars_per_kwh
        pv_cell_locations[i]["pv_pbi_cell"] = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
        ws['C{}'.format(current_row)] = pv['pv'].pbi_max_us_dollars
        pv_cell_locations[i]["pv_pbi_max_cell"] = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
        ws['D{}'.format(current_row)] = "Yes"
        pv_cell_locations[i]["pv_pbi_combined_tax_fed_cell"] = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
        ws['E{}'.format(current_row)] = pv['pv'].pbi_years
        pv_cell_locations[i]["pv_pbi_years_cell"] = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
        ws['F{}'.format(current_row)] = pv['pv'].pbi_system_max_kw
        pv_cell_locations[i]["pv_pbi_max_kw_cell"] = "\'{}\'!F{}".format(inandout_sheet_name, current_row)
        make_attribute_row(ws, current_row, length=6)
        current_row += 1
        current_row += 1

    ####################################################################################################################
    # Wind
    ####################################################################################################################

    ws['A{}'.format(current_row)] = "WIND TAX CREDITS"
    make_title_row(ws, current_row, length=4)

    current_row += 1
    ws['A{}'.format(current_row)] = "Investment tax credit (ITC)"
    ws['A{}'.format(current_row)].font = bold_font
    ws['D{}'.format(current_row)] = "Reduces depreciation and ITC basis"
    ws['D{}'.format(current_row)].border = attribute_border_left
    ws['D{}'.format(current_row)].font = attribute_font

    current_row += 1
    ws['A{}'.format(current_row)] = "As percentage"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = "%"
    ws['C{}'.format(current_row)] = "Maximum"
    ws['D{}'.format(current_row)] = "Federal"
    make_attribute_row(ws, current_row, length=4)

    current_row += 1
    ws['A{}'.format(current_row)] = "Federal"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = wind.federal_itc_pct * 100
    wind_federal_itc_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = big_number
    wind_itc_fed_percent_maxvalue_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = 'Yes'
    wind_itc_fed_percent_deprbas_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=4)
    current_row += 1
    current_row += 1

    ws['A{}'.format(current_row)] = "WIND DIRECT CASH INCENTIVES"
    make_title_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "Investment based incentive (IBI)"
    ws['D{}'.format(current_row)] = "Incentive is taxable"
    ws['E{}'.format(current_row)] = "Reduces depreciation and ITC basis"
    make_attribute_row(ws, current_row, length=5, bold_cols=[0])

    current_row += 1
    ws['A{}'.format(current_row)] = "As percentage"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = "%"
    ws['C{}'.format(current_row)] = "Maximum ($)"
    ws['D{}'.format(current_row)] = "Federal"
    ws['E{}'.format(current_row)] = "Federal"
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "State (% of total installed cost)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = wind.state_ibi_pct * 100
    wind_state_ibi_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = wind.state_ibi_max_us_dollars
    wind_state_ibi_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "No"
    wind_ibi_sta_percent_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = "Yes"
    wind_ibi_sta_percent_deprbas_fed_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "Utility (% of total installed cost)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = wind.utility_ibi_pct * 100
    wind_utility_ibi_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = wind.utility_ibi_max_us_dollars
    wind_utility_ibi_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "No"
    wind_ibi_uti_percent_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = "Yes"
    wind_ibi_uti_percent_deprbas_fed_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "Capacity based incentive (CBI)"
    ws['A{}'.format(current_row)].font = bold_font
    ws['B{}'.format(current_row)] = "Amount ($/W)"
    ws['B{}'.format(current_row)].border = attribute_border_left_and_right
    ws['C{}'.format(current_row)] = "Maximum ($)"
    ws['C{}'.format(current_row)].border = attribute_border_left_and_right
    make_attribute_row(ws, current_row, length=5, bold_cols=[0])

    current_row += 1
    ws['A{}'.format(current_row)] = "Federal ($/W)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = wind.federal_rebate_us_dollars_per_kw * 0.001
    wind_federal_cbi_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = big_number
    wind_federal_cbi_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "No"
    wind_cbi_sta_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = "No"
    wind_cbi_fed_deprbas_fed_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "State  ($/W)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = wind.state_rebate_us_dollars_per_kw * 0.001
    wind_state_cbi_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = wind.state_rebate_max_us_dollars
    wind_state_cbi_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "No"
    wind_cbi_fed_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = "Yes"
    wind_cbi_sta_deprbas_fed_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "Utility  ($/W)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = wind.utility_rebate_us_dollars_per_kw * 0.001
    wind_utility_cbi_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = wind.utility_rebate_max_us_dollars
    wind_utility_cbi_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "No"
    wind_cbi_uti_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = "Yes"
    wind_cbi_uti_deprbas_fed_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "Production based incentive (PBI)"
    ws['A{}'.format(current_row)].font = bold_font
    ws['B{}'.format(current_row)] = "Amount ($/kWh)"
    ws['B{}'.format(current_row)].border = attribute_border_left_and_right
    ws['C{}'.format(current_row)] = "Maximum ($/year)"
    ws['C{}'.format(current_row)].border = attribute_border_left_and_right
    ws['D{}'.format(current_row)] = "Federal Taxable"
    ws['D{}'.format(current_row)].border = attribute_border_left_and_right
    ws['E{}'.format(current_row)] = "Term (years)"
    ws['E{}'.format(current_row)].border = attribute_border_left_and_right
    ws['F{}'.format(current_row)] = "System Size Limit (kW)"
    ws['F{}'.format(current_row)].border = attribute_border_left_and_right
    make_attribute_row(ws, current_row, length=6, bold_cols=[0])

    current_row += 1
    ws['A{}'.format(current_row)] = "Combined ($/kWh)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = wind.pbi_us_dollars_per_kwh
    wind_pbi_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = wind.pbi_max_us_dollars
    wind_pbi_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "Yes"
    wind_pbi_combined_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = wind.pbi_years
    wind_pbi_years_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    ws['F{}'.format(current_row)] = wind.pbi_system_max_kw
    wind_pbi_max_kw_cell = "\'{}\'!F{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=6)
    current_row += 1
    current_row += 1

    ####################################################################################################################
    # Battery
    ####################################################################################################################

    ws['A{}'.format(current_row)] = "BATTERY TAX CREDITS"
    make_title_row(ws, current_row, length=4)

    current_row += 1
    ws['A{}'.format(current_row)] = "Investment tax credit (ITC)"
    ws['A{}'.format(current_row)].font = bold_font
    ws['D{}'.format(current_row)] = "Reduces depreciation and ITC basis"
    ws['D{}'.format(current_row)].border = attribute_border_left
    ws['D{}'.format(current_row)].font = attribute_font

    current_row += 1
    ws['A{}'.format(current_row)] = "As percentage"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = "%"
    ws['C{}'.format(current_row)] = "Maximum"
    ws['D{}'.format(current_row)] = "Federal"
    make_attribute_row(ws, current_row, length=4)

    current_row += 1
    ws['A{}'.format(current_row)] = "Federal"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = batt.total_itc_pct * 100
    batt_federal_itc_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = big_number
    batt_itc_fed_percent_maxvalue_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = 'Yes'
    batt_itc_fed_percent_deprbas_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=4)
    current_row += 1
    current_row += 1

    ws['A{}'.format(current_row)] = "BATTERY DIRECT CASH INCENTIVES"
    make_title_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "Investment based incentive (IBI)"
    ws['D{}'.format(current_row)] = "Incentive is taxable"
    ws['E{}'.format(current_row)] = "Reduces depreciation and ITC basis"
    make_attribute_row(ws, current_row, length=5, bold_cols=[0])

    current_row += 1
    ws['A{}'.format(current_row)] = "As percentage"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = "%"
    ws['C{}'.format(current_row)] = "Maximum ($)"
    ws['D{}'.format(current_row)] = "Federal"
    ws['E{}'.format(current_row)] = "Federal"
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "State (% of total installed cost)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = 0
    batt_state_ibi_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = big_number
    batt_state_ibi_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "No"
    batt_ibi_sta_percent_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = "No"
    batt_ibi_sta_percent_deprbas_fed_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "Utility (% of total installed cost)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = 0
    batt_utility_ibi_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = big_number
    batt_utility_ibi_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "No"
    batt_ibi_uti_percent_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = "No"
    batt_ibi_uti_percent_deprbas_fed_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "Capacity based incentive (CBI)"
    ws['A{}'.format(current_row)].font = bold_font
    ws['B{}'.format(current_row)] = "Amount ($/W)"
    ws['B{}'.format(current_row)].border = attribute_border_left_and_right
    ws['C{}'.format(current_row)] = "Maximum ($)"
    ws['C{}'.format(current_row)].border = attribute_border_left_and_right
    make_attribute_row(ws, current_row, length=5, bold_cols=[0])
    
    current_row += 1
    ws['A{}'.format(current_row)] = "Total ($/W)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = batt.total_rebate_us_dollars_per_kw / 1000.0
    batt_total_cbi_per_kw_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = big_number
    batt_total_cbi_per_kw_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "No"
    batt_total_cbi_per_kw_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = "No"
    batt_total_cbi_per_kw_deprbas_fed_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "Total  ($/Wh)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = batt.total_rebate_us_dollars_per_kwh / 1000.0 
    batt_total_cbi_per_kwh_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = big_number
    batt_total_cbi_per_kwh_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "No"
    batt_total_cbi_per_kwh_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = "No"
    batt_total_cbi_per_kwh_deprbas_fed_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=5)
    current_row += 1
    current_row += 1

    ####################################################################################################################
    # CHP
    ####################################################################################################################

    ws['A{}'.format(current_row)] = "CHP TAX CREDITS"
    make_title_row(ws, current_row, length=4)

    current_row += 1
    ws['A{}'.format(current_row)] = "Investment tax credit (ITC)"
    ws['A{}'.format(current_row)].font = bold_font
    ws['D{}'.format(current_row)] = "Reduces depreciation and ITC basis"
    ws['D{}'.format(current_row)].border = attribute_border_left
    ws['D{}'.format(current_row)].font = attribute_font

    current_row += 1
    ws['A{}'.format(current_row)] = "As percentage"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = "%"
    ws['C{}'.format(current_row)] = "Maximum"
    ws['D{}'.format(current_row)] = "Federal"
    make_attribute_row(ws, current_row, length=4)

    current_row += 1
    ws['A{}'.format(current_row)] = "Federal"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = chp.federal_itc_pct * 100
    chp_federal_itc_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = big_number
    chp_itc_fed_percent_maxvalue_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = 'Yes'
    chp_itc_fed_percent_deprbas_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=4)
    current_row += 1
    current_row += 1

    ws['A{}'.format(current_row)] = "CHP DIRECT CASH INCENTIVES"
    make_title_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "Investment based incentive (IBI)"
    ws['D{}'.format(current_row)] = "Incentive is taxable"
    ws['E{}'.format(current_row)] = "Reduces depreciation and ITC basis"
    make_attribute_row(ws, current_row, length=5, bold_cols=[0])

    current_row += 1
    ws['A{}'.format(current_row)] = "As percentage"
    ws['A{}'.format(current_row)].alignment = one_tab_indent
    ws['B{}'.format(current_row)] = "%"
    ws['C{}'.format(current_row)] = "Maximum ($)"
    ws['D{}'.format(current_row)] = "Federal"
    ws['E{}'.format(current_row)] = "Federal"
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "State (% of total installed cost)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = chp.state_ibi_pct
    chp_state_ibi_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = big_number
    chp_state_ibi_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "No"
    chp_ibi_sta_percent_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = "No"
    chp_ibi_sta_percent_deprbas_fed_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "Utility (% of total installed cost)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = chp.utility_ibi_pct
    chp_utility_ibi_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = big_number
    chp_utility_ibi_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "No"
    chp_ibi_uti_percent_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = "No"
    chp_ibi_uti_percent_deprbas_fed_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "Capacity based incentive (CBI)"
    ws['A{}'.format(current_row)].font = bold_font
    ws['B{}'.format(current_row)] = "Amount ($/W)"
    ws['B{}'.format(current_row)].border = attribute_border_left_and_right
    ws['C{}'.format(current_row)] = "Maximum ($)"
    ws['C{}'.format(current_row)].border = attribute_border_left_and_right
    make_attribute_row(ws, current_row, length=5, bold_cols=[0])

    current_row += 1
    ws['A{}'.format(current_row)] = "Federal ($/W)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = chp.federal_rebate_us_dollars_per_kw * 0.001
    chp_federal_cbi_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = big_number
    chp_federal_cbi_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "No"
    chp_cbi_sta_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = "No"
    chp_cbi_fed_deprbas_fed_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "State  ($/W)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = chp.state_rebate_us_dollars_per_kw * 0.001
    chp_state_cbi_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = chp.state_rebate_max_us_dollars
    chp_state_cbi_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "No"
    chp_cbi_fed_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = "Yes"
    chp_cbi_sta_deprbas_fed_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "Utility  ($/W)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = chp.utility_rebate_us_dollars_per_kw * 0.001
    chp_utility_cbi_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = chp.utility_rebate_max_us_dollars
    chp_utility_cbi_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "No"
    chp_cbi_uti_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = "Yes"
    chp_cbi_uti_deprbas_fed_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=5)

    current_row += 1
    ws['A{}'.format(current_row)] = "Production based incentive (PBI)"
    ws['A{}'.format(current_row)].font = bold_font
    ws['B{}'.format(current_row)] = "Amount ($/kWh)"
    ws['B{}'.format(current_row)].border = attribute_border_left_and_right
    ws['C{}'.format(current_row)] = "Maximum ($/year)"
    ws['C{}'.format(current_row)].border = attribute_border_left_and_right
    ws['D{}'.format(current_row)] = "Federal Taxable"
    ws['D{}'.format(current_row)].border = attribute_border_left_and_right
    ws['E{}'.format(current_row)] = "Term (years)"
    ws['E{}'.format(current_row)].border = attribute_border_left_and_right
    ws['F{}'.format(current_row)] = "System Size Limit (kW)"
    ws['F{}'.format(current_row)].border = attribute_border_left_and_right
    make_attribute_row(ws, current_row, length=6, bold_cols=[0])

    current_row += 1
    ws['A{}'.format(current_row)] = "Combined ($/kWh)"
    ws['A{}'.format(current_row)].alignment = two_tab_indent
    ws['B{}'.format(current_row)] = chp.pbi_us_dollars_per_kwh
    chp_pbi_cell = "\'{}\'!B{}".format(inandout_sheet_name, current_row)
    ws['C{}'.format(current_row)] = chp.pbi_max_us_dollars
    chp_pbi_max_cell = "\'{}\'!C{}".format(inandout_sheet_name, current_row)
    ws['D{}'.format(current_row)] = "Yes"
    chp_pbi_combined_tax_fed_cell = "\'{}\'!D{}".format(inandout_sheet_name, current_row)
    ws['E{}'.format(current_row)] = chp.pbi_years
    chp_pbi_years_cell = "\'{}\'!E{}".format(inandout_sheet_name, current_row)
    ws['F{}'.format(current_row)] = chp.pbi_system_max_kw
    chp_pbi_max_kw_cell = "\'{}\'!F{}".format(inandout_sheet_name, current_row)
    make_attribute_row(ws, current_row, length=6)
    current_row += 1
    current_row += 1

    ####################################################################################################################
    # Depreciation
    ####################################################################################################################

    ws['A{}'.format(current_row)] = "DEPRECIATION"
    col_idx = 1
    for pv in pv_data:
        ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = "{}".format(pv['name'])
        col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = "BATTERY"
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = "WIND"
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = "CHP"
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = "Absorption Chiller"
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = "Hot TES"
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = "Cold TES"
    col_idx += 1
    make_title_row(ws, current_row, length=col_idx)
    col_idx += 1

    macrs_table_start_col = col_idx

    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = "MACRS SCHEDULES (INFORMATIONAL ONLY)"
    macrs_cells = {0: []}
    make_title_row(ws, current_row, length=9, offset=macrs_table_start_col)
    current_row += 1
    ws['A{}'.format(current_row)] = "Federal (years)"
    col_idx = 1
    for idx, pv in enumerate(pv_data):
        ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = pv['pv'].macrs_option_years
        pv_cell_locations[idx]["pv_macrs_option_cell"] = "\'{}\'!{}{}".format(
            inandout_sheet_name, upper_case_letters[col_idx], current_row)
        col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = batt.macrs_option_years
    batt_macrs_option_cell = "\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[col_idx], current_row)
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = wind.macrs_option_years
    wind_macrs_option_cell = "\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[col_idx], current_row)
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = chp.macrs_option_years
    chp_macrs_option_cell = "\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[col_idx], current_row)
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = absorption_chiller.macrs_option_years
    absorption_chiller_macrs_option_cell = "\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[col_idx], current_row)
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = hot_tes.macrs_option_years
    hot_tes_macrs_option_cell = "\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[col_idx], current_row)
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = cold_tes.macrs_option_years
    cold_tes_macrs_option_cell = "\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[col_idx], current_row)
    col_idx += 1
    make_attribute_row(ws, current_row, length=col_idx, alignment=center_align)
    ws['{}{}'.format(upper_case_letters[macrs_table_start_col], current_row)] = "Year"

    for i in range(8):
        ws['{}{}'.format(upper_case_letters[i + macrs_table_start_col + 1], current_row)] = i + 1
    make_attribute_row(ws, current_row, length=9, bold_cols=[macrs_table_start_col], offset=macrs_table_start_col)

    current_row += 1
    ws['A{}'.format(current_row)] = "Federal bonus fraction"
    col_idx = 1

    for i, pv in enumerate(pv_data):
        ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = pv['pv'].macrs_bonus_pct
        pv_cell_locations[i]["pv_bonus_fraction_cell"] = "\'{}\'!{}{}".format(
            inandout_sheet_name, upper_case_letters[col_idx], current_row)
        col_idx += 1

    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = batt.macrs_bonus_pct
    batt_bonus_fraction_cell = "\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[col_idx], current_row)
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = wind.macrs_bonus_pct
    wind_bonus_fraction_cell = "\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[col_idx], current_row)
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = chp.macrs_bonus_pct
    chp_bonus_fraction_cell = "\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[col_idx], current_row)
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = absorption_chiller.macrs_bonus_pct
    absorption_chiller_bonus_fraction_cell = "\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[col_idx], current_row)
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = hot_tes.macrs_bonus_pct
    hot_tes_bonus_fraction_cell = "\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[col_idx], current_row)
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[col_idx], current_row)] = cold_tes.macrs_bonus_pct
    cold_tes_bonus_fraction_cell = "\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[col_idx], current_row)
    col_idx += 1
    make_attribute_row(ws, current_row, length=col_idx, alignment=center_align)
    col_idx += 1
    ws['{}{}'.format(upper_case_letters[macrs_table_start_col], current_row)] = "5-Year"

    macrs_cells[5] = []

    for i in range(6):
        macrs_cells[5].append('{}{}'.format(upper_case_letters[i + macrs_table_start_col + 1], current_row))
        ws['{}{}'.format(upper_case_letters[i + macrs_table_start_col + 1], current_row)] = macrs_schedule[5][i]
    make_attribute_row(ws, current_row, length=9, bold_cols=[macrs_table_start_col], offset=macrs_table_start_col,
                       precision=4)
    current_row += 1

    ws['{}{}'.format(upper_case_letters[macrs_table_start_col], current_row)] = "7-Year"
    macrs_cells[7] = []

    for i in range(8):
        macrs_cells[7].append('{}{}'.format(upper_case_letters[macrs_table_start_col + i + 1], current_row))
        ws['{}{}'.format(upper_case_letters[macrs_table_start_col + i + 1], current_row)] = macrs_schedule[7][i]
    make_attribute_row(ws, current_row, length=9, bold_cols=[macrs_table_start_col], offset=macrs_table_start_col,
                       precision=4)
    current_row += 1
    current_row += 1


    ####################################################################################################################
    # Annual Value Summary
    ####################################################################################################################

    def fill_in_annual_values(row):
        make_attribute_row(ws, row, length=financial.analysis_years + 2, alignment=center_align,
                           number_format='#,##0')
        fill_cols(ws, range(2, financial.analysis_years + 2), row, calculated_fill)
        fill_cols(ws, range(1, 2), row, grey_fill)
        
    ws['A{}'.format(current_row)] = "ANNUAL VALUES"

    for i in range(financial.analysis_years + 1):
        ws['{}{}'.format(upper_case_letters[i + 1], current_row)] = i
    make_title_row(ws, current_row, length=financial.analysis_years+2)
    current_row += 1

    for idx, pv in enumerate(pv_data):
        ws['A{}'.format(current_row)] = "{} Annual electricity (kWh)".format(pv['name'])
        ws['B{}'.format(current_row)] = 0
        pv_cell_locations[idx]["pv_production_series"] = list()

        for year in range(1, financial.analysis_years + 1):
            ws['{}{}'.format(upper_case_letters[year+1], current_row)] = \
                '={pv_energy} * (1 - {pv_degradation_rate}/100)^{year}'.format(
                    pv_energy=pv_cell_locations[idx]["pv_energy_cell"],
                    pv_degradation_rate=pv_cell_locations[idx]["pv_degradation_rate_cell"],
                    year=year-1,
                )
            pv_cell_locations[idx]["pv_production_series"].append("\'{}\'!{}{}".format(
                inandout_sheet_name, upper_case_letters[year+1], current_row))
            
        fill_in_annual_values(current_row)
        current_row += 1

    for idx, pv in enumerate(pv_data):
        ws['A{}'.format(current_row)] = "Existing {} Annual electricity (kWh)".format(pv['name'])
        ws['B{}'.format(current_row)] = 0
        pv_cell_locations[idx]["pv_production_series_bau"] = list()

        for year in range(1, financial.analysis_years + 1):
            ws['{}{}'.format(upper_case_letters[year+1], current_row)] = \
                '={pv_energy} * (1 - {pv_degradation_rate}/100)^{year}'.format(
                    pv_energy=pv_cell_locations[idx]["pv_energy_bau_cell"],
                    pv_degradation_rate=pv_cell_locations[idx]["pv_degradation_rate_cell"],
                    year=year-1,
                )
            pv_cell_locations[idx]["pv_production_series_bau"].append("\'{}\'!{}{}".format(
                inandout_sheet_name, upper_case_letters[year+1], current_row))

        fill_in_annual_values(current_row)
        current_row += 1

    ws['A{}'.format(current_row)] = "Wind Annual electricity (kWh)"
    ws['B{}'.format(current_row)] = 0
    wind_annual_kwh_cells = ["\'{}\'!C{}".format(inandout_sheet_name, current_row)]

    for year in range(1, financial.analysis_years + 1):
        ws['{}{}'.format(upper_case_letters[year+1], current_row)] = '={}'.format(wind_energy_cell)
        wind_annual_kwh_cells.append("\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[year+1],
                                                           current_row))
    fill_in_annual_values(current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "Backup Generator Annual electricity (kWh)"
    ws['B{}'.format(current_row)] = 0
    generator_production_series = ["\'{}\'!C{}".format(inandout_sheet_name, current_row)]

    for year in range(1, financial.analysis_years + 1):
        ws['{}{}'.format(upper_case_letters[year+1], current_row)] = '={}'.format(generator_energy_cell)
        generator_production_series.append("\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[year+1],
                                                                current_row))
    fill_in_annual_values(current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "CHP Annual electricity (kWh)"
    ws['B{}'.format(current_row)] = 0
    chp_annual_kwh_cells = ["\'{}\'!C{}".format(inandout_sheet_name, current_row)]

    for year in range(1, financial.analysis_years + 1):
        ws['{}{}'.format(upper_case_letters[year+1], current_row)] = '={}'.format(chp_electric_energy_cell)
        chp_annual_kwh_cells.append("\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[year+1],
                                                                current_row))
    fill_in_annual_values(current_row)

    current_row += 1
    ws['A{}'.format(current_row)] = "CHP Annual heat (MMBtu)"
    ws['B{}'.format(current_row)] = 0
    chp_thermal_production_series = ["\'{}\'!C{}".format(inandout_sheet_name, current_row)]

    for year in range(1, financial.analysis_years + 1):
        ws['{}{}'.format(upper_case_letters[year+1], current_row)] = '={}'.format(chp_thermal_energy_cell)
        chp_thermal_production_series.append("\'{}\'!{}{}".format(inandout_sheet_name, upper_case_letters[year+1],
                                                                current_row))
    fill_in_annual_values(current_row)
    
    current_row += 1
    for idx, pv in enumerate(pv_data):
        ws['A{}'.format(current_row)] = "{} Federal depreciation percentages (fraction)".format(pv['name'])
        ws['B{}'.format(current_row)] = 0
        pv_cell_locations[idx]["pv_macrs_cells"] = []
        for ii, c in enumerate(macrs_cells[pv['pv'].macrs_option_years]):
            ws[upper_case_letters[2 + ii] + str(current_row)] = '=' + c
            pv_cell_locations[idx]["pv_macrs_cells"].append("\'{}\'!".format(
                inandout_sheet_name) + upper_case_letters[2 + ii] + str(current_row))
        
        start_number = pv['pv'].macrs_option_years
        if start_number == 0:
            start_number = -1
        for ii in range(start_number + 1, financial.analysis_years):
            ws[upper_case_letters[2 + ii] + str(current_row)] = 0
            pv_cell_locations[idx]["pv_macrs_cells"].append("\'{}\'!".format(
                inandout_sheet_name) + upper_case_letters[2 + ii] + str(current_row))
        make_attribute_row(ws, current_row, length=financial.analysis_years+2, alignment=center_align, number_format='0.0')
        fill_cols(ws, range(2, financial.analysis_years + 2), current_row, calculated_fill)
        fill_cols(ws, range(1, 2), current_row, grey_fill)
        current_row += 1

    ws['A{}'.format(current_row)] = "Wind Federal depreciation percentages (fraction)"
    ws['B{}'.format(current_row)] = 0
    wind_macrs_cells = []
    for i, c in enumerate(macrs_cells[wind.macrs_option_years]):
        ws[upper_case_letters[2 + i] + str(current_row)] = '=' + c
        wind_macrs_cells.append("\'{}\'!".format(inandout_sheet_name) + upper_case_letters[2 + i] + str(current_row))
    
    start_number = wind.macrs_option_years
    if start_number == 0:
        start_number = -1
    for i in range(start_number+ 1, financial.analysis_years):
        ws[upper_case_letters[2 + i] + str(current_row)] = 0
        wind_macrs_cells.append("\'{}\'!".format(inandout_sheet_name) + upper_case_letters[2 + i] + str(current_row))
    make_attribute_row(ws, current_row, length=financial.analysis_years+2, alignment=center_align, number_format='0.0')
    fill_cols(ws, range(2, financial.analysis_years + 2), current_row, calculated_fill)
    fill_cols(ws, range(1, 2), current_row, grey_fill)

    current_row += 1
    ws['A{}'.format(current_row)] = "Battery Federal depreciation percentages (fraction)"
    ws['B{}'.format(current_row)] = 0
    batt_macrs_cells = []

    for i, c in enumerate(macrs_cells[batt.macrs_option_years]):
        ws[upper_case_letters[2 + i] + str(current_row)] = '=' + c
        batt_macrs_cells.append("\'{}\'!".format(inandout_sheet_name) + upper_case_letters[2 + i] + str(current_row))

    start_number = batt.macrs_option_years
    if start_number == 0:
        start_number = -1
    for i in range(start_number+ 1, financial.analysis_years):
        ws[upper_case_letters[2 + i] + str(current_row)] = 0
        batt_macrs_cells.append("\'{}\'!".format(inandout_sheet_name) + upper_case_letters[2 + i] + str(current_row))
    make_attribute_row(ws, current_row, length=financial.analysis_years+2, alignment=center_align, number_format='0.0')
    fill_cols(ws, range(2, financial.analysis_years + 2), current_row, calculated_fill)
    fill_cols(ws, range(1, 2), current_row, grey_fill)

    current_row += 1
    ws['A{}'.format(current_row)] = "CHP Federal depreciation percentages (fraction)"
    ws['B{}'.format(current_row)] = 0
    chp_macrs_cells = []

    for i, c in enumerate(macrs_cells[chp.macrs_option_years]):
        ws[upper_case_letters[2 + i] + str(current_row)] = '=' + c
        chp_macrs_cells.append("\'{}\'!".format(inandout_sheet_name) + upper_case_letters[2 + i] + str(current_row))

    start_number = chp.macrs_option_years
    if start_number == 0:
        start_number = -1
    for i in range(start_number+ 1, financial.analysis_years):
        ws[upper_case_letters[2 + i] + str(current_row)] = 0
        chp_macrs_cells.append("\'{}\'!".format(inandout_sheet_name) + upper_case_letters[2 + i] + str(current_row))
    make_attribute_row(ws, current_row, length=financial.analysis_years+2, alignment=center_align, number_format='0.0')
    fill_cols(ws, range(2, financial.analysis_years + 2), current_row, calculated_fill)
    fill_cols(ws, range(1, 2), current_row, grey_fill)

    current_row += 1
    ws['A{}'.format(current_row)] = "Absorption Chiller Federal depreciation percentages (fraction)"
    ws['B{}'.format(current_row)] = 0
    absorption_chiller_macrs_cells = []

    for i, c in enumerate(macrs_cells[absorption_chiller.macrs_option_years]):
        ws[upper_case_letters[2 + i] + str(current_row)] = '=' + c
        absorption_chiller_macrs_cells.append("\'{}\'!".format(inandout_sheet_name) + upper_case_letters[2 + i] + str(current_row))

    start_number = absorption_chiller.macrs_option_years
    if start_number == 0:
        start_number = -1
    for i in range(start_number+ 1, financial.analysis_years):
        ws[upper_case_letters[2 + i] + str(current_row)] = 0
        absorption_chiller_macrs_cells.append("\'{}\'!".format(inandout_sheet_name) + upper_case_letters[2 + i] + str(current_row))
    make_attribute_row(ws, current_row, length=financial.analysis_years+2, alignment=center_align, number_format='0.0')
    fill_cols(ws, range(2, financial.analysis_years + 2), current_row, calculated_fill)
    fill_cols(ws, range(1, 2), current_row, grey_fill)

    current_row += 1
    ws['A{}'.format(current_row)] = "Hot TES Federal depreciation percentages (fraction)"
    ws['B{}'.format(current_row)] = 0
    hot_tes_macrs_cells = []

    for i, c in enumerate(macrs_cells[hot_tes.macrs_option_years]):
        ws[upper_case_letters[2 + i] + str(current_row)] = '=' + c
        hot_tes_macrs_cells.append("\'{}\'!".format(inandout_sheet_name) + upper_case_letters[2 + i] + str(current_row))

    start_number = hot_tes.macrs_option_years
    if start_number == 0:
        start_number = -1
    for i in range(start_number+ 1, financial.analysis_years):
        ws[upper_case_letters[2 + i] + str(current_row)] = 0
        hot_tes_macrs_cells.append("\'{}\'!".format(inandout_sheet_name) + upper_case_letters[2 + i] + str(current_row))
    make_attribute_row(ws, current_row, length=financial.analysis_years+2, alignment=center_align, number_format='0.0')
    fill_cols(ws, range(2, financial.analysis_years + 2), current_row, calculated_fill)
    fill_cols(ws, range(1, 2), current_row, grey_fill)

    current_row += 1
    ws['A{}'.format(current_row)] = "Cold TES Federal depreciation percentages (fraction)"
    ws['B{}'.format(current_row)] = 0
    cold_tes_macrs_cells = []

    for i, c in enumerate(macrs_cells[cold_tes.macrs_option_years]):
        ws[upper_case_letters[2 + i] + str(current_row)] = '=' + c
        cold_tes_macrs_cells.append("\'{}\'!".format(inandout_sheet_name) + upper_case_letters[2 + i] + str(current_row))

    start_number = cold_tes.macrs_option_years
    if start_number == 0:
        start_number = -1
    for i in range(start_number+ 1, financial.analysis_years):
        ws[upper_case_letters[2 + i] + str(current_row)] = 0
        cold_tes_macrs_cells.append("\'{}\'!".format(inandout_sheet_name) + upper_case_letters[2 + i] + str(current_row))
    make_attribute_row(ws, current_row, length=financial.analysis_years+2, alignment=center_align, number_format='0.0')
    fill_cols(ws, range(2, financial.analysis_years + 2), current_row, calculated_fill)
    fill_cols(ws, range(1, 2), current_row, grey_fill)

    free_cash_flow_row = current_row + 1


    ####################################################################################################################
    ####################################################################################################################
    ####################################################################################################################
    ####################################################################################################################
    # OPTIMAL / DEVELOPER CASH FLOW SHEET
    ####################################################################################################################
    ####################################################################################################################
    ####################################################################################################################
    ####################################################################################################################

    dcs = wb.get_sheet_by_name(developer_cashflow_sheet_name)
    current_row = 2

    ####################################################################################################################
    # Operating Year
    ####################################################################################################################

    dcs['A{}'.format(current_row)] = "Operating Year"
    for year in range(financial.analysis_years + 1):
        dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = year
    make_title_row(dcs, current_row, length=financial.analysis_years+2)
    current_row += 1
    current_row += 1
    
    ####################################################################################################################
    # Operating Expenses
    ####################################################################################################################

    dcs['A{}'.format(current_row)] = "Operating Expenses"
    make_title_row(dcs, current_row, length=financial.analysis_years+2)

    current_row += 1
    
    start_om_row = current_row
    if not financial.two_party_ownership:  # Bill and export credit move to Host cashflow for two party
        
        dcs['A{}'.format(current_row)] = "Electricity bill with system before export credits"

        for year in range(1, financial.analysis_years + 1):
            dcs['{}{}'.format(upper_case_letters[year+1], current_row)] = \
                '=-{year_one_bill} * (1 + {escalation_pct}/100)^{year}'.format(
                    year_one_bill=year_one_bill_cell, year=year, escalation_pct=escalation_pct_cell
                )
        make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        current_row += 1

        dcs['A{}'.format(current_row)] = "Export credits with system"

        for year in range(1, financial.analysis_years + 1):
            dcs['{}{}'.format(upper_case_letters[year+1], current_row)] = \
                '={year_one_credits} * (1 + {escalation_pct}/100)^{year}'.format(
                    year_one_credits=year_one_credits_cell, year=year, escalation_pct=escalation_pct_cell,
                    pv_degradation_rate=pv_cell_locations[0]["pv_degradation_rate_cell"])
        make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        current_row += 1

        dcs['A{}'.format(current_row)] = "Boiler fuel bill with system"
        for year in range(1, financial.analysis_years + 1):
            dcs['{}{}'.format(upper_case_letters[year+1], current_row)] = \
                '=(-1 * ({boiler_fuel_bill_cell} * (1 + {escalation_pct}/100)^{year}))'.format(
                    boiler_fuel_bill_cell=boiler_fuel_bill_cell, year=year, escalation_pct=boiler_escalation_pct_cell)
        make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        current_row += 1

        dcs['A{}'.format(current_row)] = "CHP fuel bill"
        for year in range(1, financial.analysis_years + 1):
            dcs['{}{}'.format(upper_case_letters[year+1], current_row)] = \
                '=(-1 * ({chp_fuel_bill_cell} * (1 + {escalation_pct}/100)^{year}))'.format(
                    chp_fuel_bill_cell=chp_fuel_bill_cell, year=year, escalation_pct=chp_escalation_pct_cell)
        make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        current_row += 1

    dcs['A{}'.format(current_row)] = "Operation and Maintenance (O&M)"
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)

    current_row += 1
    for i, pv in enumerate(pv_data):
        dcs['A{}'.format(current_row)] = "New {} fixed O&M cost".format(pv['name'])
        dcs['A{}'.format(current_row)].alignment = one_tab_indent
        pv_cell_locations[i]['new_o_and_m_cashflow_cells'] = []
        for year in range(1, financial.analysis_years + 1):
            pv_cell_locations[i]['new_o_and_m_cashflow_cells'].append('\'{}\'!{}{}'.format(
                developer_cashflow_sheet_name, upper_case_letters[year + 1], current_row))
            dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = (
                '=-{pv_om_cost_us_dollars_per_kw_cell} * (1 + {om_escalation_rate_cell}/100)^{year}'
                ' * ({pv_size_kw_cell})'
            ).format(
            pv_om_cost_us_dollars_per_kw_cell=pv_cell_locations[i]["pv_om_cost_us_dollars_per_kw_cell"],
            om_escalation_rate_cell=om_escalation_rate_cell,
            year=year,
            pv_size_kw_cell=pv_cell_locations[i]["pv_size_kw_cell"]
            )
        make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        current_row += 1
        dcs['A{}'.format(current_row)] = "Existing {} fixed O&M cost".format(pv['name'])
        dcs['A{}'.format(current_row)].alignment = one_tab_indent
        pv_cell_locations[i]['existing_o_and_m_cashflow_cells'] = []
        for year in range(1, financial.analysis_years + 1):
            pv_cell_locations[i]['existing_o_and_m_cashflow_cells'].append('\'{}\'!{}{}'.format(
                developer_cashflow_sheet_name, upper_case_letters[year + 1], current_row))
            dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = (
                '=-{pv_om_cost_us_dollars_per_kw_cell} * (1 + {om_escalation_rate_cell}/100)^{year}'
                ' * ({pv_existing_kw_cell})'
            ).format(
            pv_om_cost_us_dollars_per_kw_cell=pv_cell_locations[i]["pv_om_cost_us_dollars_per_kw_cell"],
            om_escalation_rate_cell=om_escalation_rate_cell,
            year=year,
            pv_existing_kw_cell=pv_cell_locations[i]["pv_existing_kw_cell"] if not financial.two_party_ownership else 0,
            )
        make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        current_row += 1

    dcs['A{}'.format(current_row)] = "Wind fixed O&M cost"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    wind_o_and_m_cashflow_cells = []
    for year in range(1, financial.analysis_years + 1):
        wind_o_and_m_cashflow_cells.append('\'{}\'!{}{}'.format(
            developer_cashflow_sheet_name, upper_case_letters[year + 1], current_row))
        dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = '=-{} * (1 + {}/100)^{} * {}'.format(
            wind_om_cost_us_dollars_per_kw_cell, om_escalation_rate_cell, year, wind_size_kw_cell)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Generator fixed O&M cost"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    for year in range(1, financial.analysis_years + 1):
        if gen_fuel_used_gal == 0 and gen_fuel_used_gal_bau == 0:
            # generator was not modeled because it could not run (eg. no outage, can't be used with grid)
            dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = 0
        else:
            dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = (
                '=-{generator_om_cost} * (1 + {om_escalation_rate}/100)^{year} * ({gen_kw} + {existing_gen_kw})'
                ).format(
                generator_om_cost=generator_om_cost_us_dollars_per_kw_cell,
                om_escalation_rate=om_escalation_rate_cell,
                year=year,
                gen_kw=generator_size_kw_cell,
                existing_gen_kw=generator_existing_kw_cell if not financial.two_party_ownership else 0,
            )
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Generator variable O&M cost"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    for year in range(1, financial.analysis_years + 1):
        if gen_fuel_used_gal == 0 and gen_fuel_used_gal_bau == 0:
            # generator was not modeled because it could not run (eg. no outage, can't be used with grid)
            dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = 0
        else:
            dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = '=-{} * (1 + {}/100)^{} * {}'.format(
                generator_om_cost_us_dollars_per_kwh_cell, om_escalation_rate_cell, year,
                generator_energy_cell)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)

    if not financial.two_party_ownership:  # otherwise on Host's cash flow
        current_row += 1
        dcs['A{}'.format(current_row)] = "Generator diesel fuel cost ($)"
        dcs['A{}'.format(current_row)].alignment = one_tab_indent
        for year in range(1, financial.analysis_years + 1):
            dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = '=-{} * (1+{}/100)^{}'.format(
                diesel_fuel_used_cost_cell, om_escalation_rate_cell, year)
        make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                           number_format='#,##0', border=no_border)
    current_row += 1
    dcs['A{}'.format(current_row)] = "Battery kW replacement cost "
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    for i in range(1, financial.analysis_years + 1):
        dcs['{}{}'.format(upper_case_letters[i + 1], current_row)] = '=-IF({} = {}, {} * {}, 0)'.format(
            i, batt_replace_year_cell, batt_size_kw_cell, batt_replace_cost_us_dollars_per_kw_cell)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    current_row += 1
    dcs['A{}'.format(current_row)] = "Battery kWh replacement cost "
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    for i in range(1, financial.analysis_years + 1):
        dcs['{}{}'.format(upper_case_letters[i + 1], current_row)] = '=-IF({} = {}, {} * {}, 0)'.format(
            i, batt_replace_year_cell, batt_size_kwh_cell, batt_replace_cost_us_dollars_per_kwh_cell)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "CHP fixed O&M cost"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    for year in range(1, financial.analysis_years + 1):
        dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = '=-{} * {} * (1+{}/100)^{}'.format(
                chp_om_cost_us_dollars_per_kw_cell, chp_size_kw_cell, om_escalation_rate_cell, year)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "CHP variable generation O&M cost"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    for year in range(1, financial.analysis_years + 1):
        dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = '=-{} * {} * (1+{}/100)^{}'.format(
                chp_om_cost_us_dollars_per_kwh_cell, chp_electric_energy_cell, om_escalation_rate_cell, year)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "CHP runtime O&M cost"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    for year in range(1, financial.analysis_years + 1):
        dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = '=-{} * {} * {} * (1+{}/100)^{}'.format(
            chp_om_cost_us_dollars_per_kw_rated__cell, chp_size_kw_cell, chp_runtime_cell, om_escalation_rate_cell, year)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Absorption Chiller fixed O&M cost"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    for year in range(1, financial.analysis_years + 1):
        dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = '=-{} * {} * (1+{}/100)^{}'.format(
            absorption_chiller_om_cost_us_dollars_per_ton_cell, absorption_chiller_size_ton_cell, om_escalation_rate_cell, year)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Chilled water TES fixed O&M cost"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    for year in range(1, financial.analysis_years + 1):
        dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = '=-{} * {} * (1+{}/100)^{}'.format(
            cold_tes_om_cost_us_dollars_per_gal_cell, cold_tes_size_gal_cell, om_escalation_rate_cell, year)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Hot water TES fixed O&M cost"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    for year in range(1, financial.analysis_years + 1):
        dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = '=-{} * {} * (1+{}/100)^{}'.format(
            hot_tes_om_cost_us_dollars_per_gal_cell, hot_tes_size_gal_cell, om_escalation_rate_cell, year)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Total operating expenses"
    for i in range(1, financial.analysis_years + 1):
        dcs['{}{}'.format(upper_case_letters[i + 1], current_row)] = '=SUM({col}{start_row}:{col}{end_row})'.format(
            col=upper_case_letters[i + 1], start_row=start_om_row, end_row=current_row - 1)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    fill_border(dcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)
    opex_total_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Tax deductible operating expenses"

    for year in range(1, financial.analysis_years + 1):
        dcs['{}{}'.format(upper_case_letters[year+1], current_row)] = (
            "=IF({fed_tax_rate} > 0, {col}{opex_total_row}, 0)"
        ).format(
            fed_tax_rate=fed_tax_rate_cell,
            opex_total_row=opex_total_row,
            col=upper_case_letters[year+1]
        )
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    fill_border(dcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)
    opex_tax_deductible_row = current_row
    
    current_row += 1
    current_row += 1

    ####################################################################################################################
    # Incentives
    ####################################################################################################################
    ####################################################################################################################

    dcs['A{}'.format(current_row)] = "Direct Cash Incentives"
    make_title_row(dcs, current_row, length=financial.analysis_years+2)
    current_row += 1
    ibi_and_cbi_totals_rows = []

    ####################################################################################################################
    # PV
    ####################################################################################################################

    for idx, pv in enumerate(pv_data):
        dcs['A{}'.format(current_row)] = "{} Investment-based incentives (IBI)".format(pv['name'])
        make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)

        current_row += 1
        dcs['A{}'.format(current_row)] = "State IBI"
        dcs['A{}'.format(current_row)].alignment = one_tab_indent
        dcs['B{}'.format(current_row)] = "=MIN(({}/100)*({}-B{}-B{}),{})".format(
            pv_cell_locations[idx]["pv_state_ibi_pct_cell"], pv_cell_locations[idx]["pv_cost_cell"], current_row + 1,
                                                                                                     current_row + 6,
            pv_cell_locations[idx]["pv_state_ibi_max_cell"])
        make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
        pv_cell_locations[idx]["pv_state_ibi_row"] = current_row

        current_row += 1
        dcs['A{}'.format(current_row)] = "Utility IBI"
        dcs['A{}'.format(current_row)].alignment = one_tab_indent
        dcs['B{}'.format(current_row)] = "=MIN(({}/100)*{},{})".format(
            pv_cell_locations[idx]["pv_utility_ibi_pct_cell"], pv_cell_locations[idx]["pv_cost_cell"],
            pv_cell_locations[idx]["pv_utility_ibi_max_cell"])
        make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
        pv_cell_locations[idx]["pv_utility_ibi_row"] = current_row

        current_row += 1
        dcs['A{}'.format(current_row)] = "Total"
        dcs['A{}'.format(current_row)].alignment = one_tab_indent
        dcs['B{}'.format(current_row)] = "=SUM(B{},B{})".format(
            pv_cell_locations[idx]["pv_state_ibi_row"], pv_cell_locations[idx]["pv_utility_ibi_row"])
        make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
        pv_cell_locations[idx]["pv_total_ibi_cell"] = "\'{}\'!B{}".format(developer_cashflow_sheet_name, current_row)
        ibi_and_cbi_totals_rows.append(current_row)
     

        current_row += 1
        dcs['A{}'.format(current_row)] = "{} Capacity-based incentives (CBI)".format(pv['name'])
        make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)

        current_row += 1
        dcs['A{}'.format(current_row)] = "Federal CBI"
        dcs['A{}'.format(current_row)].alignment = one_tab_indent
        dcs['B{}'.format(current_row)] = "=MIN({}*{}*1000,{})".format(
            pv_cell_locations[idx]["pv_federal_cbi_cell"], pv_cell_locations[idx]["pv_size_kw_cell"],
            pv_cell_locations[idx]["pv_federal_cbi_max_cell"])
        make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
        pv_cell_locations[idx]["pv_federal_cbi_row"] = current_row

        current_row += 1
        dcs['A{}'.format(current_row)] = "State CBI"
        dcs['A{}'.format(current_row)].alignment = one_tab_indent
        dcs['B{}'.format(current_row)] = "=MIN({}*{}*1000,{})".format(
            pv_cell_locations[idx]["pv_state_cbi_cell"], pv_cell_locations[idx]["pv_size_kw_cell"],
            pv_cell_locations[idx]["pv_state_cbi_max_cell"])
        make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
        pv_cell_locations[idx]["pv_state_cbi_row"] = current_row

        current_row += 1
        dcs['A{}'.format(current_row)] = "Utility CBI"
        dcs['A{}'.format(current_row)].alignment = one_tab_indent
        dcs['B{}'.format(current_row)] = "=MIN({}*{}*1000,{})".format(
            pv_cell_locations[idx]["pv_utility_cbi_cell"], pv_cell_locations[idx]["pv_size_kw_cell"],
            pv_cell_locations[idx]["pv_utility_cbi_max_cell"])
        make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
        pv_cell_locations[idx]["pv_utility_cbi_row"] = current_row
        current_row += 1
        dcs['A{}'.format(current_row)] = "Total"
        dcs['A{}'.format(current_row)].alignment = one_tab_indent
        dcs['B{}'.format(current_row)] = "=SUM(B{},B{},B{})".format(
            pv_cell_locations[idx]["pv_federal_cbi_row"], pv_cell_locations[idx]["pv_state_cbi_row"],
            pv_cell_locations[idx]["pv_utility_cbi_row"])
        pv_cell_locations[idx]["pv_total_cbi_cell"] = "\'{}\'!B{}".format(developer_cashflow_sheet_name, current_row)
        make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
        ibi_and_cbi_totals_rows.append(current_row)
        current_row += 1

    ####################################################################################################################
    # Wind
    ####################################################################################################################

    dcs['A{}'.format(current_row)] = "Wind Investment-based incentives (IBI)"
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    current_row += 1
    dcs['A{}'.format(current_row)] = "State IBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=MIN(({}/100)*({}-B{}-B{}),{})".format(
        wind_state_ibi_cell, wind_cost_cell, current_row + 1, current_row + 6, wind_state_ibi_max_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    wind_state_ibi_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Utility IBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=MIN(({}/100)*{},{})".format(
        wind_utility_ibi_cell, wind_cost_cell, wind_utility_ibi_max_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    wind_utility_ibi_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Total"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=SUM(B{},B{})".format(wind_state_ibi_row, wind_utility_ibi_row)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    wind_total_ibi_cell = "\'{}\'!B{}".format(developer_cashflow_sheet_name, current_row) 
    ibi_and_cbi_totals_rows.append(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Wind Capacity-based incentives (CBI)"
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Federal CBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=MIN({}*{}*1000,{})".format(
        wind_federal_cbi_cell, wind_size_kw_cell, wind_federal_cbi_max_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    wind_federal_cbi_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "State CBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=MIN({}*{}*1000,{})".format(
        wind_state_cbi_cell, wind_size_kw_cell, wind_state_cbi_max_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    wind_state_cbi_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Utility CBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=MIN({}*{}*1000,{})".format(
        wind_utility_cbi_cell, wind_size_kw_cell, wind_utility_cbi_max_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    wind_utility_cbi_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Total"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=SUM(B{},B{},B{})".format(wind_federal_cbi_row, wind_state_cbi_row,
                                                                wind_utility_cbi_row)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    wind_total_cbi_cell = "\'{}\'!B{}".format(developer_cashflow_sheet_name, current_row)
    ibi_and_cbi_totals_rows.append(current_row)

    ####################################################################################################################
    # Battery
    ####################################################################################################################

    current_row += 1
    dcs['A{}'.format(current_row)] = "Battery Investment-based incentives (IBI)"
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "State IBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=MIN(({}/100)*({}),{})".format(batt_state_ibi_cell, batt_cost_cell,
                                                                     batt_state_ibi_max_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    batt_state_ibi_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Utility IBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=MIN(({}/100)*{},{})".format(batt_utility_ibi_cell, batt_cost_cell,
                                                                   batt_utility_ibi_max_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    batt_utility_ibi_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Total"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=SUM(B{},B{})".format(batt_state_ibi_row, batt_utility_ibi_row)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    batt_total_ibi_cell = "\'{}\'!B{}".format(developer_cashflow_sheet_name, current_row) 
    ibi_and_cbi_totals_rows.append(current_row)
    current_row += 1

    dcs['A{}'.format(current_row)] = "Battery Capacity-based incentives (CBI)"
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Total Power (per kW) CBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=MIN({}*{}*1000,{})".format(
        batt_total_cbi_per_kw_cell, batt_size_kw_cell, batt_total_cbi_per_kw_max_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    batt_cbi_per_kw_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Total Storage Capacity (per kWh) CBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=MIN({}*{}*1000,{})".format(
        batt_total_cbi_per_kwh_cell, batt_size_kwh_cell, batt_total_cbi_per_kwh_max_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    batt_cbi_per_kwh_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Total"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=SUM(B{},B{})".format(batt_cbi_per_kw_row, batt_cbi_per_kwh_row)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    batt_total_cbi_cell = 'B{}'.format(current_row)
    ibi_and_cbi_totals_rows.append(current_row)
    current_row += 1
    
    ####################################################################################################################
    # CHP
    ####################################################################################################################

    dcs['A{}'.format(current_row)] = "CHP Investment-based incentives (IBI)"
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    current_row += 1
    dcs['A{}'.format(current_row)] = "State IBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=MIN(({}/100)*({}-B{}-B{}),{})".format(
        chp_state_ibi_cell, chp_cost_cell, current_row + 1, current_row + 6, chp_state_ibi_max_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    chp_state_ibi_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Utility IBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=MIN(({}/100)*{},{})".format(
        chp_utility_ibi_cell, chp_cost_cell, chp_utility_ibi_max_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    chp_utility_ibi_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Total"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=SUM(B{},B{})".format(chp_state_ibi_row, chp_utility_ibi_row)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    chp_total_ibi_cell = "\'{}\'!B{}".format(developer_cashflow_sheet_name, current_row) 
    ibi_and_cbi_totals_rows.append(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "CHP Capacity-based incentives (CBI)"
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Federal CBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=MIN({}*{}*1000,{})".format(
        chp_federal_cbi_cell, chp_size_kw_cell, chp_federal_cbi_max_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    chp_federal_cbi_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "State CBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=MIN({}*{}*1000,{})".format(
        chp_state_cbi_cell, chp_size_kw_cell, chp_state_cbi_max_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    chp_state_cbi_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Utility CBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=MIN({}*{}*1000,{})".format(
        chp_utility_cbi_cell, chp_size_kw_cell, chp_utility_cbi_max_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    chp_utility_cbi_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Total"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    dcs['B{}'.format(current_row)] = "=SUM(B{},B{},B{})".format(chp_federal_cbi_row, chp_state_cbi_row,
                                                                chp_utility_cbi_row)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    chp_total_cbi_cell = "\'{}\'!B{}".format(developer_cashflow_sheet_name, current_row)
    ibi_and_cbi_totals_rows.append(current_row)
    current_row += 1
    current_row += 1


    dcs['A{}'.format(current_row)] = "Total CBI and IBI"
    dcs['B{}'.format(current_row)] = '=SUM(' + ','.join(['B{}'.format(r) for r in ibi_and_cbi_totals_rows]) + ')'
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    current_row += 1
    current_row += 1

    

    ####################################################################################################################
    # PBI
    ####################################################################################################################

    dcs['A{}'.format(current_row)] = "Production-based incentives (PBI)"
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)

    current_row += 1
    start_pbi_total_row = current_row
    for idx, pv in enumerate(pv_data):
        dcs['A{}'.format(current_row)] = "New {} Combined PBI".format(pv['name'])
        dcs['A{}'.format(current_row)].alignment = one_tab_indent
        pv_cell_locations[idx]["new_pv_pbi_series"] = []
        for year in range(financial.analysis_years):
            pv_cell_locations[idx]["new_pv_pbi_series"].append("\'{}\'!{}{}".format(
                developer_cashflow_sheet_name, upper_case_letters[year + 2], current_row))
            dcs['{}{}'.format(upper_case_letters[year + 2], current_row)] = (
                "=IF({year} < {pbi_year_limit}, "
                "MIN({dol_per_kwh} * ({pv_kwh} - {existing_pv_kwh}), "
                "{pbi_max} * (1 - {pv_degradation_rate}/100)^{year}), 0)"
                ).format(
                year=year, 
                pbi_year_limit=pv_cell_locations[idx]["pv_pbi_years_cell"], 
                dol_per_kwh=pv_cell_locations[idx]["pv_pbi_cell"],
                col=upper_case_letters[year + 2],
                pv_kwh=pv_cell_locations[idx]['pv_production_series'][year],
                existing_pv_kwh=(pv_cell_locations[idx]['pv_production_series_bau'][year]),
                pbi_max=pv_cell_locations[idx]["pv_pbi_max_cell"],
                pv_degradation_rate=pv_cell_locations[idx]["pv_degradation_rate_cell"],
            )
        make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        pv_cell_locations[idx]["new_pv_pbi_total_row"] = current_row
        current_row += 1

        dcs['A{}'.format(current_row)] = "Existing {} Combined PBI".format(pv['name'])
        dcs['A{}'.format(current_row)].alignment = one_tab_indent
        pv_cell_locations[idx]["existing_pv_pbi_series"] = []
        for year in range(financial.analysis_years):
            pv_cell_locations[idx]["existing_pv_pbi_series"].append("\'{}\'!{}{}".format(developer_cashflow_sheet_name, upper_case_letters[year + 2], current_row))
            dcs['{}{}'.format(upper_case_letters[year + 2], current_row)] = (
                "=IF({year} < {pbi_year_limit}, "
                "MIN({dol_per_kwh} * ({existing_pv_kwh}), "
                "{pbi_max} * (1 - {pv_degradation_rate}/100)^{year}), 0)"
                ).format(
                year=year, 
                pbi_year_limit=pv_cell_locations[idx]["pv_pbi_years_cell"], 
                dol_per_kwh=pv_cell_locations[idx]["pv_pbi_cell"],
                col=upper_case_letters[year + 2],
                existing_pv_kwh=(pv_cell_locations[idx]['pv_production_series_bau'][year]) if not financial.two_party_ownership else 0,
                pbi_max=pv_cell_locations[idx]["pv_pbi_max_cell"],
                pv_degradation_rate=pv_cell_locations[idx]["pv_degradation_rate_cell"],
            )
        make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        pv_cell_locations[idx]["existing_pv_pbi_total_row"] = current_row
        current_row += 1

    end_pbi_total_row = current_row
    dcs['A{}'.format(current_row)] = "Wind Combined PBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    wind_pbi_series = []
    for year in range(financial.analysis_years):
        wind_pbi_series.append("\'{}\'!{}{}".format(
            developer_cashflow_sheet_name, upper_case_letters[year + 2], current_row))
        dcs['{}{}'.format(upper_case_letters[year+2], current_row)] = (
                "=IF({year} < {pbi_year_limit}, "
                "MIN({dol_per_kwh} * {wind_kwh}, {pbi_max}), 0)"
                ).format(
                year=year,
                pbi_year_limit=wind_pbi_years_cell,
                dol_per_kwh=wind_pbi_cell,
                col=upper_case_letters[year+2],
                wind_kwh=wind_annual_kwh_cells[year+1],
                pbi_max=wind_pbi_max_cell,
            )
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    wind_pbi_total_row = current_row

    current_row += 1
    end_pbi_total_row = current_row
    dcs['A{}'.format(current_row)] = "CHP Combined PBI"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    chp_pbi_series = []
    for year in range(financial.analysis_years):
        chp_pbi_series.append("\'{}\'!{}{}".format(
            developer_cashflow_sheet_name, upper_case_letters[year + 2], current_row))
        dcs['{}{}'.format(upper_case_letters[year+2], current_row)] = (
                "=IF({year} < {pbi_year_limit}, "
                "MIN({dol_per_kwh} * {chp_kwh}, {pbi_max}), 0)"
                ).format(
                year=year,
                pbi_year_limit=chp_pbi_years_cell,
                dol_per_kwh=chp_pbi_cell,
                col=upper_case_letters[year+2],
                chp_kwh=chp_annual_kwh_cells[year+1],
                pbi_max=chp_pbi_max_cell,
            )
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    chp_pbi_total_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Total taxable cash incentives"

    for i in range(financial.analysis_years+1):
        pv_cells = list()
        if i == 0:
            for idx in range(len(pv_data)):  # could be multiple PVs
                pv_cells.append(
                    ('IF({pv_ibi_sta_percent_tax_fed_cell}="Yes", {col}{pv_state_ibi_row}, 0)'
                     ' + IF({pv_ibi_uti_percent_tax_fed_cell}="Yes", {col}{pv_utility_ibi_row}, 0)'
                     ' + IF({pv_cbi_fed_tax_fed_cell}="Yes", {col}{pv_federal_cbi_row} ,0)'
                     ' + IF({pv_cbi_sta_tax_fed_cell}="Yes", {col}{pv_state_cbi_row}, 0)'
                     ' + IF({pv_cbi_uti_tax_fed_cell}="Yes", {col}{pv_utility_cbi_row}, 0)'
                     ).format(
                        col=upper_case_letters[i + 1],
                        pv_ibi_sta_percent_tax_fed_cell=pv_cell_locations[idx]["pv_ibi_sta_percent_tax_fed_cell"],
                        pv_state_ibi_row=pv_cell_locations[idx]["pv_state_ibi_row"],
                        pv_ibi_uti_percent_tax_fed_cell=pv_cell_locations[idx]["pv_ibi_uti_percent_tax_fed_cell"],
                        pv_utility_ibi_row=pv_cell_locations[idx]["pv_utility_ibi_row"],
                        pv_cbi_fed_tax_fed_cell=pv_cell_locations[idx]["pv_cbi_fed_tax_fed_cell"],
                        pv_federal_cbi_row=pv_cell_locations[idx]["pv_federal_cbi_row"],
                        pv_cbi_sta_tax_fed_cell=pv_cell_locations[idx]["pv_cbi_sta_tax_fed_cell"],
                        pv_state_cbi_row=pv_cell_locations[idx]["pv_state_cbi_row"],
                        pv_cbi_uti_tax_fed_cell=pv_cell_locations[idx]["pv_cbi_uti_tax_fed_cell"],
                        pv_utility_cbi_row=pv_cell_locations[idx]["pv_utility_cbi_row"]
                    )
                )
            pv_string = '+'.join(pv_cells)

            dcs['{}{}'.format(upper_case_letters[i + 1], current_row)] = (
                '={pv_string}+IF({batt_ibi_sta_percent_tax_fed_cell}="Yes", {col}{batt_state_ibi_row}, 0)'
                '+IF({batt_ibi_uti_percent_tax_fed_cell}="Yes", {col}{batt_utility_ibi_row}, 0)'
                '+IF({batt_total_cbi_per_kw_tax_fed_cell}="Yes", {col}{batt_cbi_per_kw_row}, 0)'
                '+IF({batt_total_cbi_per_kwh_tax_fed_cell}="Yes", {col}{batt_cbi_per_kwh_row}, 0)'
                '+IF({wind_ibi_sta_percent_tax_fed_cell}="Yes", {col}{wind_state_ibi_row}, 0)'
                '+IF({wind_ibi_uti_percent_tax_fed_cell}="Yes", {col}{wind_utility_ibi_row}, 0)'
                '+IF({wind_cbi_fed_tax_fed_cell}="Yes", {col}{wind_federal_cbi_row}, 0)'
                '+IF({wind_cbi_sta_tax_fed_cell}="Yes", {col}{wind_state_cbi_row}, 0)'
                '+IF({wind_cbi_uti_tax_fed_cell}="Yes", {col}{wind_utility_cbi_row}, 0)'
                '+IF({chp_ibi_sta_percent_tax_fed_cell}="Yes", {col}{chp_state_ibi_row}, 0)'
                '+IF({chp_ibi_uti_percent_tax_fed_cell}="Yes", {col}{chp_utility_ibi_row}, 0)'
                '+IF({chp_cbi_fed_tax_fed_cell}="Yes", {col}{chp_federal_cbi_row}, 0)'
                '+IF({chp_cbi_sta_tax_fed_cell}="Yes", {col}{chp_state_cbi_row}, 0)'
                '+IF({chp_cbi_uti_tax_fed_cell}="Yes", {col}{chp_utility_cbi_row}, 0)'
            ).format(
                col=upper_case_letters[i + 1],
                pv_string=pv_string,
                batt_ibi_sta_percent_tax_fed_cell=batt_ibi_sta_percent_tax_fed_cell,
                batt_state_ibi_row=batt_state_ibi_row,
                batt_ibi_uti_percent_tax_fed_cell=batt_ibi_uti_percent_tax_fed_cell,
                batt_utility_ibi_row=batt_utility_ibi_row,
                batt_total_cbi_per_kw_tax_fed_cell=batt_total_cbi_per_kw_tax_fed_cell,
                batt_total_cbi_per_kwh_tax_fed_cell=batt_total_cbi_per_kwh_tax_fed_cell,
                batt_cbi_per_kw_row=batt_cbi_per_kw_row,
                batt_cbi_per_kwh_row=batt_cbi_per_kwh_row,
                wind_ibi_sta_percent_tax_fed_cell=wind_ibi_sta_percent_tax_fed_cell,
                wind_state_ibi_row=wind_state_ibi_row,
                wind_ibi_uti_percent_tax_fed_cell=wind_ibi_uti_percent_tax_fed_cell,
                wind_utility_ibi_row=wind_utility_ibi_row,
                wind_cbi_fed_tax_fed_cell=wind_cbi_fed_tax_fed_cell,
                wind_federal_cbi_row=wind_federal_cbi_row,
                wind_cbi_sta_tax_fed_cell=wind_cbi_sta_tax_fed_cell,
                wind_state_cbi_row=wind_state_cbi_row,
                wind_cbi_uti_tax_fed_cell=wind_cbi_uti_tax_fed_cell,
                wind_utility_cbi_row=wind_utility_cbi_row,
                chp_ibi_sta_percent_tax_fed_cell=chp_ibi_sta_percent_tax_fed_cell,
                chp_state_ibi_row=chp_state_ibi_row,
                chp_ibi_uti_percent_tax_fed_cell=chp_ibi_uti_percent_tax_fed_cell,
                chp_utility_ibi_row=chp_utility_ibi_row,
                chp_cbi_fed_tax_fed_cell=chp_cbi_fed_tax_fed_cell,
                chp_federal_cbi_row=chp_federal_cbi_row,
                chp_cbi_sta_tax_fed_cell=chp_cbi_sta_tax_fed_cell,
                chp_state_cbi_row=chp_state_cbi_row,
                chp_cbi_uti_tax_fed_cell=chp_cbi_uti_tax_fed_cell,
                chp_utility_cbi_row=chp_utility_cbi_row
            )
        else:
            for idx in range(len(pv_data)):  # could be multiple PVs
                pv_cells.append(
                    ('IF({pv_pbi_combined_tax_fed_cell}="Yes", {col}{new_pv_pbi_total_row}, 0)'
                     ' + IF({pv_pbi_combined_tax_fed_cell}="Yes", {col}{existing_pv_pbi_total_row}, 0)'
                     ).format(
                        col=upper_case_letters[i + 1],
                        pv_pbi_combined_tax_fed_cell=pv_cell_locations[idx]["pv_pbi_combined_tax_fed_cell"],
                        new_pv_pbi_total_row=pv_cell_locations[idx]["new_pv_pbi_total_row"],
                        existing_pv_pbi_total_row=pv_cell_locations[idx]["existing_pv_pbi_total_row"],
                    )
                )
            pv_string = '+'.join(pv_cells)

            dcs['{}{}'.format(upper_case_letters[i + 1], current_row)] = (
                '={pv_string} + IF({wind_pbi_combined_tax_fed_cell}="Yes", {col}{wind_pbi_total_row}, 0)'
                '+IF({chp_pbi_combined_tax_fed_cell}="Yes", {col}{chp_pbi_total_row}, 0)'
            ).format(
                col=upper_case_letters[i + 1],
                pv_string=pv_string,
                wind_pbi_combined_tax_fed_cell=wind_pbi_combined_tax_fed_cell,
                wind_pbi_total_row=wind_pbi_total_row,
                chp_pbi_combined_tax_fed_cell=chp_pbi_combined_tax_fed_cell,
                chp_pbi_total_row=chp_pbi_total_row
            )
            make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    fill_border(dcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)
    total_taxable_cash_incentives_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Total non-taxable cash incentives"
    for i in range(financial.analysis_years +1):
        if i == 0:
            pv_cells = list()
            pv_cells.append(
                ('IF({pv_ibi_sta_percent_tax_fed_cell}="No", {col}{pv_state_ibi_row}, 0)'
                 ' + IF({pv_ibi_uti_percent_tax_fed_cell}="No", {col}{pv_utility_ibi_row}, 0)'
                 ' + IF({pv_cbi_fed_tax_fed_cell}="No", {col}{pv_federal_cbi_row} ,0)'
                 ' + IF({pv_cbi_sta_tax_fed_cell}="No", {col}{pv_state_cbi_row}, 0)'
                 ' + IF({pv_cbi_uti_tax_fed_cell}="No", {col}{pv_utility_cbi_row}, 0)'
                 ' + IF({pv_pbi_combined_tax_fed_cell}="No", {col}{new_pv_pbi_total_row}, 0)'
                 ' + IF({pv_pbi_combined_tax_fed_cell}="No", {col}{existing_pv_pbi_total_row}, 0)'
                 ).format(
                    col=upper_case_letters[i + 1],
                    pv_ibi_sta_percent_tax_fed_cell=pv_cell_locations[idx]["pv_ibi_sta_percent_tax_fed_cell"],
                    pv_state_ibi_row=pv_cell_locations[idx]["pv_state_ibi_row"],
                    pv_ibi_uti_percent_tax_fed_cell=pv_cell_locations[idx]["pv_ibi_uti_percent_tax_fed_cell"],
                    pv_utility_ibi_row=pv_cell_locations[idx]["pv_utility_ibi_row"],
                    pv_cbi_fed_tax_fed_cell=pv_cell_locations[idx]["pv_cbi_fed_tax_fed_cell"],
                    pv_federal_cbi_row=pv_cell_locations[idx]["pv_federal_cbi_row"],
                    pv_cbi_sta_tax_fed_cell=pv_cell_locations[idx]["pv_cbi_sta_tax_fed_cell"],
                    pv_state_cbi_row=pv_cell_locations[idx]["pv_state_cbi_row"],
                    pv_cbi_uti_tax_fed_cell=pv_cell_locations[idx]["pv_cbi_uti_tax_fed_cell"],
                    pv_utility_cbi_row=pv_cell_locations[idx]["pv_utility_cbi_row"],
                    pv_pbi_combined_tax_fed_cell=pv_cell_locations[idx]["pv_pbi_combined_tax_fed_cell"],
                    new_pv_pbi_total_row=pv_cell_locations[idx]["new_pv_pbi_total_row"],
                    existing_pv_pbi_total_row=pv_cell_locations[idx]["existing_pv_pbi_total_row"]
                )
               )

            pv_string = '+'.join(pv_cells)

            dcs['{}{}'.format(upper_case_letters[i + 1], current_row)] = (
                '={pv_string}+IF({batt_ibi_sta_percent_tax_fed_cell}="No", {col}{batt_state_ibi_row}, 0)'
                '+IF({batt_ibi_uti_percent_tax_fed_cell}="No", {col}{batt_utility_ibi_row}, 0)'
                '+IF({batt_total_cbi_per_kw_tax_fed_cell}="No", {col}{batt_cbi_per_kw_row}, 0)'
                '+IF({batt_total_cbi_per_kwh_tax_fed_cell}="No", {col}{batt_cbi_per_kwh_row}, 0)'
                '+IF({wind_ibi_sta_percent_tax_fed_cell}="No", {col}{wind_state_ibi_row}, 0)'
                '+IF({wind_ibi_uti_percent_tax_fed_cell}="No", {col}{wind_utility_ibi_row}, 0)'
                '+IF({wind_cbi_fed_tax_fed_cell}="No", {col}{wind_federal_cbi_row}, 0)'
                '+IF({wind_cbi_sta_tax_fed_cell}="No", {col}{wind_state_cbi_row}, 0)'
                '+IF({wind_cbi_uti_tax_fed_cell}="No", {col}{wind_utility_cbi_row}, 0)'
                '+IF({wind_pbi_combined_tax_fed_cell}="No", {col}{wind_pbi_total_row}, 0)'
                '+IF({chp_ibi_sta_percent_tax_fed_cell}="No", {col}{chp_state_ibi_row}, 0)'
                '+IF({chp_ibi_uti_percent_tax_fed_cell}="No", {col}{chp_utility_ibi_row}, 0)'
                '+IF({chp_cbi_fed_tax_fed_cell}="No", {col}{chp_federal_cbi_row}, 0)'
                '+IF({chp_cbi_sta_tax_fed_cell}="No", {col}{chp_state_cbi_row}, 0)'
                '+IF({chp_cbi_uti_tax_fed_cell}="No", {col}{chp_utility_cbi_row}, 0)'
                '+IF({chp_pbi_combined_tax_fed_cell}="No", {col}{chp_pbi_total_row}, 0)'
            ).format(
                col=upper_case_letters[i + 1],
                pv_string=pv_string,
                batt_ibi_sta_percent_tax_fed_cell=batt_ibi_sta_percent_tax_fed_cell,
                batt_state_ibi_row=batt_state_ibi_row,
                batt_ibi_uti_percent_tax_fed_cell=batt_ibi_uti_percent_tax_fed_cell,
                batt_utility_ibi_row=batt_utility_ibi_row,
                batt_total_cbi_per_kw_tax_fed_cell=batt_total_cbi_per_kw_tax_fed_cell,
                batt_total_cbi_per_kwh_tax_fed_cell=batt_total_cbi_per_kwh_tax_fed_cell,
                batt_cbi_per_kw_row=batt_cbi_per_kw_row,
                batt_cbi_per_kwh_row=batt_cbi_per_kwh_row,
                wind_ibi_sta_percent_tax_fed_cell=wind_ibi_sta_percent_tax_fed_cell,
                wind_state_ibi_row=wind_state_ibi_row,
                wind_ibi_uti_percent_tax_fed_cell=wind_ibi_uti_percent_tax_fed_cell,
                wind_utility_ibi_row=wind_utility_ibi_row,
                wind_cbi_fed_tax_fed_cell=wind_cbi_fed_tax_fed_cell,
                wind_federal_cbi_row=wind_federal_cbi_row,
                wind_cbi_sta_tax_fed_cell=wind_cbi_sta_tax_fed_cell,
                wind_state_cbi_row=wind_state_cbi_row,
                wind_cbi_uti_tax_fed_cell=wind_cbi_uti_tax_fed_cell,
                wind_utility_cbi_row=wind_utility_cbi_row,
                wind_pbi_combined_tax_fed_cell=wind_pbi_combined_tax_fed_cell,
                wind_pbi_total_row=wind_pbi_total_row,
                chp_ibi_sta_percent_tax_fed_cell=chp_ibi_sta_percent_tax_fed_cell,
                chp_state_ibi_row=chp_state_ibi_row,
                chp_ibi_uti_percent_tax_fed_cell=chp_ibi_uti_percent_tax_fed_cell,
                chp_utility_ibi_row=chp_utility_ibi_row,
                chp_cbi_fed_tax_fed_cell=chp_cbi_fed_tax_fed_cell,
                chp_federal_cbi_row=chp_federal_cbi_row,
                chp_cbi_sta_tax_fed_cell=chp_cbi_sta_tax_fed_cell,
                chp_state_cbi_row=chp_state_cbi_row,
                chp_cbi_uti_tax_fed_cell=chp_cbi_uti_tax_fed_cell,
                chp_utility_cbi_row=chp_utility_cbi_row,
                chp_pbi_combined_tax_fed_cell=chp_pbi_combined_tax_fed_cell,
                chp_pbi_total_row=chp_pbi_total_row
             )
        else:
            pv_cells = list()
            for idx in range(len(pv_data)):  # could be multiple PVs
                pv_cells.append(
                    ('   IF({pv_pbi_combined_tax_fed_cell}="No", {col}{new_pv_pbi_total_row}, 0)'
                     ' + IF({pv_pbi_combined_tax_fed_cell}="No", {col}{existing_pv_pbi_total_row}, 0)' 
                     ).format(
                        col=upper_case_letters[i + 1],
                        pv_pbi_combined_tax_fed_cell=pv_cell_locations[idx]["pv_pbi_combined_tax_fed_cell"],
                        new_pv_pbi_total_row=pv_cell_locations[idx]["new_pv_pbi_total_row"],
                        existing_pv_pbi_total_row=pv_cell_locations[idx]["existing_pv_pbi_total_row"]
                    )
                )
            pv_string = '+'.join(pv_cells)
     
            dcs['{}{}'.format(upper_case_letters[i + 1], current_row)] = (
                '={pv_string}'
                '+IF({wind_pbi_combined_tax_fed_cell}="No", {col}{wind_pbi_total_row}, 0)'
                '+IF({chp_pbi_combined_tax_fed_cell}="No", {col}{chp_pbi_total_row}, 0)'
            ).format(
                col=upper_case_letters[i + 1],
                pv_string=pv_string,
                wind_pbi_combined_tax_fed_cell=wind_pbi_combined_tax_fed_cell,
                wind_pbi_total_row=wind_pbi_total_row,
                chp_pbi_combined_tax_fed_cell=chp_pbi_combined_tax_fed_cell,
                chp_pbi_total_row=chp_pbi_total_row
            )
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    fill_border(dcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)
    total_nontaxable_cash_incentives_row = current_row
    current_row += 1
    current_row += 1

    ####################################################################################################################
    # Capital Depreciation
    ####################################################################################################################
    dcs['A{}'.format(current_row)] = "Capital Depreciation"
    make_title_row(dcs, current_row, length=financial.analysis_years+2)

    ####################################################################################################################
    # PV depreciation
    ####################################################################################################################

    current_row += 1
    for idx, pv in enumerate(pv_data):
        dcs['A{}'.format(current_row)] = "{} Depreciation, Commercial only".format(pv['name'])
        dcs['A{}'.format(current_row)].alignment = one_tab_indent
        make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)

        current_row += 1
        dcs['A{}'.format(current_row)] = "Percentage"
        dcs['A{}'.format(current_row)].alignment = two_tab_indent
        for i in range(len(pv_cell_locations[idx]["pv_macrs_cells"])):
            dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = '={}'.format(
                pv_cell_locations[idx]["pv_macrs_cells"][i])
        make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                           number_format='#,##0.0000', border=no_border)
        pv_cell_locations[idx]["pv_macrs_percent_row"] = current_row

        current_row += 1
        dcs['A{}'.format(current_row)] = "Bonus Basis"
        dcs['A{}'.format(current_row)].alignment = two_tab_indent
        make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
        pv_cell_locations[idx]["pv_bonus_basis_cell"] = 'B{}'.format(current_row)

        current_row += 1
        dcs['A{}'.format(current_row)] = "Basis"
        dcs['A{}'.format(current_row)].alignment = two_tab_indent
        dcs['B{}'.format(current_row)] = '={}*(1-{})'.format(pv_cell_locations[idx]["pv_bonus_basis_cell"],
                                                             pv_cell_locations[idx]["pv_bonus_fraction_cell"])
        make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
        pv_cell_locations[idx]["pv_basis_cell"] = 'B{}'.format(current_row)

        current_row += 1
        dcs['A{}'.format(current_row)] = "Amount"
        dcs['A{}'.format(current_row)].alignment = two_tab_indent
        pv_cell_locations[idx]["pv_depreciation_benefit"] = list()
        for i in range(financial.analysis_years):
            pv_cell_locations[idx]["pv_depreciation_benefit"].append("\'{}\'!{}{}".format(
                    developer_cashflow_sheet_name, upper_case_letters[i+2], current_row))
            if i == 0:
                dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = \
                    '={basis_cell}*{col}{macrs_row} + ({bonus_basis_cell}*{bonus_basis_pct_cell})'.format(
                        basis_cell=pv_cell_locations[idx]["pv_basis_cell"],
                        col=upper_case_letters[i + 2],
                        macrs_row=pv_cell_locations[idx]["pv_macrs_percent_row"],
                        bonus_basis_cell=pv_cell_locations[idx]["pv_bonus_basis_cell"],
                        bonus_basis_pct_cell=pv_cell_locations[idx]["pv_bonus_fraction_cell"]
                    )
            else:
                dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = \
                    '={basis_cell}*{col}{macrs_row}'.format(
                        basis_cell=pv_cell_locations[idx]["pv_basis_cell"],
                        col=upper_case_letters[i + 2],
                        macrs_row=pv_cell_locations[idx]["pv_macrs_percent_row"]
                    )
        make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        pv_cell_locations[idx]["pv_fed_income_total"] = current_row
        current_row += 1

    ####################################################################################################################
    # Wind depreciation
    ####################################################################################################################

    dcs['A{}'.format(current_row)] = "Wind Depreciation, Commercial only"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Percentage"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    for i in range(len(wind_macrs_cells)):
        dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = '={}'.format(wind_macrs_cells[i])
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0.0000', border=no_border)
    wind_macrs_percent_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Bonus Basis"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    wind_bonus_basis_cell = 'B{}'.format(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Basis"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    dcs['B{}'.format(current_row)] = '={}*(1-{})'.format(wind_bonus_basis_cell, wind_bonus_fraction_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    wind_basis_cell = 'B{}'.format(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Amount"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    wind_depreciation_benefit = list()
    for i in range(financial.analysis_years):
        wind_depreciation_benefit.append("\'{}\'!{}{}".format(
            developer_cashflow_sheet_name, upper_case_letters[i+2], current_row))
        
        if i == 0:
            dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = \
                '={basis_cell}*{col}{macrs_row} + ({bonus_basis_cell}*{bonus_basis_pct_cell})'.format(
                    basis_cell=wind_basis_cell, col=upper_case_letters[i + 2], macrs_row=wind_macrs_percent_row,
                    bonus_basis_cell=wind_bonus_basis_cell, bonus_basis_pct_cell=wind_bonus_fraction_cell)
        else:
            dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = '={basis_cell}*{col}{macrs_row}'.format(
                basis_cell=wind_basis_cell, col=upper_case_letters[i + 2], macrs_row=wind_macrs_percent_row)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    wind_fed_income_total = current_row
    current_row += 1

    ####################################################################################################################
    # Battery depreciation
    ####################################################################################################################

    dcs['A{}'.format(current_row)] = "Battery Depreciation, Commercial only"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Percentage"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    for i in range(len(batt_macrs_cells)):
        dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = '={}'.format(batt_macrs_cells[i])
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0.0000', border=no_border)
    batt_macrs_percent_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Bonus Basis"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    batt_bonus_basis_cell = 'B{}'.format(current_row)
    current_row += 1
    dcs['A{}'.format(current_row)] = "Basis"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    dcs['B{}'.format(current_row)] = '={}*(1-{})'.format(batt_bonus_basis_cell, batt_bonus_fraction_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    batt_basis_cell = 'B{}'.format(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Amount"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    for i in range(financial.analysis_years):
        if i == 0:
            dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = \
                '={basis_cell}*{col}{macrs_row} + ({bonus_basis_cell}*{bonus_basis_pct_cell})'.format(
                    basis_cell=batt_basis_cell, col=upper_case_letters[i + 2], macrs_row=batt_macrs_percent_row,
                    bonus_basis_cell=batt_bonus_basis_cell, bonus_basis_pct_cell=batt_bonus_fraction_cell)
        else:
            dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = '={basis_cell}*{col}{macrs_row}'.format(
                basis_cell=batt_basis_cell, col=upper_case_letters[i + 2], macrs_row=batt_macrs_percent_row)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    batt_fed_income_total = current_row
    current_row += 1
    
    ####################################################################################################################
    # CHP depreciation
    ####################################################################################################################

    dcs['A{}'.format(current_row)] = "CHP Depreciation, Commercial only"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Percentage"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    for i in range(len(chp_macrs_cells)):
        dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = '={}'.format(chp_macrs_cells[i])
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0.0000', border=no_border)
    chp_macrs_percent_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Bonus Basis"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    chp_bonus_basis_cell = 'B{}'.format(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Basis"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    dcs['B{}'.format(current_row)] = '={}*(1-{})'.format(chp_bonus_basis_cell, chp_bonus_fraction_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    chp_basis_cell = 'B{}'.format(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Amount"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    chp_depreciation_benefit = list()
    for i in range(financial.analysis_years):
        chp_depreciation_benefit.append("\'{}\'!{}{}".format(
            developer_cashflow_sheet_name, upper_case_letters[i+2], current_row))
        
        if i == 0:
            dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = \
                '={basis_cell}*{col}{macrs_row} + ({bonus_basis_cell}*{bonus_basis_pct_cell})'.format(
                    basis_cell=chp_basis_cell, col=upper_case_letters[i + 2], macrs_row=chp_macrs_percent_row,
                    bonus_basis_cell=chp_bonus_basis_cell, bonus_basis_pct_cell=chp_bonus_fraction_cell)
        else:
            dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = '={basis_cell}*{col}{macrs_row}'.format(
                basis_cell=chp_basis_cell, col=upper_case_letters[i + 2], macrs_row=chp_macrs_percent_row)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    chp_fed_income_total = current_row
    current_row += 1

    ####################################################################################################################
    # Absorption Chiller depreciation
    ####################################################################################################################

    dcs['A{}'.format(current_row)] = "Absorption Chiller Depreciation, Commercial only"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Percentage"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    for i in range(len(absorption_chiller_macrs_cells)):
        dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = '={}'.format(absorption_chiller_macrs_cells[i])
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0.0000', border=no_border)
    absorption_chiller_macrs_percent_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Bonus Basis"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    dcs['B{}'.format(current_row)] = (
            "=IF(OR({absorption_chiller_macrs_option_cell}=5,{absorption_chiller_macrs_option_cell}=7),"
            "{absorption_chiller_cost_cell},0)").format(
            absorption_chiller_macrs_option_cell=absorption_chiller_macrs_option_cell, 
            absorption_chiller_cost_cell=absorption_chiller_cost_cell)
    absorption_chiller_bonus_basis_cell = 'B{}'.format(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Basis"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    dcs['B{}'.format(current_row)] = '={}*(1-{})'.format(absorption_chiller_bonus_basis_cell, absorption_chiller_bonus_fraction_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    absorption_chiller_basis_cell = 'B{}'.format(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Amount"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    absorption_chiller_depreciation_benefit = list()
    for i in range(financial.analysis_years):
        absorption_chiller_depreciation_benefit.append("\'{}\'!{}{}".format(
            developer_cashflow_sheet_name, upper_case_letters[i+2], current_row))
        
        if i == 0:
            dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = \
                '={basis_cell}*{col}{macrs_row} + ({bonus_basis_cell}*{bonus_basis_pct_cell})'.format(
                    basis_cell=absorption_chiller_basis_cell, col=upper_case_letters[i + 2], macrs_row=absorption_chiller_macrs_percent_row,
                    bonus_basis_cell=absorption_chiller_bonus_basis_cell, bonus_basis_pct_cell=absorption_chiller_bonus_fraction_cell)
        else:
            dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = '={basis_cell}*{col}{macrs_row}'.format(
                basis_cell=absorption_chiller_basis_cell, col=upper_case_letters[i + 2], macrs_row=absorption_chiller_macrs_percent_row)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    absorption_chiller_fed_income_total = current_row
    current_row += 1

    ####################################################################################################################
    # Hot TES depreciation
    ####################################################################################################################

    dcs['A{}'.format(current_row)] = "Hot TES Depreciation, Commercial only"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Percentage"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    for i in range(len(hot_tes_macrs_cells)):
        dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = '={}'.format(hot_tes_macrs_cells[i])
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0.0000', border=no_border)
    hot_tes_macrs_percent_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Bonus Basis"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    dcs['B{}'.format(current_row)] = (
            "=IF(OR({hot_tes_macrs_option_cell}=5,{hot_tes_macrs_option_cell}=7),"
            "{hot_tes_cost_cell},0)").format(
            hot_tes_macrs_option_cell=hot_tes_macrs_option_cell, 
            hot_tes_cost_cell=hot_tes_cost_cell)
    hot_tes_bonus_basis_cell = 'B{}'.format(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Basis"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    dcs['B{}'.format(current_row)] = '={}*(1-{})'.format(hot_tes_bonus_basis_cell, hot_tes_bonus_fraction_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    hot_tes_basis_cell = 'B{}'.format(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Amount"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    hot_tes_depreciation_benefit = list()
    for i in range(financial.analysis_years):
        hot_tes_depreciation_benefit.append("\'{}\'!{}{}".format(
            developer_cashflow_sheet_name, upper_case_letters[i+2], current_row))
        
        if i == 0:
            dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = \
                '={basis_cell}*{col}{macrs_row} + ({bonus_basis_cell}*{bonus_basis_pct_cell})'.format(
                    basis_cell=hot_tes_basis_cell, col=upper_case_letters[i + 2], macrs_row=hot_tes_macrs_percent_row,
                    bonus_basis_cell=hot_tes_bonus_basis_cell, bonus_basis_pct_cell=hot_tes_bonus_fraction_cell)
        else:
            dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = '={basis_cell}*{col}{macrs_row}'.format(
                basis_cell=hot_tes_basis_cell, col=upper_case_letters[i + 2], macrs_row=hot_tes_macrs_percent_row)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    hot_tes_fed_income_total = current_row
    current_row += 1

    ####################################################################################################################
    # Cold TES depreciation
    ####################################################################################################################

    dcs['A{}'.format(current_row)] = "Cold TES Depreciation, Commercial only"
    dcs['A{}'.format(current_row)].alignment = one_tab_indent
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Percentage"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    for i in range(len(cold_tes_macrs_cells)):
        dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = '={}'.format(cold_tes_macrs_cells[i])
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0.0000', border=no_border)
    cold_tes_macrs_percent_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Bonus Basis"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    dcs['B{}'.format(current_row)] = (
            "=IF(OR({cold_tes_macrs_option_cell}=5,{cold_tes_macrs_option_cell}=7),"
            "{cold_tes_cost_cell},0)").format(
            cold_tes_macrs_option_cell=cold_tes_macrs_option_cell, 
            cold_tes_cost_cell=cold_tes_cost_cell)
    cold_tes_bonus_basis_cell = 'B{}'.format(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Basis"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    dcs['B{}'.format(current_row)] = '={}*(1-{})'.format(cold_tes_bonus_basis_cell, cold_tes_bonus_fraction_cell)
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    cold_tes_basis_cell = 'B{}'.format(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Amount"
    dcs['A{}'.format(current_row)].alignment = two_tab_indent
    cold_tes_depreciation_benefit = list()
    for i in range(financial.analysis_years):
        cold_tes_depreciation_benefit.append("\'{}\'!{}{}".format(
            developer_cashflow_sheet_name, upper_case_letters[i+2], current_row))
        
        if i == 0:
            dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = \
                '={basis_cell}*{col}{macrs_row} + ({bonus_basis_cell}*{bonus_basis_pct_cell})'.format(
                    basis_cell=cold_tes_basis_cell, col=upper_case_letters[i + 2], macrs_row=cold_tes_macrs_percent_row,
                    bonus_basis_cell=cold_tes_bonus_basis_cell, bonus_basis_pct_cell=cold_tes_bonus_fraction_cell)
        else:
            dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = '={basis_cell}*{col}{macrs_row}'.format(
                basis_cell=cold_tes_basis_cell, col=upper_case_letters[i + 2], macrs_row=cold_tes_macrs_percent_row)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    cold_tes_fed_income_total = current_row
    current_row += 1
    ####################################################################################################################
    # Total depreciation
    ####################################################################################################################

    dcs['A{}'.format(current_row)] = "Total depreciation"
    dcs['A{}'.format(current_row)].font = bold_font
    # dcs['A{}'.format(current_row)].alignment = one_tab_indent
    for i in range(financial.analysis_years):
        pv_string = ','.join(["{col}{pv}".format(
            col=upper_case_letters[i + 2],
            pv=pv_cell_locations[idx]["pv_fed_income_total"]) for idx in range(len(pv_data))])
        dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = \
            '=SUM({pv_string},{col}{wind},{col}{batt},{col}{chp},{col}{absorption_chiller},{col}{hot_tes},{col}{cold_tes})'.format(
                col=upper_case_letters[i + 2], pv_string=pv_string,
                wind=wind_fed_income_total, batt=batt_fed_income_total, 
                chp=chp_fed_income_total, absorption_chiller=absorption_chiller_fed_income_total,
                hot_tes=hot_tes_fed_income_total, cold_tes=cold_tes_fed_income_total)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    fill_border(dcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)
    total_depreciation_row = current_row
    current_row += 1
    current_row += 1

    ####################################################################################################################
    # Tax benefits for LCOE Calculations
    ####################################################################################################################
    dcs['A{}'.format(current_row)] = "Tax benefits for LCOE Calculations"
    make_title_row(dcs, current_row, length=financial.analysis_years+2)
    current_row += 1

    for idx, pv in enumerate(pv_data):
        dcs['A{}'.format(current_row)] = "{} income tax savings, $".format(pv['name'])
        pv_cell_locations[idx]["pv_income_tax_benefits"] = list()
        for year in range(1, financial.analysis_years + 1):
            dcs['{}{}'.format(upper_case_letters[year+1], current_row)] = \
                '=((-1 * {pv_om_cell}) + {pv_depr_cell}) * {fed_tax_rate_cell}/100'.format(
                    pv_om_cell = pv_cell_locations[idx]['new_o_and_m_cashflow_cells'][year -1],
                    pv_depr_cell = pv_cell_locations[idx]["pv_depreciation_benefit"][year -1],
                    fed_tax_rate_cell=fed_tax_rate_cell
                )
            pv_cell_locations[idx]["pv_income_tax_benefits"].append("\'{}\'!{}{}".format(
                developer_cashflow_sheet_name, upper_case_letters[year+1], current_row))
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    fill_border(dcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)
    current_row += 1
    
    dcs['A{}'.format(current_row)] = "Wind income tax savings, $"
    wind_income_tax_benefits = list()
    for year in range(1, financial.analysis_years + 1):
        dcs['{}{}'.format(upper_case_letters[year+1], current_row)] = \
            '=((-1 * {wind_om_cell}) + {wind_depr_cell}) * {fed_tax_rate_cell}/100'.format(
                wind_om_cell = wind_o_and_m_cashflow_cells[year -1],
                wind_depr_cell = wind_depreciation_benefit[year -1],
                fed_tax_rate_cell=fed_tax_rate_cell
            )
        wind_income_tax_benefits.append("\'{}\'!{}{}".format(
            developer_cashflow_sheet_name, upper_case_letters[year+1], current_row))
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    fill_border(dcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)
    current_row += 1
    current_row += 1

    ####################################################################################################################
    # Federal ITC
    ####################################################################################################################

    dcs['A{}'.format(current_row)] = "Federal Investment Tax Credit"
    make_title_row(dcs, current_row, length=financial.analysis_years+2)

    current_row += 1
    for idx, pv in enumerate(pv_data):
        dcs['A{}'.format(current_row)] = "Federal ITC basis: {}".format(pv["name"])
        dcs['B{}'.format(current_row)] = (
            '={pv_cost_cell}-IF({pv_ibi_sta_percent_deprbas_fed_cell}="Yes",'
            '{col}{pv_state_ibi_row},0)-IF({pv_ibi_uti_percent_deprbas_fed_cell}="Yes",'
            '{col}{pv_utility_ibi_row},0)-IF({pv_cbi_fed_deprbas_fed_cell}="Yes",'
            '{col}{pv_federal_cbi_row},0)-IF({pv_cbi_sta_deprbas_fed_cell}="Yes",'
            '{col}{pv_state_cbi_row},0)-IF({pv_cbi_uti_deprbas_fed_cell}="Yes",'
            '{col}{pv_utility_cbi_row},0)'
        ).format(
            pv_cost_cell=pv_cell_locations[idx]["pv_cost_cell"],
            pv_ibi_sta_percent_deprbas_fed_cell=pv_cell_locations[idx]["pv_ibi_sta_percent_deprbas_fed_cell"],
            col='B',
            pv_state_ibi_row=pv_cell_locations[idx]["pv_state_ibi_row"],
            pv_ibi_uti_percent_deprbas_fed_cell=pv_cell_locations[idx]["pv_ibi_uti_percent_deprbas_fed_cell"],
            pv_utility_ibi_row=pv_cell_locations[idx]["pv_utility_ibi_row"],
            pv_cbi_fed_deprbas_fed_cell=pv_cell_locations[idx]["pv_cbi_fed_deprbas_fed_cell"],
            pv_federal_cbi_row=pv_cell_locations[idx]["pv_federal_cbi_row"],
            pv_cbi_sta_deprbas_fed_cell=pv_cell_locations[idx]["pv_cbi_sta_deprbas_fed_cell"],
            pv_state_cbi_row=pv_cell_locations[idx]["pv_state_cbi_row"],
            pv_cbi_uti_deprbas_fed_cell=pv_cell_locations[idx]["pv_cbi_uti_deprbas_fed_cell"],
            pv_utility_cbi_row=pv_cell_locations[idx]["pv_utility_cbi_row"]
        )
        dcs[pv_cell_locations[idx]["pv_bonus_basis_cell"]] = (
            '=IF(OR({pv_macrs_option_cell}=5,{pv_macrs_option_cell}=7),'
            '({pv_itc_basis_cell}-IF({pv_itc_fed_percent_deprbas_fed_cell}="Yes",'
            '0.5*MIN({pv_itc_fed_percent_cell}/100*{pv_itc_basis_cell},{pv_itc_fed_percent_maxvalue_cell}),0)),0)'
        ).format(
            pv_macrs_option_cell=pv_cell_locations[idx]["pv_macrs_option_cell"],
            pv_itc_basis_cell='B{}'.format(current_row),
            pv_itc_fed_percent_deprbas_fed_cell=pv_cell_locations[idx]["pv_itc_fed_percent_deprbas_fed_cell"],
            pv_itc_fed_percent_cell=pv_cell_locations[idx]["pv_itc_fed_percent_cell"],
            pv_itc_fed_percent_maxvalue_cell=pv_cell_locations[idx]["pv_itc_fed_percent_maxvalue_cell"],
        )
        pv_cell_locations[idx]["pv_federal_itc_basis_cell"] = 'B{}'.format(current_row)
        make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
        current_row += 1
        dcs['A{}'.format(current_row)] = "Federal ITC amount: {}".format(pv["name"])
        dcs['C{}'.format(current_row)] = (
            '=MIN({pv_itc_fed_percent_cell}/100*{pv_federal_itc_basis_cell},{pv_itc_fed_percent_maxvalue_cell})'
        ).format(
            pv_itc_fed_percent_cell=pv_cell_locations[idx]["pv_itc_fed_percent_cell"],
            pv_federal_itc_basis_cell=pv_cell_locations[idx]["pv_federal_itc_basis_cell"],
            pv_itc_fed_percent_maxvalue_cell=pv_cell_locations[idx]["pv_itc_fed_percent_maxvalue_cell"]
        )
        make_attribute_row(dcs, current_row, length=3, alignment=right_align, number_format='#,##0', border=no_border)
        pv_cell_locations[idx]["pv_federal_itc_amount_row"] = current_row
        pv_cell_locations[idx]["pv_federal_itc_amount_cell"] = "\'{}\'!C{}".format(
                developer_cashflow_sheet_name, current_row)
        current_row += 1

    dcs['A{}'.format(current_row)] = "Federal ITC basis: Wind"
    dcs['B{}'.format(current_row)] = (
        '={wind_cost_cell}-IF({wind_ibi_sta_percent_deprbas_fed_cell}="Yes",'
        '{col}{wind_state_ibi_row},0)-IF({wind_ibi_uti_percent_deprbas_fed_cell}="Yes",'
        '{col}{wind_utility_ibi_row},0)-IF({wind_cbi_fed_deprbas_fed_cell}="Yes",'
        '{col}{wind_federal_cbi_row},0)-IF({wind_cbi_sta_deprbas_fed_cell}="Yes",'
        '{col}{wind_state_cbi_row},0)-IF({wind_cbi_uti_deprbas_fed_cell}="Yes",'
        '{col}{wind_utility_cbi_row},0)'
    ).format(
        wind_cost_cell=wind_cost_cell,
        wind_ibi_sta_percent_deprbas_fed_cell=wind_ibi_sta_percent_deprbas_fed_cell,
        col='B',
        wind_state_ibi_row=wind_state_ibi_row,
        wind_ibi_uti_percent_deprbas_fed_cell=wind_ibi_uti_percent_deprbas_fed_cell,
        wind_utility_ibi_row=wind_utility_ibi_row,
        wind_cbi_fed_deprbas_fed_cell=wind_cbi_fed_deprbas_fed_cell,
        wind_federal_cbi_row=wind_federal_cbi_row,
        wind_cbi_sta_deprbas_fed_cell=wind_cbi_sta_deprbas_fed_cell,
        wind_state_cbi_row=wind_state_cbi_row,
        wind_cbi_uti_deprbas_fed_cell=wind_cbi_uti_deprbas_fed_cell,
        wind_utility_cbi_row=wind_utility_cbi_row
    )
    dcs[wind_bonus_basis_cell] = (
        '=IF(OR({wind_macrs_option_cell}=5,{wind_macrs_option_cell}=7),'
        '({wind_itc_basis_cell}-IF({wind_itc_fed_percent_deprbas_fed_cell}="Yes",'
        '0.5*MIN({wind_federal_itc_cell}/100*{wind_itc_basis_cell},{wind_itc_fed_percent_maxvalue_cell}),0)),0)'
    ).format(
        wind_macrs_option_cell=wind_macrs_option_cell,
        wind_itc_basis_cell='B{}'.format(current_row),
        wind_itc_fed_percent_deprbas_fed_cell=wind_itc_fed_percent_deprbas_fed_cell,
        wind_federal_itc_cell=wind_federal_itc_cell,
        wind_itc_fed_percent_maxvalue_cell=wind_itc_fed_percent_maxvalue_cell
    )
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    wind_federal_itc_basis_cell = 'B{}'.format(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Federal ITC amount: Wind"
    dcs['C{}'.format(current_row)] = (
        '=MIN({wind_federal_itc_cell}/100*{wind_federal_itc_basis_cell},{wind_itc_fed_percent_maxvalue_cell})'
    ).format(
        wind_federal_itc_cell=wind_federal_itc_cell,
        wind_federal_itc_basis_cell=wind_federal_itc_basis_cell,
        wind_itc_fed_percent_maxvalue_cell=wind_itc_fed_percent_maxvalue_cell
    )
    make_attribute_row(dcs, current_row, length=3, alignment=right_align, number_format='#,##0', border=no_border)
    wind_federal_itc_amount_row = current_row
    wind_federal_itc_amount_cell = "\'{}\'!C{}".format(
                developer_cashflow_sheet_name, current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Federal ITC basis: Battery"
    dcs['B{}'.format(current_row)] = (
        '={batt_cost_cell}-IF({batt_ibi_sta_percent_deprbas_fed_cell}="Yes",'
        '{col}{batt_state_ibi_row},0)-IF({batt_ibi_uti_percent_deprbas_fed_cell}="Yes",'
        '{col}{batt_utility_ibi_row},0)-IF({batt_total_cbi_per_kw_deprbas_fed_cell}="Yes",'
        '{col}{batt_cbi_per_kw_row},0)-IF({batt_total_cbi_per_kwh_deprbas_fed_cell}="Yes",'
        '{col}{batt_cbi_per_kwh_row},0)'
    ).format(
        batt_cost_cell=batt_cost_cell,
        batt_ibi_sta_percent_deprbas_fed_cell=batt_ibi_sta_percent_deprbas_fed_cell,
        col='B',
        batt_state_ibi_row=batt_state_ibi_row,
        batt_ibi_uti_percent_deprbas_fed_cell=batt_ibi_uti_percent_deprbas_fed_cell,
        batt_utility_ibi_row=batt_utility_ibi_row,
        batt_total_cbi_per_kw_deprbas_fed_cell=batt_total_cbi_per_kw_deprbas_fed_cell,
        batt_total_cbi_per_kwh_deprbas_fed_cell=batt_total_cbi_per_kwh_deprbas_fed_cell,
        batt_cbi_per_kw_row=batt_cbi_per_kw_row,
        batt_cbi_per_kwh_row=batt_cbi_per_kwh_row
    )
    dcs[batt_bonus_basis_cell] = (
        '=IF(OR({batt_macrs_option_cell}=5,{batt_macrs_option_cell}=7),'
        '({batt_itc_basis_cell}-IF({batt_itc_fed_percent_deprbas_fed_cell}="Yes",'
        '0.5*MIN({batt_federal_itc_cell}/100*{batt_itc_basis_cell},{batt_itc_fed_percent_maxvalue_cell}),0)),0)'
    ).format(
        batt_macrs_option_cell=batt_macrs_option_cell,
        batt_itc_basis_cell='B{}'.format(current_row),
        batt_itc_fed_percent_deprbas_fed_cell=batt_itc_fed_percent_deprbas_fed_cell,
        batt_federal_itc_cell=batt_federal_itc_cell,
        batt_itc_fed_percent_maxvalue_cell=batt_itc_fed_percent_maxvalue_cell
    )
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    batt_federal_itc_basis_cell = 'B{}'.format(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Federal ITC amount: Battery"
    dcs['C{}'.format(current_row)] = (
        '=MIN({batt_federal_itc_cell}/100*{batt_federal_itc_basis_cell},{batt_itc_fed_percent_maxvalue_cell})'
    ).format(
        batt_federal_itc_cell=batt_federal_itc_cell,
        batt_federal_itc_basis_cell=batt_federal_itc_basis_cell,
        batt_itc_fed_percent_maxvalue_cell=batt_itc_fed_percent_maxvalue_cell
    )
    make_attribute_row(dcs, current_row, length=3, alignment=right_align, number_format='#,##0', border=no_border)
    batt_federal_itc_amount_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Federal ITC basis: CHP"
    dcs['B{}'.format(current_row)] = (
        '={chp_cost_cell}-IF({chp_ibi_sta_percent_deprbas_fed_cell}="Yes",'
        '{col}{chp_state_ibi_row},0)-IF({chp_ibi_uti_percent_deprbas_fed_cell}="Yes",'
        '{col}{chp_utility_ibi_row},0)-IF({chp_cbi_fed_deprbas_fed_cell}="Yes",'
        '{col}{chp_federal_cbi_row},0)-IF({chp_cbi_sta_deprbas_fed_cell}="Yes",'
        '{col}{chp_state_cbi_row},0)-IF({chp_cbi_uti_deprbas_fed_cell}="Yes",'
        '{col}{chp_utility_cbi_row},0)'
    ).format(
        chp_cost_cell=chp_cost_cell,
        chp_ibi_sta_percent_deprbas_fed_cell=chp_ibi_sta_percent_deprbas_fed_cell,
        col='B',
        chp_state_ibi_row=chp_state_ibi_row,
        chp_ibi_uti_percent_deprbas_fed_cell=chp_ibi_uti_percent_deprbas_fed_cell,
        chp_utility_ibi_row=chp_utility_ibi_row,
        chp_cbi_fed_deprbas_fed_cell=chp_cbi_fed_deprbas_fed_cell,
        chp_federal_cbi_row=chp_federal_cbi_row,
        chp_cbi_sta_deprbas_fed_cell=chp_cbi_sta_deprbas_fed_cell,
        chp_state_cbi_row=chp_state_cbi_row,
        chp_cbi_uti_deprbas_fed_cell=chp_cbi_uti_deprbas_fed_cell,
        chp_utility_cbi_row=chp_utility_cbi_row
    )
    dcs[chp_bonus_basis_cell] = (
        '=IF(OR({chp_macrs_option_cell}=5,{chp_macrs_option_cell}=7),'
        '({chp_itc_basis_cell}-IF({chp_itc_fed_percent_deprbas_fed_cell}="Yes",'
        '0.5*MIN({chp_federal_itc_cell}/100*{chp_itc_basis_cell},{chp_itc_fed_percent_maxvalue_cell}),0)),0)'
    ).format(
        chp_macrs_option_cell=chp_macrs_option_cell,
        chp_itc_basis_cell='B{}'.format(current_row),
        chp_itc_fed_percent_deprbas_fed_cell=chp_itc_fed_percent_deprbas_fed_cell,
        chp_federal_itc_cell=chp_federal_itc_cell,
        chp_itc_fed_percent_maxvalue_cell=chp_itc_fed_percent_maxvalue_cell
    )
    make_attribute_row(dcs, current_row, length=2, alignment=right_align, number_format='#,##0', border=no_border)
    chp_federal_itc_basis_cell = 'B{}'.format(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Federal ITC amount: CHP"
    dcs['C{}'.format(current_row)] = (
        '=MIN({chp_federal_itc_cell}/100*{chp_federal_itc_basis_cell},{chp_itc_fed_percent_maxvalue_cell})'
    ).format(
        chp_federal_itc_cell=chp_federal_itc_cell,
        chp_federal_itc_basis_cell=chp_federal_itc_basis_cell,
        chp_itc_fed_percent_maxvalue_cell=chp_itc_fed_percent_maxvalue_cell
    )
    make_attribute_row(dcs, current_row, length=3, alignment=right_align, number_format='#,##0', border=no_border)
    chp_federal_itc_amount_row = current_row
    chp_federal_itc_amount_cell = "\'{}\'!C{}".format(
                developer_cashflow_sheet_name, current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Total Federal ITC"
    dcs['A{}'.format(current_row)].font = bold_font
    fill_border(dcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)
    pv_string = ",".join([
        "C{pv_federal_itc_amount_row}".format(
            pv_federal_itc_amount_row=pv_cell_locations[idx]["pv_federal_itc_amount_row"])
        for idx in range(len(pv_data))
    ])
    dcs['C{}'.format(current_row)] = (
        '=SUM({pv_string}, {col}{wind_federal_itc_amount_row}, '
        '{col}{batt_federal_itc_amount_row}, {col}{chp_federal_itc_amount_row})'
    ).format(
        col=upper_case_letters[0 + 2],
        pv_string=pv_string,
        wind_federal_itc_amount_row=wind_federal_itc_amount_row,
        batt_federal_itc_amount_row=batt_federal_itc_amount_row,
        chp_federal_itc_amount_row=chp_federal_itc_amount_row
    )
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    fill_border(dcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)
    total_fed_itc_row = current_row
    current_row += 1
    current_row += 1

    ####################################################################################################################
    # After-tax Cash Flows
    ####################################################################################################################

    dcs['A{}'.format(current_row)] = "Total Cash Flows"
    make_title_row(dcs, current_row, length=financial.analysis_years + 2)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Upfront Capital Cost"
    dcs['B{}'.format(current_row)] = "=-{}".format(installed_costs_cell)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    upfront_cost_cell = "B{}".format(current_row)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Operating expenses, after-tax"

    for year in range(1, financial.analysis_years + 1):
        dcs['{}{}'.format(upper_case_letters[year+1], current_row)] = (
            "=({col}{opex_total} - {col}{opex_tax_deductible}) + {col}{opex_tax_deductible} * (1 - {tax_rate}/100)"
            ).format(
            col=upper_case_letters[year+1],
            opex_total=opex_total_row,
            opex_tax_deductible=opex_tax_deductible_row,
            tax_rate=fed_tax_rate_cell,
        )
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    opex_after_tax_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Total Cash incentives, after-tax"
    for year in range(financial.analysis_years + 1):
        dcs["{}{}".format(upper_case_letters[year+1], current_row)] = (
            "= ({col}{taxed_incentives} * (1 - {tax_rate}/100)) + "  
            "{col}{total_nontaxable_cash_incentives_row}"
            ).format(
            col=upper_case_letters[year+1],
            taxed_incentives=total_taxable_cash_incentives_row,
            tax_rate=fed_tax_rate_cell,
            total_nontaxable_cash_incentives_row=total_nontaxable_cash_incentives_row
        )
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    incentives_after_tax_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Depreciation Tax Shield"
    for i in range(financial.analysis_years):
        dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = '={col}{row} * {tax_rate}/100'.format(
            col=upper_case_letters[i + 2], row=total_depreciation_row, tax_rate=fed_tax_rate_cell)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    depreciation_tax_shield_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Investment Tax Credit"
    for i in range(financial.analysis_years):
        dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = '={col}{row}'.format(
            col=upper_case_letters[i + 2],
            row=total_fed_itc_row,
        )
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    itc_row = current_row

    current_row += 1
    dcs['A{}'.format(current_row)] = "Free Cash Flow before income"
    dcs['B{}'.format(current_row)] = "={}+B{}".format(upfront_cost_cell, incentives_after_tax_row)

    for i in range(financial.analysis_years):
        dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = (
            "={col}{opex} + {col}{incentives} + {col}{depr} + {col}{itc}"
            ).format(
            col=upper_case_letters[i+2],
            opex=opex_after_tax_row,
            incentives=incentives_after_tax_row,
            depr=depreciation_tax_shield_row,
            itc=itc_row,
        )
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    optimal_fcf_row = current_row
    fill_border(dcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Discounted Cash Flow"
    dcs['B{}'.format(current_row)] = "={}+B{}".format(upfront_cost_cell, incentives_after_tax_row)
    for year in range(financial.analysis_years):
        dcs['{}{}'.format(upper_case_letters[year+2], current_row)] = (
            "={col}{fcf} / (1 + {disc_rate}/100)^{year}"
            ).format(
            col=upper_case_letters[year+2],
            fcf=optimal_fcf_row,
            disc_rate=discount_rate_cell,
            year=year+1,
        )
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    optimal_dcf_row = current_row
    fill_border(dcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)

    current_row += 1
    dcs['A{}'.format(current_row)] = "Optimal Life Cycle Cost"
    if financial.two_party_ownership:
        dcs['A{}'.format(current_row)] = "Developer Net Present Cost"
    dcs['B{}'.format(current_row)] = "=SUM(B{row}:{col}{row})".format(
        row=optimal_dcf_row,
        col=upper_case_letters[financial.analysis_years+1]
    )
    optimal_LCC_cell = "\'{}\'!B{}".format(developer_cashflow_sheet_name, current_row)
    make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    dcs['A{}'.format(current_row)].font = title_font
    dcs['B{}'.format(current_row)].font = title_font

    ####################################################################################################################
    # Income From Host
    ####################################################################################################################

    if financial.two_party_ownership:

        current_row += 1
        dcs['A{}'.format(current_row)] = "Capital Recovery Factor"
        dcs['B{}'.format(current_row)] = (
            "={discount_rate}/100 * POWER(1 + {discount_rate}/100, {years}) / "
            "(POWER(1 + {discount_rate}/100, {years}) - 1) / (1 - {tax_rate}/100)"
            ).format(
            discount_rate=discount_rate_cell,
            years=analysis_period_cell,
            tax_rate=fed_tax_rate_cell,
        )
        capital_recovery_factor_cell = 'B{}'.format(current_row)
        make_attribute_row(dcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='0.000', border=no_border)

        current_row += 1
        dcs['A{}'.format(current_row)] = "Income from Host"
        for year in range(financial.analysis_years):
            dcs['{}{}'.format(upper_case_letters[year + 2], current_row)] = (
                "=-{npc} * {crf}"
            ).format(
                npc=optimal_LCC_cell,
                crf=capital_recovery_factor_cell,
            )
        income_from_host_cell = "\'{}\'!C{}".format(developer_cashflow_sheet_name, current_row)
        make_attribute_row(dcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)

        current_row += 1
        dcs['A{}'.format(current_row)] = "Income from Host, after-tax"
        for year in range(financial.analysis_years):
            dcs['{}{}'.format(upper_case_letters[year + 2], current_row)] = (
                "={income} * (1 - {tax_rate}/100)"
            ).format(
                income=income_from_host_cell,
                tax_rate=fed_tax_rate_cell,
            )
        income_from_host_after_tax_cell = "\'{}\'!C{}".format(developer_cashflow_sheet_name, current_row)
        make_attribute_row(dcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        income_after_tax_row = current_row

        current_row += 1
        dcs['A{}'.format(current_row)] = "Discounted Income from Host"
        for year in range(financial.analysis_years):
            dcs['{}{}'.format(upper_case_letters[year + 2], current_row)] = (
                "={income} / (1 + {disc_rate}/100)^{year}"
            ).format(
                income=income_from_host_after_tax_cell,
                disc_rate=discount_rate_cell,
                year=year+1,
            )
        make_attribute_row(dcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        discounted_income_row = current_row

        current_row += 1
        dcs['A{}'.format(current_row)] = "NPV of Income from Host"
        dcs['B{}'.format(current_row)] = "=SUM(C{row}:{col}{row})".format(
            row=discounted_income_row,
            col=upper_case_letters[financial.analysis_years + 1]
        )
        make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                       number_format='#,##0', border=no_border)
        dcs['A{}'.format(current_row)].font = title_font
        dcs['B{}'.format(current_row)].font = title_font

        current_row += 1
        dcs['A{}'.format(current_row)] = "Free Cash Flow after income"
        dcs['B{}'.format(current_row)] = "=B{}".format(optimal_fcf_row)
        free_cashflow_cells = ["\'{}\'!{}{}".format(
                developer_cashflow_sheet_name, "B", current_row)]
        for i in range(financial.analysis_years):
            free_cashflow_cells.append("\'{}\'!{}{}".format(
                developer_cashflow_sheet_name, upper_case_letters[i + 2], current_row))
            dcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = (
                "={col}{fcf} + {col}{income}"
                ).format(
                col=upper_case_letters[i+2],
                fcf=optimal_fcf_row,
                income=income_after_tax_row,
            )
        make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        developer_fcf_row = current_row
        fill_border(dcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)

        current_row += 1
        dcs['A{}'.format(current_row)] = "Cumulative Free Cash Flow after income"
        cumulative_cashflow_cells = []
        for year in range(financial.analysis_years+1):
            cumulative_cashflow_cells.append("\'{}\'!{}{}".format(
                developer_cashflow_sheet_name, upper_case_letters[year + 1], current_row))
            if year == 0:
                dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = (
                    "={yr0_cashflow}"
                    ).format(
                        yr0_cashflow="\'{}\'!{}{}".format(
                        developer_cashflow_sheet_name, upper_case_letters[year + 1], current_row -1)
                    )
            else:                    
                dcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = (
                    "={yr_cashflow} + {prev_year_cashflow}"
                    ).format(
                        yr_cashflow="\'{}\'!{}{}".format(
                        developer_cashflow_sheet_name, upper_case_letters[year + 1], current_row -1),
                        prev_year_cashflow="\'{}\'!{}{}".format(
                        developer_cashflow_sheet_name, upper_case_letters[year], current_row)
                    )
        make_attribute_row(dcs, current_row, length=financial.analysis_years+2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        fill_border(dcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)

    ####################################################################################################################
    ####################################################################################################################
    # HOST / BAU CASH FLOW SHEET
    ####################################################################################################################
    ####################################################################################################################
    ####################################################################################################################
    ####################################################################################################################

    hcs = wb.get_sheet_by_name(host_cashflow_sheet_name)
    current_row = 2

    ##################################################################################################
    # Operating Year
    ##################################################################################################

    hcs['A{}'.format(current_row)] = "Operating Year"
    for year in range(financial.analysis_years + 1):
        hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = year
    make_title_row(hcs, current_row, length=financial.analysis_years+2)

    ####################################################################################################################
    # BAU/Host Operating Expenses
    ####################################################################################################################

    hcs['A{}'.format(current_row)] = "Operating Expenses"
    make_title_row(hcs, current_row, length=financial.analysis_years + 2)

    current_row += 1
    start_om_row = current_row
    hcs['A{}'.format(current_row)] = "Business as Usual electricity bill ($)"

    for year in range(1, financial.analysis_years + 1):
        hcs['{}{}'.format(upper_case_letters[year+1], current_row)] = \
            '=-{year_one_bau_bill} * (1 + {escalation_pct}/100)^{year}'.format(
                year_one_bau_bill=year_one_bau_bill_cell,
                escalation_pct=escalation_pct_cell,
                year=year,
            )
    make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    bau_bill_row = current_row

    current_row += 1
    hcs['A{}'.format(current_row)] = "Business as Usual export credits ($)"

    for year in range(1, financial.analysis_years + 1):
        hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = \
            '={year_one_bau_credits} * (1 + {escalation_pct}/100)^{year}'.format(
                year_one_bau_credits=year_one_bau_credits_cell,
                escalation_pct=escalation_pct_cell,
                year=year,
            )
    make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    bau_export_credit_row = current_row

    if not financial.two_party_ownership:  # Show BAU costs
        """
        BAU O&M costs do not matter in two party proforma. They are sunk costs. They are shown for single party model
        to account for BAU LCC. However, in the two party proforma we don't show optimal and BAU LCC's; instead we show
        the developer's costs (plus income from host benefit for break even) and the host's cost/benefits. The host's
        costs/benefits are the net bill savings and the payment to the developer.
        """
        current_row += 1
        hcs['A{}'.format(current_row)] = "Business as Usual boiler fuel bill ($)"

        for year in range(1, financial.analysis_years + 1):
            hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = \
                '=(-1 * ({boiler_fuel_bill_bau_cell} * (1 + {escalation_pct}/100)^{year}))'.format(
                    boiler_fuel_bill_bau_cell=boiler_fuel_bill_bau_cell,
                    escalation_pct=boiler_escalation_pct_cell,
                    year=year,
                )
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        bau_export_credit_row = current_row

        current_row += 1
        hcs['A{}'.format(current_row)] = "Operation and Maintenance (O&M)"
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)

        current_row += 1
        for i, pv in enumerate(pv_data):
            hcs['A{}'.format(current_row)] = "Existing {} cost in $/kW".format(pv['name'])
            hcs['A{}'.format(current_row)].alignment = one_tab_indent
            for year in range(1, financial.analysis_years + 1):
                hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = (
                    '=-{pv_om_cost_us_dollars_per_kw_cell} * (1 + {om_escalation_rate_cell}/100)^{year}'
                    ' * {pv_existing_kw_cell}'
                ).format(
                pv_om_cost_us_dollars_per_kw_cell=pv_cell_locations[i]["pv_om_cost_us_dollars_per_kw_cell"],
                om_escalation_rate_cell=om_escalation_rate_cell,
                year=year,
                pv_size_kw_cell=pv_cell_locations[i]["pv_size_kw_cell"],
                pv_existing_kw_cell=pv_cell_locations[i]["pv_existing_kw_cell"],
                )
            make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                               number_format='#,##0', border=no_border)
            current_row += 1

        hcs['A{}'.format(current_row)] = "Existing Generator fixed O&M cost"
        hcs['A{}'.format(current_row)].alignment = one_tab_indent
        for year in range(1, financial.analysis_years + 1):
            if gen_fuel_used_gal_bau == 0:
                # generator was not modeled because it could not run (eg. no outage, can't be used with grid)
                hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = 0
            else:
                hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = (
                    '=-{generator_om_cost} * (1 + {om_escalation_rate}/100)^{year} * {existing_gen_kw}'
                    ).format(
                    generator_om_cost=generator_om_cost_us_dollars_per_kw_cell,
                    om_escalation_rate=om_escalation_rate_cell,
                    year=year,
                    existing_gen_kw=generator_existing_kw_cell,
                )
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        """
        Generator could be used for outage or during grid connected times. Either way it is used every year in the 
        analysis period so do not use outage one time or every year for cost calculations (it only applies to outage cost
        calculations, which are done outside of REopt).
        TODO: test proforma with generator that can be used anytime (not just outage)
        """
        current_row += 1
        hcs['A{}'.format(current_row)] = "Existing Generator variable O&M cost"
        hcs['A{}'.format(current_row)].alignment = one_tab_indent
        for year in range(1, financial.analysis_years + 1):
            if gen_fuel_used_gal_bau == 0:
                # generator was not modeled because it could not run (eg. no outage, can't be used with grid)
                hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = 0
            else:
                hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = '=-{} * (1 + {}/100)^{}'.format(
                    generator.existing_gen_year_one_variable_om_cost_us_dollars or 0, om_escalation_rate_cell, year,
                    )
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        current_row += 1
        hcs['A{}'.format(current_row)] = "Existing Generator fuel cost ($)"
        hcs['A{}'.format(current_row)].alignment = one_tab_indent
        for year in range(1, financial.analysis_years + 1):
            hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = '=-{} * (1 + {}/100)^{}'.format(
                gen_fuel_used_cost_bau_cell, om_escalation_rate_cell, year)
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)

        current_row += 1
        hcs['A{}'.format(current_row)] = "Total operating expenses"

        for i in range(1, financial.analysis_years + 1):
            hcs['{}{}'.format(upper_case_letters[i + 1], current_row)] = '=SUM({col}{start_row}:{col}{end_row})'.format(
                col=upper_case_letters[i + 1], start_row=start_om_row, end_row=current_row - 1)
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        fill_border(hcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)
        bau_opex_total_row = current_row

        current_row += 1
        hcs['A{}'.format(current_row)] = "Tax deductible operating expenses"

        for year in range(1, financial.analysis_years + 1):
            hcs['{}{}'.format(upper_case_letters[year+1], current_row)] = (
                "=IF({fed_tax_rate} > 0, {col}{opex_total_row}, 0)"
            ).format(
                fed_tax_rate=fed_tax_rate_cell,
                opex_total_row=bau_opex_total_row,
                col=upper_case_letters[year+1]
            )
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        fill_border(hcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)
        bau_opex_tax_deductible_row = current_row

        current_row += 1
        current_row += 1
    else:  # show Host's costs
        current_row += 1
        hcs['A{}'.format(current_row)] = "Electricity bill with optimal system before export credits"

        for year in range(1, financial.analysis_years + 1):
            hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = \
                '=-{year_one_bill} * (1 + {escalation_pct}/100)^{year}'.format(
                    year_one_bill=year_one_bill_cell, year=year, escalation_pct=escalation_pct_cell
                )
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        bill_with_sys_row = current_row
    
        current_row += 1
        hcs['A{}'.format(current_row)] = "Export credits with optimal system"
        for year in range(1, financial.analysis_years + 1):
            hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = (
                '={year_one_credits} * (1 + {escalation_pct}/100)^{year}'
                ).format(
                    year_one_credits=year_one_credits_cell, year=year, escalation_pct=escalation_pct_cell)
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        export_credits_row = current_row

        current_row += 1
        hcs['A{}'.format(current_row)] = "Business as Usual boiler fuel bill ($)"
        for year in range(1, financial.analysis_years + 1):
            hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = (
                '= (-1 * ({fuel_bill} * (1 + {escalation_pct}/100)^{year}))'
                ).format(
                    fuel_bill=boiler_fuel_bill_bau_cell, year=year, escalation_pct=boiler_escalation_pct_cell)
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        boiler_bill_bau_row = current_row

        current_row += 1
        hcs['A{}'.format(current_row)] = "Boiler fuel bill with optimal systems ($)"
        for year in range(1, financial.analysis_years + 1):
            hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = (
                '=(-1 * ({fuel_bill} * (1 + {escalation_pct}/100)^{year}))'
                ).format(
                    fuel_bill=boiler_fuel_bill_cell, year=year, escalation_pct=boiler_escalation_pct_cell)
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        boiler_bill_row = current_row

        current_row += 1
        hcs['A{}'.format(current_row)] = "CHP fuel bill with optimal systems ($)"
        for year in range(1, financial.analysis_years + 1):
            hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = (
                '=(-1 * ({fuel_bill} * (1 + {escalation_pct}/100)^{year}))'
                ).format(
                    fuel_bill=chp_fuel_bill_cell, year=year, escalation_pct=chp_escalation_pct_cell)
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        chp_bill_row = current_row

        current_row += 1
        hcs['A{}'.format(current_row)] = "Payment to Developer ($)"
        for year in range(1, financial.analysis_years + 1):
            hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = "=-{}".format(income_from_host_cell)
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        developer_payment_row = current_row

        current_row += 1
        hcs['A{}'.format(current_row)] = "Existing Generator fuel cost ($)"
        for year in range(1, financial.analysis_years + 1):
            hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = '=-{} * (1 + {}/100)^{}'.format(
                gen_fuel_used_cost_bau_cell, om_escalation_rate_cell, year)
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        bau_gen_fuel_costs_row = current_row
        
        current_row += 1
        hcs['A{}'.format(current_row)] = "Generator fuel cost ($)"
        for year in range(1, financial.analysis_years + 1):
            hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = '=-{} * (1+{}/100)^{}'.format(
                diesel_fuel_used_cost_cell, om_escalation_rate_cell, year)
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        gen_fuel_costs_row = current_row

        current_row += 1
        hcs['A{}'.format(current_row)] = "Net Energy Costs"
        for year in range(1, financial.analysis_years + 1):
            hcs['{}{}'.format(upper_case_letters[year+1], current_row)] = (
                "=-{bau_bill} - {bau_export_credit} + {with_sys_bill} + {export_credits} + {developer_payment} "
                " - {bau_gen_fuel} + {gen_fuel}"
                " - {boiler_bill_bau} + {boiler_bill} + {chp_bill}"
                ).format(
                bau_bill="{}{}".format(upper_case_letters[year+1], bau_bill_row),
                bau_export_credit="{}{}".format(upper_case_letters[year+1], bau_export_credit_row),
                with_sys_bill="{}{}".format(upper_case_letters[year+1], bill_with_sys_row),
                export_credits="{}{}".format(upper_case_letters[year+1], export_credits_row),
                developer_payment="{}{}".format(upper_case_letters[year+1], developer_payment_row),
                gen_fuel="{}{}".format(upper_case_letters[year+1], gen_fuel_costs_row),
                bau_gen_fuel="{}{}".format(upper_case_letters[year+1], bau_gen_fuel_costs_row),
                chp_bill="{}{}".format(upper_case_letters[year+1], chp_bill_row),
                boiler_bill="{}{}".format(upper_case_letters[year+1], boiler_bill_row),
                boiler_bill_bau="{}{}".format(upper_case_letters[year+1], boiler_bill_bau_row)
            )
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        fill_border(hcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)
        net_elec_costs_row = current_row
        bau_opex_total_row = current_row

        current_row += 1
        hcs['A{}'.format(current_row)] = "Tax deductible Electricity Costs"

        for year in range(1, financial.analysis_years + 1):
            hcs['{}{}'.format(upper_case_letters[year+1], current_row)] = (
                "=IF({fed_tax_rate} > 0, {col}{opex_total_row}, 0)"
            ).format(
                fed_tax_rate=fed_tax_rate_cell,
                opex_total_row=net_elec_costs_row,
                col=upper_case_letters[year+1]
            )
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        fill_border(hcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)
        bau_opex_tax_deductible_row = current_row

        current_row += 1
        current_row += 1

    ####################################################################################################################
    # Existing PBI
    ####################################################################################################################

    if not financial.two_party_ownership:

        hcs['A{}'.format(current_row)] = "Production-based incentives (PBI)"
        make_title_row(hcs, current_row, length=financial.analysis_years + 2)

        current_row += 1
        for idx, pv in enumerate(pv_data):
            hcs['A{}'.format(current_row)] = "Existing {} Combined PBI".format(pv['name'])
            hcs['A{}'.format(current_row)].alignment = one_tab_indent

            for year in range(financial.analysis_years):
                hcs['{}{}'.format(upper_case_letters[year + 2], current_row)] = (
                    "=IF({year} < {pbi_year_limit}, "
                    "MIN({dol_per_kwh} * {pv_kwh}, {pbi_max} * (1 - {pv_degradation_rate}/100)^{year}), 0)"
                ).format(
                    year=year,
                    pbi_year_limit=pv_cell_locations[idx]["pv_pbi_years_cell"],
                    dol_per_kwh=pv_cell_locations[idx]["pv_pbi_cell"],
                    col=upper_case_letters[year + 2],
                    pv_kwh=pv_cell_locations[idx]['pv_production_series_bau'][year],
                    pbi_max=pv_cell_locations[idx]["pv_pbi_max_cell"],
                    pv_degradation_rate=pv_cell_locations[idx]["pv_degradation_rate_cell"],
                )
            make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                               number_format='#,##0', border=no_border)
            pv_cell_locations[idx]["existing_pv_pbi_row"] = current_row
            current_row += 1

        hcs['A{}'.format(current_row)] = "Total taxable cash incentives"

        for i in range(financial.analysis_years):
            pv_cells = list()
            for idx in range(len(pv_data)):  # could be multiple PVs
                pv_cells.append(
                    ('IF({pv_pbi_combined_tax_fed_cell}="Yes", {col}{pv_pbi_total_row}, 0)'
                     ).format(
                        col=upper_case_letters[i + 2],
                        pv_pbi_combined_tax_fed_cell=pv_cell_locations[idx]["pv_pbi_combined_tax_fed_cell"],
                        pv_pbi_total_row=pv_cell_locations[idx]["existing_pv_pbi_row"]
                    )
                )
            pv_string = '+'.join(pv_cells)

            hcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = (
                '={pv_string}'
            ).format(
                pv_string=pv_string,
            )
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        fill_border(hcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)
        bau_total_taxable_cash_incentives_row = current_row
        current_row += 1
        hcs['A{}'.format(current_row)] = "Total non-taxable cash incentives"

        for i in range(financial.analysis_years):
            pv_cells = list()
            for idx in range(len(pv_data)):  # could be multiple PVs
                pv_cells.append(
                    ('IF({pv_pbi_combined_tax_fed_cell}="No", {col}{pv_pbi_total_row}, 0)'
                     ).format(
                        col=upper_case_letters[i + 2],
                        pv_pbi_combined_tax_fed_cell=pv_cell_locations[idx]["pv_pbi_combined_tax_fed_cell"],
                        pv_pbi_total_row=pv_cell_locations[idx]["existing_pv_pbi_row"]
                    )
                )
            pv_string = '+'.join(pv_cells)

            hcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = (
                '={pv_string}'
            ).format(
                pv_string=pv_string,
            )
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        fill_border(hcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)
        bau_total_nontaxable_cash_incentives_row = current_row
        current_row += 1
        current_row += 1
    
    ####################################################################################################################
    # After-tax Cash Flows
    ####################################################################################################################

    hcs['A{}'.format(current_row)] = "Total Cash Flows"
    make_title_row(hcs, current_row, length=financial.analysis_years + 2)

    current_row += 1
    hcs['A{}'.format(current_row)] = "Net Operating expenses, after-tax"

    for year in range(1, financial.analysis_years + 1):
        hcs['{}{}'.format(upper_case_letters[year + 1], current_row)] = (
            "=({col}{opex_total} - {col}{opex_tax_deductible}) + {col}{opex_tax_deductible} * (1 - {tax_rate}/100)"
        ).format(
            col=upper_case_letters[year + 1],
            opex_total=bau_opex_total_row,
            opex_tax_deductible=bau_opex_tax_deductible_row,
            tax_rate=host_fed_tax_rate_cell,
        )
    make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    bau_opex_after_tax_row = current_row

    if financial.two_party_ownership:
        bau_fcf_row = current_row  # for discounted cash flow

    if not financial.two_party_ownership:
        current_row += 1
        hcs['A{}'.format(current_row)] = "Total Cash incentives, after-tax"
        for year in range(1, financial.analysis_years + 1):
            hcs["{}{}".format(upper_case_letters[year + 1], current_row)] = (
                "= ({col}{taxed_incentives} * (1 - {tax_rate}/100)) + "
                "{col}{total_nontaxable_cash_incentives_row}"
            ).format(
                col=upper_case_letters[year + 1],
                taxed_incentives=bau_total_taxable_cash_incentives_row,
                tax_rate=host_fed_tax_rate_cell,
                total_nontaxable_cash_incentives_row=bau_total_nontaxable_cash_incentives_row
            )
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        bau_incentives_after_tax_row = current_row

        current_row += 1
        hcs['A{}'.format(current_row)] = "Free Cash Flow"

        for i in range(financial.analysis_years):
            hcs['{}{}'.format(upper_case_letters[i + 2], current_row)] = (
                "={col}{opex} + {col}{incentives}"
            ).format(
                col=upper_case_letters[i + 2],
                opex=bau_opex_after_tax_row,
                incentives=bau_incentives_after_tax_row,
            )
        make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                           number_format='#,##0', border=no_border)
        bau_fcf_row = current_row
        fill_border(hcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)

    current_row += 1
    hcs['A{}'.format(current_row)] = "Discounted Cash Flow"

    for year in range(financial.analysis_years):
        hcs['{}{}'.format(upper_case_letters[year + 2], current_row)] = (
            "={col}{fcf} / (1 + {disc_rate}/100)^{year}"
        ).format(
            col=upper_case_letters[year + 2],
            fcf=bau_fcf_row,
            disc_rate=host_discount_rate_cell,
            year=year + 1,
        )
    make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    bau_dcf_row = current_row
    fill_border(hcs, range(financial.analysis_years + 2), current_row, border_top_and_bottom)

    current_row += 1
    hcs['A{}'.format(current_row)] = "BAU Life Cycle Cost"
    if financial.two_party_ownership:
        hcs['A{}'.format(current_row)] = "Host Net Present Value"
    hcs['B{}'.format(current_row)] = "=SUM(B{row}:{col}{row})".format(
        row=bau_dcf_row,
        col=upper_case_letters[financial.analysis_years + 1]
    )
    bau_LCC_cell = "\'{}\'!B{}".format(host_cashflow_sheet_name, current_row)
    make_attribute_row(hcs, current_row, length=financial.analysis_years + 2, alignment=right_align,
                       number_format='#,##0', border=no_border)
    hcs['A{}'.format(current_row)].font = title_font
    hcs['B{}'.format(current_row)].font = title_font



    ####################################################################################################################
    ####################################################################################################################
    # INPUTS AND OUTPUTS - LCC's, NPV, IRR, LCOE and Simple Payback
    ####################################################################################################################
    ####################################################################################################################
    ####################################################################################################################
    ####################################################################################################################

    current_row = 5
    ws['D{}'.format(current_row)] = "RESULTS"
    make_title_row(ws, current_row, length=2, offset=3)

    current_row += 1
    ws['D{}'.format(current_row)] = "Business as usual LCC, $"
    ws['E{}'.format(current_row)] = "=-{}".format(bau_LCC_cell)
    if financial.two_party_ownership:
        ws['D{}'.format(current_row)] = "Annual payment to developer, $"
        ws['E{}'.format(current_row)] = "={}".format(income_from_host_cell)
    make_attribute_row(ws, current_row, length=2, offset=3, number_format="#,##0")
    fill_cols(ws, range(4, 5), current_row, calculated_fill)

    current_row += 1
    ws['D{}'.format(current_row)] = "Optimal LCC, $"
    if financial.two_party_ownership:
        ws['D{}'.format(current_row)] = "Developer NPC, $"
    ws['E{}'.format(current_row)] = "=-{}".format(optimal_LCC_cell)
    make_attribute_row(ws, current_row, length=2, offset=3, number_format="#,##0")
    fill_cols(ws, range(4, 5), current_row, calculated_fill)
    ws['F{}'.format(current_row)] = (
        'NOTE: A negative LCC indicates a profit (for example when production based incentives are greater than costs.'
    )

    current_row += 1
    if not financial.two_party_ownership:
        ws['D{}'.format(current_row)] = "NPV, $"
        ws['E{}'.format(current_row)] = (
            "=E{} - E{}"
        ).format(
            current_row - 2,
            current_row - 1,
        )
    else:
        ws['D{}'.format(current_row)] = "Host's NPV, $"
        ws['E{}'.format(current_row)] = "={}".format(bau_LCC_cell)
    npv_cell = 'E{}'.format(current_row)
    ws['F{}'.format(current_row)] = (
        'NOTE: This NPV can differ slightly (<1%) from the Webtool/API results due to rounding and the tolerance in the'
        ' optimizer.'
    )
    make_attribute_row(ws, current_row, length=2, offset=3, number_format="#,##0")
    fill_cols(ws, range(4, 5), current_row, calculated_fill)


    if not financial.two_party_ownership:
        current_row += 1
        irr_row = current_row
        ws['D{}'.format(current_row)] = "IRR, %"
        ws['E{}'.format(current_row)] = (
            '=IF({npv_cell}=0,"",IRR({col1}{row}:{colN}{row}, {discount_rate}/100))'
            ).format(npv_cell=npv_cell,
            col1="B",
            colN=upper_case_letters[financial.analysis_years + 1],
            row=free_cash_flow_row,
            discount_rate=discount_rate_cell,
        )
        make_attribute_row(ws, current_row, length=2, offset=3, number_format="0.00%")
        fill_cols(ws, range(4, 5), current_row, calculated_fill)

        
        current_row = free_cash_flow_row
        ws['A{}'.format(current_row)] = "Free Cash Flow"
 
        free_cashflow_cells = []
        for year in range(financial.analysis_years+1):
            free_cashflow_cells.append("\'{}\'!{}{}".format(
                inandout_sheet_name, upper_case_letters[year + 1], current_row))
            ws['{}{}'.format(upper_case_letters[year + 1], current_row)] = (
                "={optimal_fcf} - {bau_fcf}"
            ).format(
                optimal_fcf="\'{}\'!{}{}".format(
                    developer_cashflow_sheet_name, upper_case_letters[year + 1], optimal_fcf_row),
                bau_fcf="\'{}\'!{}{}".format(
                    host_cashflow_sheet_name, upper_case_letters[year + 1], bau_fcf_row),
            )
        fill_in_annual_values(current_row)

        current_row += 1
        ws['A{}'.format(current_row)] = "Cumulative Free Cash Flow"
        cumulative_cashflow_cells = []
        for year in range(financial.analysis_years+1):
            cumulative_cashflow_cells.append("\'{}\'!{}{}".format(
                inandout_sheet_name, upper_case_letters[year + 1], current_row))
            if year == 0:
                ws['{}{}'.format(upper_case_letters[year + 1], current_row)] = (
                    "={yr0_cashflow}"
                    ).format(
                        yr0_cashflow="\'{}\'!{}{}".format(
                        inandout_sheet_name, upper_case_letters[year + 1], current_row -1)
                    )
            else:                    
                ws['{}{}'.format(upper_case_letters[year + 1], current_row)] = (
                    "={yr_cashflow} + {prev_year_cashflow}"
                    ).format(
                        yr_cashflow="\'{}\'!{}{}".format(
                        inandout_sheet_name, upper_case_letters[year + 1], current_row -1),
                        prev_year_cashflow="\'{}\'!{}{}".format(
                        inandout_sheet_name, upper_case_letters[year], current_row)
                    )
        fill_in_annual_values(current_row)

        simple_payback = ""
        simple_payback += '=IF({}=0,"", IF({} < 0, "Over {}", '.format(cumulative_cashflow_cells[0], cumulative_cashflow_cells[-1], financial.analysis_years)
        
        simple_payback += "SUM("
        simple_payback_terms = []
        for year in range(1, financial.analysis_years+1):
            simple_payback_terms.append(
                    "IF({current_cumulative}<0,1,"
                    "IF({prev_cumulative}>0,0,"
                        "-{prev_cumulative}/{current_free_cashflow}))".format(
                            current_cumulative=cumulative_cashflow_cells[year],
                            prev_cumulative=cumulative_cashflow_cells[year-1],
                            current_free_cashflow=free_cashflow_cells[year]
                            ))
        simple_payback += ', '.join(simple_payback_terms) + ")))"

        current_row = irr_row + 1
        ws['D{}'.format(current_row)] = "Simple Payback Period, years"
        ws['E{}'.format(current_row)] = simple_payback
        make_attribute_row(ws, current_row, length=2, offset=3, number_format="0.00")
        fill_cols(ws, range(4, 5), current_row, calculated_fill)
    else:
        current_row += 1
        irr_row = current_row
        ws['D{}'.format(current_row)] = "Developer IRR, %"
        ws['E{}'.format(current_row)] = (
             "=IF({npv_cell}=0,"",IRR({optimal_fcf_B}:{optimal_fcf_N},{discount_rate_cell}/100))"
            ).format(npv_cell = npv_cell,
            optimal_fcf_B="\'{}\'!B{}".format(
                developer_cashflow_sheet_name, developer_fcf_row),
            optimal_fcf_N="\'{}\'!{}{}".format(
                developer_cashflow_sheet_name, upper_case_letters[financial.analysis_years + 1], developer_fcf_row),
            discount_rate_cell=discount_rate_cell,
        )
        make_attribute_row(ws, current_row, length=2, offset=3, number_format="0.00%")
        fill_cols(ws, range(4, 5), current_row, calculated_fill)

        simple_payback = ""
        simple_payback += '=IF({} < 0, "Over {}", '.format(cumulative_cashflow_cells[-1], financial.analysis_years)
        
        simple_payback += "SUM("
        simple_payback_terms = []
        for year in range(1, financial.analysis_years+1):
            simple_payback_terms.append(
                    "IF({current_cumulative}<0,1,"
                    "IF({prev_cumulative}>0,0,"
                        "-{prev_cumulative}/{current_free_cashflow}))".format(
                            current_cumulative=cumulative_cashflow_cells[year],
                            prev_cumulative=cumulative_cashflow_cells[year-1],
                            current_free_cashflow=free_cashflow_cells[year]
                            ))
        simple_payback += ', '.join(simple_payback_terms) + "))"

        current_row = irr_row + 1
        ws['D{}'.format(current_row)] = "Developer Simple Payback Period, years"
        ws['E{}'.format(current_row)] = simple_payback
        make_attribute_row(ws, current_row, length=2, offset=3, number_format="0.00")
        fill_cols(ws, range(4, 5), current_row, calculated_fill)


    


    def two_party_factor(pwf_offtaker_cell, offtaker_tax_pct_cell, pwf_owner_cell, owner_tax_pct_cell):
        """
        :param pwf_offtaker_cell: present worth factor offtacker cell in spreadsheet
        :param offtaker_tax_pct_cell: offtacker tax percent cell in spreadsheet
        :param pwf_owner_cell: present worth factor owner cell in spreadsheet
        :param owner_tax_pct_cell: owner tax percent cell in spreadsheet
        :return: two_party_factor
        """
        tpf =   ("ROUND((({pwf_offtaker_cell}/100) * (1 - ({offtaker_tax_pct_cell}/100))) / "
                "(({pwf_owner_cell}/100) * (1 - ({owner_tax_pct_cell}/100))),4)".format(
                    pwf_offtaker_cell=pwf_offtaker_cell,
                    offtaker_tax_pct_cell=offtaker_tax_pct_cell,
                    pwf_owner_cell=pwf_owner_cell,
                    owner_tax_pct_cell=owner_tax_pct_cell)
                    )
        return tpf

    


    def degradation_factor(analysis_period, rate_degradation_cell):
        """
        :param analysis_period_cell: analysis period cell, years
        :param rate_degradation_cell: degredation rate cell
        :return: energy levelization factor
        """

        if analysis_period == 0:
            return '0'
        if analysis_period == 1:
            return '1'
        terms = ['1']
        for yr in range(1, int(analysis_period)):
            terms.append('(1 - ({rate_degradation}/100))^{yr}'.format(rate_degradation=rate_degradation_cell, yr=yr))
        return '=ROUND(AVERAGE({terms}),4)'.format(terms=','.join(terms))


    def present_worth(analysis_period_cell, esc_cell,  disc_cell):
        """
        :param analysis_period_cell: analysis period cell, years
        :param esc_cell: escalation rate cell
        :param disc_cell: discount rate cell 
        :return: present worth factor equation
        """
        if esc_cell != disc_cell:
            x = '((1 + {esc_cell}/100) / (1 + {disc_cell}/100))'.format(esc_cell=esc_cell, disc_cell=disc_cell)
            result = 'ROUND({x} * (1 - {x}^{analysis_period_cell}) / (1 - {x}),4)'.format(
                x=x, analysis_period_cell=analysis_period_cell)
            return result
        return analysis_period_cell

    for idx, pv in enumerate(pv_data):
        
        if (pv['pv_installed_kw'] or 0) > 0:
            npv_om = '-1*(NPV({}/100,{}))'.format(discount_rate_cell, 
                pv_cell_locations[idx]["new_o_and_m_cashflow_cells"][0] + ":" + pv_cell_locations[idx]["new_o_and_m_cashflow_cells"][-1])

            npv_pbi = 'NPV({}/100,{})'.format(discount_rate_cell, 
                pv_cell_locations[idx]["new_pv_pbi_series"][0] + ":" + pv_cell_locations[idx]["new_pv_pbi_series"][-1])

            npv_macrs_benefit = 'NPV({}/100,{})'.format(discount_rate_cell, 
                pv_cell_locations[idx]["pv_federal_itc_amount_cell"])

            npv_tax_benefits = 'NPV({}/100,{})'.format(discount_rate_cell, 
                pv_cell_locations[idx]["pv_income_tax_benefits"][0] + ":" + pv_cell_locations[idx]["pv_income_tax_benefits"][-1])

            npv_total_pv = 'NPV({}/100,{})'.format(discount_rate_cell, 
                pv_cell_locations[idx]["pv_production_series"][0] + ":" + pv_cell_locations[idx]["pv_production_series"][-1])

            npv_existing_pv = 'NPV({}/100,{})'.format(discount_rate_cell, 
                pv_cell_locations[idx]["pv_production_series_bau"][0] + ":" + pv_cell_locations[idx]["pv_production_series_bau"][-1])

            pv_lcoe = ("=ROUND((({total_installed_cost} + {npv_om} - "
                        "SUM({cbi},{ibi},{npv_pbi},{npv_macrs_benefit},{npv_tax_benefits})) "
                ") / (({npv_total_pv} - {npv_existing_pv})), 4)"
                       ).format(
                total_installed_cost=pv_cell_locations[idx]["pv_cost_cell"],
                npv_om=npv_om,
                cbi=pv_cell_locations[idx]["pv_total_cbi_cell"],
                ibi=pv_cell_locations[idx]['pv_total_ibi_cell'],
                npv_pbi=npv_pbi, 
                year_one_energy=pv_cell_locations[idx]["pv_energy_cell"], 
                existing_energy_cell=pv_cell_locations[idx]["pv_energy_bau_cell"], 
                npv_macrs_benefit=npv_macrs_benefit,
                npv_tax_benefits=npv_tax_benefits,
                npv_total_pv=npv_total_pv,
                npv_existing_pv=npv_existing_pv
                )
            ws[pv_cell_locations[idx]["pv_lcoe_cell"]] = pv_lcoe

            
    if (wind.size_kw or 0) > 0:
        npv_wind_om = '(-1 * NPV({}/100,{}))'.format(discount_rate_cell, 
                wind_o_and_m_cashflow_cells[0] + ":" + wind_o_and_m_cashflow_cells[-1])

        npv_wind_pbi = 'NPV({}/100,{})'.format(discount_rate_cell, 
            wind_pbi_series[0] + ":" + wind_pbi_series[-1])

        npv_macrs_benefit = 'NPV({}/100,{})'.format(discount_rate_cell, 
                wind_federal_itc_amount_cell)

        npv_tax_benefits = 'NPV({}/100,{})'.format(discount_rate_cell, 
            wind_income_tax_benefits[0] + ":" + wind_income_tax_benefits[-1])

        npv_total_wind = 'NPV({}/100,{})'.format(discount_rate_cell, 
            wind_annual_kwh_cells[0] + ":" + wind_annual_kwh_cells[-1])

        wind_lcoe = ("=ROUND(" 
        "({total_cost} + {npv_om} - SUM({cbi},{ibi},{npv_pbi},{npv_macrs_benefit},{npv_tax_benefits}))" 
        "/ {npv_total_wind}" 
        ",4)".format(
                total_cost=wind_cost_cell,  npv_om=npv_wind_om,
                cbi=wind_total_cbi_cell, ibi=wind_total_ibi_cell,
                npv_pbi=npv_wind_pbi,
                npv_macrs_benefit=npv_macrs_benefit,
                npv_tax_benefits=npv_tax_benefits,
                npv_total_wind=npv_total_wind,
            )
        )
        ws[wind_lcoe_cell] = wind_lcoe

    ####################################################################################################################
    ####################################################################################################################
    wb.save(output_file_path)