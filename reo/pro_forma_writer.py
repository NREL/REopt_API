import os
import math
import numpy as np
from openpyxl import load_workbook

# logging
from log_levels import log

class ProForma(object):

    file_template = "REoptCashFlowTemplate.xlsm"
    file_output = "ProForma.xlsm"

    def __init__(self, path_templates, path_output, econ, results):

        # paths
        self.file_template = os.path.join(path_templates, self.file_template)
        self.file_output = os.path.join(path_output, self.file_output)

        # data
        self.econ = econ

        # supported techs for cash flow
        self.techs = ['PV', 'BATT']

        # system sizes
        self.pv_kw = results['pv_kw']
        self.batt_kw = results['batt_kw']
        self.batt_kwh = results['batt_kwh']

        if self.pv_kw is None:
            self.pv_kw = 0
        if self.batt_kw is None:
            self.batt_kw = 0
        if self.batt_kwh is None:
            self.batt_kwh = 0

        # input to template
        self.pv_installed_cost = self.pv_kw * econ.pv_cost
        self.battery_installed_cost = self.batt_kw * econ.batt_cost_kw + self.batt_kwh * econ.batt_cost_kwh
        self.total_capital_costs = self.pv_installed_cost + self.battery_installed_cost

        self.year_one_bill = results['year_one_demand_cost'] + results['year_one_energy_cost']
        self.year_one_exports = results['year_one_export_benefit']
        self.year_one_bill_bau = results['year_one_demand_cost_bau'] + results['year_one_energy_cost_bau']
        self.year_one_energy_produced = results['year_one_energy_produced']
        self.year_one_savings = self.year_one_bill_bau - self.year_one_bill

        # no state taxes for now
        self.fed_tax_owner = self.econ.owner_tax_rate

        # Unit test outputs
        self.fed_pv_depr_basis_calc = 0
        self.fed_pv_itc_basis_calc = 0
        self.fed_batt_depr_basis_calc = 0
        self.fed_batt_itc_basis_calc = 0
        self.pre_tax_cash_flow = list()
        self.direct_cash_incentives = list()
        self.federal_tax_liability = list()
        self.bill_with_sys = list()
        self.bill_bau = list()
        self.exports_with_sys = list()
        self.total_operating_expenses = list()
        self.after_tax_annual_costs = list()
        self.after_tax_value = list()
        self.after_tax_cash_flow = list()
        self.net_annual_cost_without_sys = list()
        self.net_annual_cost_with_sys = list()

        # ProForma outputs
        self.irr = 0
        self.npv = 0
        self.lcc = 0
        self.lcc_bau = 0

        # Incentives data structure
        self.incentives = dict()
        for t in self.techs:

            self.incentives[t] = dict()

            if t == 'PV':

                self.incentives[t]['tech_cost'] = self.pv_installed_cost
                self.incentives[t]['tech_size'] = self.pv_kw

                # MACRS
                self.incentives[t]['macrs_schedule'] = self.econ.pv_macrs_schedule
                self.incentives[t]['bonus_fraction'] = self.econ.pv_macrs_bonus_fraction
                self.incentives[t]['macrs_itc_reduction'] = self.econ.pv_macrs_itc_reduction

                # tax credits
                self.incentives[t]['itc_fed_percent'] = self.econ.pv_itc_federal
                self.incentives[t]['itc_fed_percent_maxvalue'] = self.econ.pv_itc_federal_max

                # cash incentives (need to rename state, util from ITC)
                self.incentives[t]['ibi_sta_percent'] = self.econ.pv_itc_state
                self.incentives[t]['ibi_sta_percent_maxvalue'] = self.econ.pv_itc_state_max
                self.incentives[t]['ibi_uti_percent'] = self.econ.pv_itc_utility
                self.incentives[t]['ibi_uti_percent_maxvalue'] = self.econ.pv_itc_utility_max

                # capacity based incentives
                self.incentives[t]['cbi_fed_amount'] = self.econ.pv_rebate_federal
                self.incentives[t]['cbi_fed_maxvalue'] = self.econ.pv_rebate_federal_max
                self.incentives[t]['cbi_sta_amount'] = self.econ.pv_rebate_state
                self.incentives[t]['cbi_sta_maxvalue'] = self.econ.pv_rebate_state_max
                self.incentives[t]['cbi_uti_amount'] = self.econ.pv_rebate_utility
                self.incentives[t]['cbi_uti_maxvalue'] = self.econ.pv_rebate_utility_max

                # production based incentives
                self.incentives[t]['pbi_combined_amount'] = self.econ.pv_pbi
                self.incentives[t]['pbi_combined_maxvalue'] = self.econ.pv_pbi_max
                self.incentives[t]['pbi_combined_term'] = self.econ.pv_pbi_years

            elif t == 'BATT':

                self.incentives[t]['tech_cost'] = self.battery_installed_cost
                self.incentives[t]['tech_size'] = self.batt_kw

                # MACRS
                self.incentives[t]['macrs_schedule'] = self.econ.batt_macrs_schedule
                self.incentives[t]['bonus_fraction'] = self.econ.batt_macrs_bonus_fraction
                self.incentives[t]['macrs_itc_reduction'] = self.econ.batt_macrs_itc_reduction

                # tax credits
                self.incentives[t]['itc_fed_percent'] = self.econ.batt_itc_federal
                self.incentives[t]['itc_fed_percent_maxvalue'] = self.econ.batt_itc_federal_max

                # cash incentives (need to rename state, util from ITC)
                self.incentives[t]['ibi_sta_percent'] = self.econ.batt_itc_state
                self.incentives[t]['ibi_sta_percent_maxvalue'] = self.econ.batt_itc_state_max
                self.incentives[t]['ibi_uti_percent'] = self.econ.batt_itc_utility
                self.incentives[t]['ibi_uti_percent_maxvalue'] = self.econ.batt_itc_utility_max

                # capacity based incentives
                self.incentives[t]['cbi_fed_amount'] = self.econ.batt_rebate_federal
                self.incentives[t]['cbi_fed_maxvalue'] = self.econ.batt_rebate_federal_max
                self.incentives[t]['cbi_sta_amount'] = self.econ.batt_rebate_state
                self.incentives[t]['cbi_sta_maxvalue'] = self.econ.batt_rebate_state_max
                self.incentives[t]['cbi_uti_amount'] = self.econ.batt_rebate_utility
                self.incentives[t]['cbi_uti_maxvalue'] = self.econ.batt_rebate_utility_max

                # production based incentives
                self.incentives[t]['pbi_combined_amount'] = 0
                self.incentives[t]['pbi_combined_maxvalue'] = 0
                self.incentives[t]['pbi_combined_term'] = 0


            # tax credits reduce depreciation and ITC basis
            self.incentives[t]['itc_fed_percent_deprbas_fed'] = True

            # cash incentive is taxable
            self.incentives[t]['ibi_sta_percent_tax_fed'] = True
            self.incentives[t]['ibi_uti_percent_tax_fed'] = True

            self.incentives[t]['cbi_fed_tax_fed'] = True
            self.incentives[t]['cbi_sta_tax_fed'] = True
            self.incentives[t]['cbi_uti_tax_fed'] = True

            self.incentives[t]['pbi_combined_tax_fed'] = True

            # incentive reduces depreciation and ITC basis
            self.incentives[t]['ibi_sta_percent_deprbas_fed'] = False
            self.incentives[t]['ibi_uti_percent_deprbas_fed'] = False

            self.incentives[t]['cbi_fed_deprbas_fed'] = False
            self.incentives[t]['cbi_sta_deprbas_fed'] = False
            self.incentives[t]['cbi_uti_deprbas_fed'] = False

            # Assume state MACRS the same as fed
            self.incentives[t]['state_depreciation_equals_federal'] = True

            # calculated values
            self.incentives[t]['federal_itc_basis'] = self.federal_itc_basis(t)
            self.incentives[t]['federal_depreciation_basis'] = self.federal_depreciation_basis(t)


    def get_irr(self):
        return self.irr

    def get_npv(self):
        return round(self.npv)

    def get_lcc(self):
        return round(self.lcc)

    def get_lcc_bau(self):
        return round(self.lcc_bau)

    def update_template(self):

        sheet_io = "Inputs and Outputs"

        # Open file for reading
        wb = load_workbook(self.file_template, read_only=False, keep_vba=True)

        # Open Inputs Sheet
        ws = wb.get_sheet_by_name(sheet_io)

        # System Design
        ws['B3'] = self.pv_kw
        ws['B4'] = self.econ.pv_degradation_rate * 100
        ws['B5'] = self.batt_kw
        ws['B6'] = self.batt_kwh

        # Year 1 Results
        ws['B9'] = self.year_one_bill_bau
        ws['B10'] = self.year_one_bill
        ws['B11'] = self.year_one_exports
        ws['B12'] = self.year_one_energy_produced

        # System Costs
        ws['B15'] = self.total_capital_costs
        ws['B16'] = self.pv_installed_cost
        ws['B17'] = self.battery_installed_cost
        ws['B19'] = self.econ.pv_om
        ws['B20'] = self.econ.batt_replacement_cost_kw
        ws['B21'] = self.econ.batt_replacement_year_kw
        ws['B22'] = self.econ.batt_replacement_cost_kwh
        ws['B23'] = self.econ.batt_replacement_year_kwh

        # Analysis Parameters
        ws['B31'] = self.econ.analysis_period
        ws['B32'] = self.econ.rate_inflation * 100
        ws['B33'] = self.econ.rate_escalation * 100
        ws['B35'] = self.econ.owner_discount_rate * 100

        # Tax rates
        ws['B39'] = self.fed_tax_owner * 100

        # PV Tax Credits and Incentives
        ws['B44'] = self.econ.pv_itc_federal * 100
        ws['C44'] = self.econ.pv_itc_federal_max
        ws['B49'] = self.econ.pv_itc_state * 100
        ws['C49'] = self.econ.pv_itc_state_max
        ws['B50'] = self.econ.pv_itc_utility * 100
        ws['C50'] = self.econ.pv_itc_utility_max
        ws['B52'] = self.econ.pv_rebate_federal * 0.001
        ws['C52'] = self.econ.pv_rebate_federal_max
        ws['B53'] = self.econ.pv_rebate_state * 0.001
        ws['C53'] = self.econ.pv_rebate_state_max
        ws['B54'] = self.econ.pv_rebate_utility * 0.001
        ws['C54'] = self.econ.pv_rebate_utility_max
        ws['B56'] = self.econ.pv_pbi
        ws['C56'] = self.econ.pv_pbi_max
        ws['E56'] = self.econ.pv_pbi_years
        ws['F56'] = self.econ.pv_pbi_system_max

        # Battery Tax Credits and Incentives
        ws['B61'] = self.econ.batt_itc_federal * 100
        ws['C61'] = self.econ.batt_itc_federal_max
        ws['B66'] = self.econ.batt_itc_state * 100
        ws['C66'] = self.econ.batt_itc_state_max
        ws['B67'] = self.econ.batt_itc_utility * 100
        ws['C67'] = self.econ.batt_itc_utility_max
        ws['B69'] = self.econ.batt_rebate_federal
        ws['C69'] = self.econ.batt_rebate_federal_max
        ws['B70'] = self.econ.batt_rebate_state
        ws['C70'] = self.econ.batt_rebate_state_max
        ws['B71'] = self.econ.batt_rebate_utility
        ws['C71'] = self.econ.batt_rebate_utility_max

        # Depreciation
        if self.econ.pv_macrs_schedule > 0:
            ws['B74'] = self.econ.pv_macrs_schedule
            ws['B75'] = self.econ.pv_macrs_bonus_fraction
        if self.econ.batt_macrs_schedule > 0:
            ws['C74'] = self.econ.batt_macrs_schedule
            ws['C75'] = self.econ.batt_macrs_bonus_fraction

        # Save
        wb.save(self.file_output)

    def compute_cashflow(self):

        n_cols = self.econ.analysis_period + 1

        # row_vectors to match template spreadsheet
        zero_list = [0] * n_cols
        annual_energy = list(zero_list)
        bill_without_system = list(zero_list)
        bill_with_system = list(zero_list)
        exports_with_system = list(zero_list)
        value_of_savings = list(zero_list)
        o_and_m_capacity_cost = list(zero_list)
        batt_kw_replacement_cost = list(zero_list)
        batt_kwh_replacement_cost = list(zero_list)
        total_operating_expenses = list(zero_list)
        tech_operating_expenses = dict()
        total_deductible_expenses = list(zero_list)
        debt_amount = list(zero_list)
        pre_tax_cash_flow = list(zero_list)
        total_cash_incentives = list(zero_list)
        total_production_based_incentives = list(zero_list)
        federal_taxable_income_before_deductions = list(zero_list)
        federal_depreciation_amount = list(zero_list)
        federal_total_deductions = list(zero_list)
        federal_income_tax = list(zero_list)
        federal_tax_liability = list(zero_list)
        after_tax_annual_costs = list(zero_list)
        after_tax_value_of_energy = list(zero_list)
        after_tax_cash_flow = list(zero_list)
        net_annual_costs_with_system = list(zero_list)
        net_annual_costs_without_system = list(zero_list)

        # incentives, tax credits, depreciation
        federal_itc = self.federal_itc_total()

        # year 0 initializations
        total_cash_incentives[0] = sum([self.total_cash_incentives(tech) for tech in self.techs])
        debt_amount[0] = self.total_capital_costs - total_cash_incentives[0]
        pre_tax_cash_flow[0] = -debt_amount[0]
        after_tax_annual_costs[0] = total_cash_incentives[0] - self.total_capital_costs
        after_tax_cash_flow[0] = after_tax_annual_costs[0]
        net_annual_costs_with_system[0] = after_tax_annual_costs[0]
        net_annual_costs_without_system[0] = 0

        # year 1 initializations
        annual_energy[1] = self.year_one_energy_produced
        bill_with_system[1] = self.year_one_bill
        bill_without_system[1] = self.year_one_bill_bau
        exports_with_system[1] = self.year_one_exports
        value_of_savings[1] = self.year_one_savings - self.year_one_exports
        o_and_m_capacity_cost[1] = self.econ.pv_om * self.pv_kw
        inflation_modifier = 1 + self.econ.rate_inflation + self.econ.rate_escalation
        federal_taxable_income_before_deductions[1] = sum([self.federal_taxable_income_before_deductions(tech) for tech in self.techs])

        for year in range(1, n_cols):

            inflation_modifier_n = inflation_modifier ** (year-1)
            degradation_modifier = 1 - self.econ.pv_degradation_rate

            if year > 1:

                # Annual energy
                annual_energy[year] = annual_energy[year - 1] * (1 - self.econ.pv_degradation_rate)

                # Bill savings
                bill_without_system[year] = bill_without_system[year - 1] * inflation_modifier
                bill_with_system[year] = bill_with_system[year - 1] * inflation_modifier
                exports_with_system[year] = exports_with_system[year - 1] * inflation_modifier * degradation_modifier
                value_of_savings[year] = bill_without_system[year] - (bill_with_system[year] + exports_with_system[year])

                # Operating Expenses
                o_and_m_capacity_cost[year] = o_and_m_capacity_cost[year - 1] * inflation_modifier

            if self.econ.batt_replacement_year_kw == year:
                batt_kw_replacement_cost[year] = self.econ.batt_replacement_cost_kw * self.batt_kw * inflation_modifier_n

            if self.econ.batt_replacement_year_kwh == year:
                batt_kwh_replacement_cost[year] = self.econ.batt_replacement_cost_kwh * self.batt_kwh * inflation_modifier_n

            tech_operating_expenses['PV'] = o_and_m_capacity_cost[year]
            tech_operating_expenses['BATT'] = batt_kw_replacement_cost[year] + batt_kwh_replacement_cost[year]
            total_operating_expenses[year] = sum(tech_operating_expenses.values())

            # Tax deductible operating expenses
            for tech in self.techs:
                total_deductible_expenses[year] += tech_operating_expenses[tech]

            # Pre-tax cash flow
            pre_tax_cash_flow[year] = -(total_operating_expenses[year])

            # Direct cash incentives
            total_production_based_incentives[year] = self.pbi_calculate('PV', year, annual_energy[year])
            total_cash_incentives[year] += total_production_based_incentives[year]

            # Federal income tax
            if self.incentives['PV']['pbi_combined_tax_fed']:
                federal_taxable_income_before_deductions[year] = total_cash_incentives[year]

            federal_itc_to_apply = 0
            if year == 1:
                federal_itc_to_apply = federal_itc

            federal_depreciation_amount[year] = self.federal_depreciation_amount_total(year)
            federal_total_deductions[year] = total_deductible_expenses[year] + federal_depreciation_amount[year]
            federal_income_tax[year] = (federal_taxable_income_before_deductions[year]-federal_total_deductions[year]) * self.fed_tax_owner
            federal_tax_liability[year] = -federal_income_tax[year] + federal_itc_to_apply

            # After tax calculation
            after_tax_annual_costs[year] = pre_tax_cash_flow[year] + \
                                           total_cash_incentives[year] + \
                                           federal_tax_liability[year]

            after_tax_value_of_energy[year] = value_of_savings[year] * (1 - self.fed_tax_owner)
            after_tax_cash_flow[year] = after_tax_annual_costs[year] + after_tax_value_of_energy[year]

            net_annual_costs_without_system[year] = -bill_without_system[year] * (1 - self.fed_tax_owner)
            net_annual_costs_with_system[year] = after_tax_annual_costs[year] - (bill_with_system[year] + exports_with_system[year]) * (1 - self.fed_tax_owner)

        # Additional Unit test outputs
        self.fed_pv_depr_basis_calc = self.incentives['PV']['federal_depreciation_basis']
        self.fed_pv_itc_basis_calc = self.incentives['PV']['federal_itc_basis']
        self.fed_batt_depr_basis_calc = self.incentives['BATT']['federal_depreciation_basis']
        self.fed_batt_itc_basis_calc = self.incentives['BATT']['federal_itc_basis']
        self.pre_tax_cash_flow = pre_tax_cash_flow
        self.direct_cash_incentives = total_cash_incentives
        self.federal_tax_liability = federal_tax_liability
        self.bill_with_sys = bill_with_system
        self.exports_with_sys = exports_with_system
        self.bill_bau = bill_without_system
        self.total_operating_expenses = total_operating_expenses
        self.after_tax_annual_costs = after_tax_annual_costs
        self.after_tax_value = after_tax_value_of_energy
        self.after_tax_cash_flow = after_tax_cash_flow
        self.net_annual_cost_without_sys = net_annual_costs_without_system
        self.net_annual_cost_with_sys = net_annual_costs_with_system

        # compute outputs
        try:
            self.irr = np.irr(after_tax_cash_flow)
        except ValueError:
            self.irr = 0
            log("ERROR", "IRR calculation failed to compute a real number")

        if math.isnan(self.irr):
            self.irr = 0

        self.npv = np.npv(self.econ.owner_discount_rate_nominal, after_tax_cash_flow)
        self.lcc = -np.npv(self.econ.owner_discount_rate_nominal, net_annual_costs_with_system)
        self.lcc_bau = -np.npv(self.econ.owner_discount_rate_nominal, net_annual_costs_without_system)

    def pbi_calculate(self, tech, year, annual_energy):

        pbi_amount = 0
        if year <= self.incentives[tech]['pbi_combined_term']:
            pbi_amount = min(self.incentives[tech]['pbi_combined_amount'] * annual_energy, self.incentives[tech]['pbi_combined_maxvalue'])
        return pbi_amount

    def federal_taxable_income_before_deductions(self, tech):

        taxable_income = 0
        tech_cost = self.incentives[tech]['tech_cost']
        tech_kw = self.incentives[tech]['tech_size']

        state_ibi = min(self.incentives[tech]['ibi_sta_percent'] * tech_cost,
                        self.incentives[tech]['ibi_sta_percent_maxvalue'])
        utility_ibi = min(self.incentives[tech]['ibi_uti_percent'] * tech_cost,
                          self.incentives[tech]['ibi_uti_percent_maxvalue'])
        federal_cbi = min(self.incentives[tech]['cbi_fed_amount'] * tech_kw,
                          self.incentives[tech]['cbi_fed_maxvalue'])
        state_cbi = min(self.incentives[tech]['cbi_sta_amount'] * tech_kw,
                        self.incentives[tech]['cbi_sta_maxvalue'])
        utility_cbi = min(self.incentives[tech]['cbi_uti_amount'] * tech_kw,
                          self.incentives[tech]['cbi_uti_maxvalue'])

        if self.incentives[tech]['ibi_sta_percent_tax_fed']:
            taxable_income += state_ibi
        if self.incentives[tech]['ibi_uti_percent_tax_fed']:
            taxable_income += utility_ibi
        if self.incentives[tech]['cbi_fed_tax_fed']:
            taxable_income += federal_cbi
        if self.incentives[tech]['cbi_sta_tax_fed']:
            taxable_income += state_cbi
        if self.incentives[tech]['cbi_uti_tax_fed']:
            taxable_income += utility_cbi

        return taxable_income

    def federal_depreciation_basis(self, tech):

        tech_cost = self.incentives[tech]['tech_cost']

        federal_itc = min(self.incentives[tech]['itc_fed_percent'] * tech_cost,
                          self.incentives[tech]['itc_fed_percent_maxvalue'])
        itc_federal = 0

        if self.incentives[tech]['itc_fed_percent_deprbas_fed']:
            itc_federal = self.incentives[tech]['macrs_itc_reduction'] * federal_itc

        federal_itc_basis = self.federal_itc_basis(tech)
        basis = federal_itc_basis - itc_federal

        return basis

    def federal_depreciation_amount_total(self, year):

        depreciation_amount = 0
        for tech in self.techs:
            bonus_fraction = 0
            macrs_schedule = self.incentives[tech]['macrs_schedule']

            macrs_schedule_array = []
            if macrs_schedule == 5:
                macrs_schedule_array = self.econ.macrs_five_year
            elif macrs_schedule == 7:
                macrs_schedule_array = self.econ.macrs_seven_year

            depreciation_basis = self.incentives[tech]['federal_depreciation_basis']
            if year == 1:
                bonus_fraction = self.incentives[tech]['bonus_fraction']
            if year <= len(macrs_schedule_array) and macrs_schedule > 0:
                depreciation_amount += depreciation_basis * (macrs_schedule_array[year - 1] + bonus_fraction)

        return depreciation_amount

    def federal_itc_basis(self, tech):

        # reduce the itc basis
        ibi_state = 0
        ibi_util = 0
        cbi_fed = 0
        cbi_state = 0
        cbi_util = 0

        tech_cost = self.incentives[tech]['tech_cost']
        tech_kw = self.incentives[tech]['tech_size']

        state_ibi = min(self.incentives[tech]['ibi_sta_percent'] * tech_cost,
                        self.incentives[tech]['ibi_sta_percent_maxvalue'])
        utility_ibi = min(self.incentives[tech]['ibi_uti_percent'] * tech_cost,
                          self.incentives[tech]['ibi_uti_percent_maxvalue'])
        federal_cbi = min(self.incentives[tech]['cbi_fed_amount'] * tech_kw,
                          self.incentives[tech]['cbi_fed_maxvalue'])
        state_cbi = min(self.incentives[tech]['cbi_sta_amount'] * tech_kw,
                        self.incentives[tech]['cbi_sta_maxvalue'])
        utility_cbi = min(self.incentives[tech]['cbi_uti_amount'] * tech_kw,
                          self.incentives[tech]['cbi_uti_maxvalue'])

        if self.incentives[tech]['ibi_sta_percent_deprbas_fed']:
            ibi_state = state_ibi
        if self.incentives[tech]['ibi_uti_percent_deprbas_fed']:
            ibi_util = utility_ibi
        if self.incentives[tech]['cbi_fed_deprbas_fed']:
            cbi_fed = federal_cbi
        if self.incentives[tech]['cbi_sta_deprbas_fed']:
            cbi_state = state_cbi
        if self.incentives[tech]['cbi_uti_deprbas_fed']:
            cbi_util = utility_cbi

        return tech_cost - ibi_state - ibi_util - cbi_fed - cbi_state - cbi_util

    def federal_itc_total(self):

        federal_itc = 0
        for tech in self.techs:
            tech_cost = self.incentives[tech]['tech_cost']
            federal_itc += min(self.incentives[tech]['itc_fed_percent'] * tech_cost,
                               self.incentives[tech]['itc_fed_percent_maxvalue'])
        return federal_itc

    def total_cash_incentives(self, tech):

        tech_cost = self.incentives[tech]['tech_cost']
        tech_kw = self.incentives[tech]['tech_size']

        state_ibi = float(min(self.incentives[tech]['ibi_sta_percent'] * tech_cost,
                        self.incentives[tech]['ibi_sta_percent_maxvalue']))
        utility_ibi = float(min(self.incentives[tech]['ibi_uti_percent'] * tech_cost,
                          self.incentives[tech]['ibi_uti_percent_maxvalue']))
        federal_cbi = float(min(self.incentives[tech]['cbi_fed_amount'] * tech_kw,
                          self.incentives[tech]['cbi_fed_maxvalue']))
        state_cbi = float(min(self.incentives[tech]['cbi_sta_amount'] * tech_kw,
                        self.incentives[tech]['cbi_sta_maxvalue']))
        utility_cbi = float(min(self.incentives[tech]['cbi_uti_amount'] * tech_kw,
                          self.incentives[tech]['cbi_uti_maxvalue']))

        return state_ibi + utility_ibi + federal_cbi + state_cbi + utility_cbi





